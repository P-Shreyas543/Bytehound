"""100Hz Multi-Cell BMS Integration Test with physical S32K144 MCU.

Loads the MultiCell-BMS-OverAllFrame.xlsx configuration, connects to the board,
exercises active polling and 100Hz streaming, logs data to raw & decoded Excel workbooks,
runs a timeline of commands, and validates spreadsheet consistency and timings.
"""

from __future__ import annotations

import argparse
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

sys.path.insert(0, str(Path(__file__).parent))

from PySide6.QtCore import QCoreApplication
import openpyxl
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from app.commands.tx_command_builder import build_tx_command
from app.decoder.config_loader import load_config
from app.decoder.frame_decoder import decode_frame
from app.protocol.packet_builder import build_packet
from app.serial_io.serial_worker import PollingWorker, SerialSettings
from app.serial_logging.decoded_logger import DecodedLogger
from app.serial_logging.raw_logger import RawLogger


class Report:
    def __init__(self) -> None:
        self.passed: List[str] = []
        self.failed: List[str] = []

    def ok(self, label: str, detail: str = "") -> None:
        print(f"[ OK ] {label}" + (f"  ({detail})" if detail else ""))
        self.passed.append(label)

    def fail(self, label: str, detail: str = "") -> None:
        print(f"[FAIL] {label}" + (f"  ({detail})" if detail else ""))
        self.failed.append(f"{label} - {detail}" if detail else label)

    def note(self, msg: str) -> None:
        print(f"[note] {msg}")


def run_for(app: QCoreApplication, seconds: float) -> None:
    """Spin the Qt event loop for `seconds` and return."""
    end = time.perf_counter() + seconds
    while time.perf_counter() < end:
        app.processEvents()
        time.sleep(0.005)


def generate_plot(decoded_path: Path, plot_output_path: Path) -> None:
    """Reads the generated decoded Excel sheet and plots the key telemetry parameters."""
    print(f"Reading decoded log from {decoded_path} for plotting...")
    
    df = pd.read_excel(decoded_path, sheet_name="Data")
    
    time_col = None
    for col in df.columns:
        if "elapsed_ms" in col:
            time_col = col
            break
            
    if time_col is None:
        print("Could not find elapsed_ms column. Plotting against index.")
        df["Time (s)"] = df.index
    else:
        df["Time (s)"] = df[time_col] / 1000.0

    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["axes.edgecolor"] = "#CBD5E1"
    plt.rcParams["axes.linewidth"] = 1.2
    plt.rcParams["grid.color"] = "#F1F5F9"
    plt.rcParams["grid.linewidth"] = 0.8

    fig, axs = plt.subplots(4, 1, figsize=(12, 16), sharex=True)
    fig.suptitle("NXP S32K144 100Hz BMS Integration Telemetry Log", fontsize=16, fontweight="bold", color="#1E293B", y=0.96)

    def find_col(name_part):
        for col in df.columns:
            if name_part.lower() in col.lower():
                return col
        return None

    # Plot 1: Pack Voltage & Load Voltage
    v_pack_col = find_col("Pack Voltage")
    v_load_col = find_col("Load Voltage")
    ax1 = axs[0]
    if v_pack_col:
        ax1.plot(df["Time (s)"], df[v_pack_col], label="Pack Voltage", color="#2563EB", linewidth=2)
    if v_load_col:
        ax1.plot(df["Time (s)"], df[v_load_col], label="Load Voltage", color="#059669", linewidth=1.5, linestyle="--")
    ax1.set_ylabel("Voltage (V)", fontsize=11, fontweight="semibold", color="#475569")
    ax1.set_title("BMS Pack and Load Voltages", fontsize=12, fontweight="semibold", color="#1E293B", loc="left")
    ax1.legend(loc="upper right")
    ax1.grid(True)

    # Plot 2: Cell Voltages (14 Cells)
    ax2 = axs[1]
    cell_cols = [c for c in df.columns if "Cell Voltage " in c and not any(x in c for x in ["min", "max", "diff", "avg"])]
    if cell_cols:
        colors = plt.cm.viridis(plt.cm.colors.Normalize(0, len(cell_cols))(range(len(cell_cols))))
        for idx, col in enumerate(sorted(cell_cols)):
            ax2.plot(df["Time (s)"], df[col], label=col.split(".")[-1], color=colors[idx], linewidth=1, alpha=0.8)
    ax2.set_ylabel("Cell Voltage (V)", fontsize=11, fontweight="semibold", color="#475569")
    ax2.set_title("Individual Cell Voltages (14 Channels)", fontsize=12, fontweight="semibold", color="#1E293B", loc="left")
    ax2.grid(True)

    # Plot 3: Temperatures
    ax3 = axs[2]
    pack_temp_cols = [c for c in df.columns if "Pack Temperature " in c and not any(x in c for x in ["min", "max", "avg"])]
    mosfet_temp_col = find_col("MOSFET Temperature")
    ambient_temp_col = find_col("Ambient Temperature")
    ic_temp_col = find_col("IC Temperature")

    if pack_temp_cols:
        ax3.plot(df["Time (s)"], df[pack_temp_cols[0]], label="Pack Temp 1", color="#EF4444", linewidth=1.5)
    if mosfet_temp_col:
        ax3.plot(df["Time (s)"], df[mosfet_temp_col], label="MOSFET", color="#F59E0B", linewidth=1.5)
    if ambient_temp_col:
        ax3.plot(df["Time (s)"], df[ambient_temp_col], label="Ambient", color="#10B981", linewidth=1.5)
    if ic_temp_col:
        ax3.plot(df["Time (s)"], df[ic_temp_col], label="BCC IC", color="#8B5CF6", linewidth=1.5)
    ax3.set_ylabel("Temperature (°C)", fontsize=11, fontweight="semibold", color="#475569")
    ax3.set_title("System Temperature Sensors", fontsize=12, fontweight="semibold", color="#1E293B", loc="left")
    ax3.legend(loc="upper right")
    ax3.grid(True)

    # Plot 4: Capacity and State of Charge (SOC)
    ax4 = axs[3]
    soc_col = find_col("SOC (CC)")
    capacity_col = find_col("Cumulative Capacity")
    if soc_col:
        ax4.plot(df["Time (s)"], df[soc_col], label="SOC (%)", color="#3B82F6", linewidth=2)
        ax4.set_ylabel("SOC (%)", fontsize=11, fontweight="semibold", color="#3B82F6")
        ax4.tick_params(axis="y", labelcolor="#3B82F6")
    if capacity_col:
        ax4_twin = ax4.twinx()
        ax4_twin.plot(df["Time (s)"], df[capacity_col], label="Capacity (Ah)", color="#EC4899", linewidth=1.5, linestyle=":")
        ax4_twin.set_ylabel("Capacity (Ah)", fontsize=11, fontweight="semibold", color="#EC4899")
        ax4_twin.tick_params(axis="y", labelcolor="#EC4899")
        
    ax4.set_xlabel("Elapsed Time (seconds)", fontsize=11, fontweight="semibold", color="#475569")
    ax4.set_title("BMS State of Charge and Capacity", fontsize=12, fontweight="semibold", color="#1E293B", loc="left")
    ax4.grid(True)

    plt.tight_layout()
    fig.subplots_adjust(top=0.92, bottom=0.06)
    
    plt.savefig(plot_output_path, dpi=150, facecolor="white")
    plt.close()
    print(f"Plot successfully saved to {plot_output_path}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", default="COM9")
    parser.add_argument("--seconds", type=int, default=60,
                        help="duration of the test in seconds (default 60s)")
    parser.add_argument("--boot-delay", type=float, default=3.0,
                        help="delay after opening serial port before starting operations")
    args = parser.parse_args()

    rep = Report()
    
    # Load MultiCell overall frame configuration
    print("Loading config from MultiCell-BMS-OverAllFrame.xlsx...")
    cfg = load_config(Path("MultiCell-BMS-OverAllFrame.xlsx"))

    # Output paths
    Path("scratch").mkdir(exist_ok=True)
    raw_path = Path("scratch/smoke_overall_raw.xlsx")
    dec_path = Path("scratch/smoke_overall_decoded.xlsx")
    plot_path = Path("scratch/smoke_overall_plot.png")

    raw_path.unlink(missing_ok=True)
    dec_path.unlink(missing_ok=True)
    plot_path.unlink(missing_ok=True)

    # Initialize PySide application
    app = QCoreApplication.instance() or QCoreApplication(sys.argv)

    # Initialize loggers
    raw_logger = RawLogger(raw_path, hex_format=cfg.protocol.raw_log_format)
    decoded_logger = DecodedLogger(dec_path, cfg)
    
    # Force polling mode
    decoded_logger.polling_mode = True

    # Counters
    rx_packet_count = 0
    tx_packet_count = 0
    crc_errors = 0
    tx_recorded_data = []

    # Track 0x1000 arrival times to calculate packet interval
    t1000_arrivals = []

    def on_packets(batch):
        nonlocal rx_packet_count, crc_errors
        for item in batch:
            pkt, pre_decoded = item if isinstance(item, tuple) else (item, None)
            rx_packet_count += 1
            raw_logger.log("RX", pkt.raw)
            if not pkt.ok:
                crc_errors += 1
                print(f"[warning] Received corrupt packet: {pkt.error}")
                continue
            
            if pkt.frame_id == 0x1000:
                t1000_arrivals.append(time.perf_counter())

            decoded = pre_decoded if pre_decoded is not None else decode_frame(cfg, pkt.frame_id, pkt.payload)
            elapsed_ms = int((time.perf_counter() - log_start) * 1000)
            decoded_logger.log_frame(decoded, elapsed_ms)

    def on_tx(data: bytes):
        nonlocal tx_packet_count
        tx_packet_count += 1
        tx_recorded_data.append(data)
        raw_logger.log("TX", data)
        
        # Decode and log TX commands in the decoded log
        try:
            if len(data) >= 5 and data[0] == 0xAA and data[1] == 0x55:
                frame_id = int.from_bytes(data[2:4], "little")
                payload_len = data[4]
                payload = data[5:5+payload_len]
                decoded_tx = decode_frame(cfg, frame_id, payload)
                elapsed_ms = int((time.perf_counter() - log_start) * 1000)
                decoded_logger.log_frame(decoded_tx, elapsed_ms)
        except Exception as e:
            rep.note(f"Failed to decode TX packet: {e}")

    def on_error(msg: str):
        rep.note(f"Worker error: {msg}")

    # Set up serial worker
    settings = SerialSettings(port=args.port, baud_rate=cfg.serial_defaults.baud_rate)
    worker = PollingWorker(settings, cfg.protocol, cfg.polling_schedules)
    
    # Enable pipelining for faster high-frequency polling
    worker.set_pipelining(True, depth=2, gap_ms=10) # 10 ms TX gap

    worker.packets_received.connect(on_packets)
    worker.tx_recorded.connect(on_tx)
    worker.error_occurred.connect(on_error)

    print(f"\n=================== STARTING 100Hz MULTICELL TEST ON {args.port} ===================")
    print(f"Target duration: {args.seconds} seconds")
    print(f"Writing raw log to: {raw_path}")
    print(f"Writing decoded log to: {dec_path}")

    # Open loggers and worker
    raw_logger.open()
    decoded_logger.open()
    worker.open()
    worker.set_polling_global(False) # start passive
    
    log_start = time.perf_counter()

    # Settle delay
    print(f"Waiting {args.boot_delay} seconds for board settle...")
    run_for(app, args.boot_delay)

    # Enable active polling / listen to streams
    worker.set_polling_global(True)
    print("Active polling enabled.")

    # Send Fault Reset command (AA 55 00 10 02 FF FF + CRC big endian + padded)
    try:
        pkt = build_tx_command(cfg, "Fault Reset", {})
        worker.enqueue_priority_tx(pkt)
        print("Fault Reset command queued.")
    except Exception as e:
        rep.note(f"Failed to build Fault Reset command: {e}")

    # Timeline of actions
    command_schedule = {
        10: ("Stress ON", True),
        25: ("Fault Reset", None),
        40: ("Stress OFF", False),
        50: ("Fault Reset", None)
    }
    commands_sent = set()

    test_end_time = log_start + args.seconds
    next_status_print = log_start + 5.0

    try:
        while time.perf_counter() < test_end_time:
            run_for(app, 0.05)
            now = time.perf_counter()
            elapsed = now - log_start

            # Print status periodically
            if now >= next_status_print:
                print(f"[STATUS] {elapsed:.1f}s elapsed | "
                      f"RX Packets: {rx_packet_count} | "
                      f"TX Packets: {tx_packet_count} | "
                      f"CRC Errors: {crc_errors}")
                next_status_print = now + 10.0

            # Check command schedule
            for sched_sec, cmd_info in list(command_schedule.items()):
                if elapsed >= sched_sec and sched_sec not in commands_sent:
                    commands_sent.add(sched_sec)
                    cmd_type, cmd_val = cmd_info
                    print(f"\n[ACTION] Elapsed {elapsed:.1f}s: Triggering {cmd_type}")
                    
                    try:
                        if cmd_type == "Fault Reset":
                            pkt = build_tx_command(cfg, "Fault Reset", {})
                            worker.enqueue_priority_tx(pkt)
                        elif cmd_type == "Stress ON":
                            pkt = build_packet(cfg.protocol, 0x1002, b"\x01")
                            worker.enqueue_priority_tx(pkt)
                        elif cmd_type == "Stress OFF":
                            pkt = build_packet(cfg.protocol, 0x1002, b"\x00")
                            worker.enqueue_priority_tx(pkt)
                    except Exception as err:
                        rep.fail(f"Sending command scheduled at {sched_sec}s", str(err))

    except KeyboardInterrupt:
        print("\nTest interrupted by user.")
    finally:
        print("\nStopping serial worker and loggers...")
        worker.close()
        raw_logger.close()
        decoded_logger.close()
        
    print("Logs closed. Awaiting thread drain...")
    raw_logger.await_drain(timeout=30.0)
    decoded_logger.await_drain(timeout=30.0)

    # ========================== VERIFICATION PHASE ==========================
    print("\n=================== VERIFICATION PHASE ===================")
    
    # 1. Raw log verification
    if raw_path.exists() and raw_path.stat().st_size > 0:
        rep.ok("Raw log file created", f"{raw_path.name} ({raw_path.stat().st_size} bytes)")
        try:
            wb_raw = openpyxl.load_workbook(raw_path, read_only=True)
            if "Metadata" in wb_raw.sheetnames and "Data" in wb_raw.sheetnames:
                rep.ok("Raw log contains Metadata and Data sheets")
                data_ws = wb_raw["Data"]
                headers = [cell.value for cell in next(data_ws.iter_rows(max_row=1))]
                expected_cols = ["timestamp", "direction", "hex", "delta_t_ms"]
                if headers == expected_cols:
                    rep.ok("Raw log Data sheet headers match perfectly")
                else:
                    rep.fail("Raw log headers mismatch", f"got {headers} expected {expected_cols}")
            wb_raw.close()
        except Exception as e:
            rep.fail("Loading raw log Excel file", str(e))
    else:
        rep.fail("Raw log file missing or empty")

    # 2. Decoded log verification
    if dec_path.exists() and dec_path.stat().st_size > 0:
        rep.ok("Decoded log file created", f"{dec_path.name} ({dec_path.stat().st_size} bytes)")
        try:
            wb_dec = openpyxl.load_workbook(dec_path, read_only=True)
            if "Metadata" in wb_dec.sheetnames and "Data" in wb_dec.sheetnames:
                rep.ok("Decoded log contains Metadata and Data sheets")
                data_ws = wb_dec["Data"]
                headers = [cell.value for cell in next(data_ws.iter_rows(max_row=1))]
                
                # Check for expanded cell voltages (14 channels) and temps (7 channels)
                has_cell1 = any("Cell Voltage 1" in h for h in headers if h)
                has_cell14 = any("Cell Voltage 14" in h for h in headers if h)
                has_temp7 = any("Pack Temperature 7" in h for h in headers if h)
                
                if has_cell1 and has_cell14 and has_temp7:
                    rep.ok("Decoded log headers successfully mapped expanded signals")
                else:
                    rep.fail("Decoded log headers missing expanded signals", 
                             f"cell1={has_cell1}, cell14={has_cell14}, temp7={has_temp7}")
            wb_dec.close()
        except Exception as e:
            rep.fail("Loading decoded log Excel file", str(e))
    else:
        rep.fail("Decoded log file missing or empty")

    # 3. 100Hz Timing interval verification
    if len(t1000_arrivals) > 10:
        intervals = [t1000_arrivals[i] - t1000_arrivals[i-1] for i in range(1, len(t1000_arrivals))]
        avg_interval_ms = (sum(intervals) / len(intervals)) * 1000
        rep.ok("100Hz packet stream verified", f"Average interval for 0x1000: {avg_interval_ms:.2f} ms")
        if avg_interval_ms < 15.0:
            rep.ok("Streaming interval meets the 100Hz target (<15ms average)")
        else:
            rep.fail("Streaming interval is too slow", f"Average interval: {avg_interval_ms:.2f} ms")
    else:
        rep.fail("Insufficient 0x1000 packets received to verify 100Hz timing", f"count={len(t1000_arrivals)}")

    # 4. Generate matplotlib plot
    try:
        generate_plot(dec_path, plot_path)
        if plot_path.exists() and plot_path.stat().st_size > 0:
            rep.ok("Telemetry plot generated successfully", f"{plot_path.name}")
        else:
            rep.fail("Plot file size is 0 bytes")
    except Exception as e:
        rep.fail("Generating matplotlib plot", str(e))

    # ========================== FINAL REPORT ==========================
    print(f"\n=================== TEST RESULT SUMMARY ===================")
    print(f"PASSED: {len(rep.passed)}")
    print(f"FAILED: {len(rep.failed)}")
    if rep.failed:
        for f in rep.failed:
            print(f"  FAIL: {f}")
        return len(rep.failed)
    
    print("ALL TESTS PASSED SUCCESSFULLY!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
