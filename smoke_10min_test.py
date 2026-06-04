"""10-Minute Comprehensive Integration Test with physical NXP S32K144.

Exercises the serial pipeline, polling, TX commands, and logging systems over 
a realistic 10-minute timeframe. Generates Excel spreadsheet logs and plots 
the decoded data using matplotlib.

Usage:
    .venv\\Scripts\\python.exe smoke_10min_test.py --port COM9
"""

from __future__ import annotations

import argparse
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

sys.path.insert(0, str(Path(__file__).parent))

from PySide6.QtCore import QCoreApplication
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt

from app.commands.tx_command_builder import build_tx_command
from app.decoder.config_loader import load_config
from app.decoder.frame_decoder import decode_frame
from app.protocol.packet_builder import build_packet
from app.serial_io.serial_worker import PollingWorker, SerialSettings
from app.serial_logging.decoded_logger import DecodedLogger
from app.serial_logging.raw_logger import RawLogger


class Report:
    def __init__(self) -> None:
        self.passed: List[str] = []
        self.failed: List[str] = []

    def ok(self, label: str, detail: str = "") -> None:
        print(f"[ OK ] {label}" + (f"  ({detail})" if detail else ""))
        self.passed.append(label)

    def fail(self, label: str, detail: str = "") -> None:
        print(f"[FAIL] {label}" + (f"  ({detail})" if detail else ""))
        self.failed.append(f"{label} - {detail}" if detail else label)

    def note(self, msg: str) -> None:
        print(f"[note] {msg}")


def run_for(app: QCoreApplication, seconds: float) -> None:
    """Spin the Qt event loop for `seconds` and return."""
    end = time.perf_counter() + seconds
    while time.perf_counter() < end:
        app.processEvents()
        time.sleep(0.005)


def generate_plot(decoded_path: Path, plot_output_path: Path) -> None:
    """Reads the generated decoded Excel sheet and plots the key telemetry parameters."""
    print(f"Reading decoded log from {decoded_path}...")
    
    # Read the Data sheet from Excel
    df = pd.read_excel(decoded_path, sheet_name="Data")
    
    # Determine the time column (we look for elapsed_ms or elapsed time)
    # The columns are prefixed by frame name, e.g., 'BMS Status.elapsed_ms'
    time_col = None
    for col in df.columns:
        if "elapsed_ms" in col:
            time_col = col
            break
            
    if time_col is None:
        print("Could not find elapsed_ms column in the decoded log sheet. Plotting against index.")
        df["Time (s)"] = df.index
    else:
        df["Time (s)"] = df[time_col] / 1000.0

    print("Columns available:", list(df.columns))

    # Set up styling for a modern, high-quality look
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["font.sans-serif"] = ["DejaVu Sans", "Arial", "Liberation Sans"]
    plt.rcParams["axes.edgecolor"] = "#CBD5E1"
    plt.rcParams["axes.linewidth"] = 1.2
    plt.rcParams["grid.color"] = "#F1F5F9"
    plt.rcParams["grid.linewidth"] = 0.8

    fig, axs = plt.subplots(4, 1, figsize=(12, 16), sharex=True)
    fig.suptitle("NXP S32K144 BMS Simulator 10-Minute Telemetry log", fontsize=16, fontweight="bold", color="#1E293B", y=0.96)

    # Helper function to find matching columns
    def find_col(name_part):
        for col in df.columns:
            if name_part.lower() in col.lower():
                return col
        return None

    # Plot 1: Pack Voltage & Voltage Limit
    v_pack_col = find_col("Pack Voltage")
    v_lim_col = find_col("Voltage Limit")
    ax1 = axs[0]
    if v_pack_col:
        ax1.plot(df["Time (s)"], df[v_pack_col], label="Pack Voltage", color="#3B82F6", linewidth=2)
    if v_lim_col:
        ax1.plot(df["Time (s)"], df[v_lim_col], label="Voltage Limit", color="#EF4444", linewidth=2, linestyle="--")
    ax1.set_ylabel("Voltage (V)", fontsize=11, fontweight="semibold", color="#475569")
    ax1.set_title("Pack Voltage & Configured Voltage Limit", fontsize=12, fontweight="semibold", color="#1E293B", loc="left")
    ax1.legend(loc="upper right", frameon=True, facecolor="white", edgecolor="#E2E8F0")
    ax1.grid(True)

    # Plot 2: Cell Voltages
    ax2 = axs[1]
    cell_cols = [c for c in df.columns if "Cell Voltage" in c]
    if cell_cols:
        colors = plt.cm.plasma(plt.cm.colors.Normalize(0, len(cell_cols))(range(len(cell_cols))))
        for idx, col in enumerate(sorted(cell_cols)):
            ax2.plot(df["Time (s)"], df[col], label=col.split(".")[-1], color=colors[idx], linewidth=1, alpha=0.8)
    ax2.set_ylabel("Cell Voltage (V)", fontsize=11, fontweight="semibold", color="#475569")
    ax2.set_title("Individual Cell Voltages (8 Channels)", fontsize=12, fontweight="semibold", color="#1E293B", loc="left")
    # Show legend only for first and last to keep it tidy
    if len(cell_cols) > 0:
        ax2.legend(loc="upper right", frameon=True, facecolor="white", edgecolor="#E2E8F0", ncol=4)
    ax2.grid(True)

    # Plot 3: Pack Current & SOC
    ax3 = axs[2]
    current_col = find_col("Pack Current")
    soc_col = find_col("Pack SOC")
    
    if current_col:
        ax3.plot(df["Time (s)"], df[current_col], label="Current (A)", color="#10B981", linewidth=2)
        ax3.set_ylabel("Current (A)", fontsize=11, fontweight="semibold", color="#10B981")
        ax3.tick_params(axis="y", labelcolor="#10B981")
        
    if soc_col:
        ax3_twin = ax3.twinx()
        ax3_twin.plot(df["Time (s)"], df[soc_col], label="SOC (%)", color="#F59E0B", linewidth=1.5, linestyle=":")
        ax3_twin.set_ylabel("State of Charge (%)", fontsize=11, fontweight="semibold", color="#F59E0B")
        ax3_twin.tick_params(axis="y", labelcolor="#F59E0B")
        ax3_twin.grid(False)
        
    ax3.set_title("Current & State of Charge (SOC)", fontsize=12, fontweight="semibold", color="#1E293B", loc="left")
    ax3.grid(True)

    # Plot 4: BMS State & Mode
    ax4 = axs[3]
    state_col = find_col("BMS State.label")
    mode_col = find_col("mode") # or Mode.label
    
    if state_col:
        # Map labels to numbers for plotting
        unique_labels = sorted(df[state_col].dropna().unique())
        label_map = {l: i for i, l in enumerate(unique_labels)}
        mapped_states = df[state_col].map(label_map)
        ax4.step(df["Time (s)"], mapped_states, where="post", label="BMS State", color="#8B5CF6", linewidth=2)
        ax4.set_yticks(range(len(unique_labels)))
        ax4.set_yticklabels(unique_labels)
        ax4.set_ylabel("BMS State", fontsize=11, fontweight="semibold", color="#8B5CF6")
        
    ax4.set_xlabel("Elapsed Time (seconds)", fontsize=11, fontweight="semibold", color="#475569")
    ax4.set_title("BMS Operating State Machine Transitions", fontsize=12, fontweight="semibold", color="#1E293B", loc="left")
    ax4.grid(True)

    plt.tight_layout()
    fig.subplots_adjust(top=0.92, bottom=0.06)
    
    plt.savefig(plot_output_path, dpi=150, facecolor="white")
    plt.close()
    print(f"Plot saved to {plot_output_path}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", default="COM9")
    parser.add_argument("--seconds", type=int, default=600,
                        help="duration of the test in seconds (default 600s / 10mins)")
    parser.add_argument("--boot-delay", type=float, default=3.0,
                        help="delay after opening serial port before starting operations")
    args = parser.parse_args()

    rep = Report()
    cfg = load_config(Path("app/resources/config_template"))

    # Output paths
    Path("scratch").mkdir(exist_ok=True)
    raw_path = Path("scratch/smoke_10min_raw.xlsx")
    dec_path = Path("scratch/smoke_10min_decoded.xlsx")
    plot_path = Path("scratch/smoke_10min_plot.png")

    raw_path.unlink(missing_ok=True)
    dec_path.unlink(missing_ok=True)
    plot_path.unlink(missing_ok=True)

    # Initialize PySide application
    app = QCoreApplication.instance() or QCoreApplication(sys.argv)

    # Initialize loggers
    raw_logger = RawLogger(raw_path, hex_format=cfg.protocol.raw_log_format)
    decoded_logger = DecodedLogger(dec_path, cfg)
    
    # We set polling_mode to True since we will poll the frames
    decoded_logger.polling_mode = True

    # Collectors
    rx_packet_count = 0
    tx_packet_count = 0
    crc_errors = 0
    tx_recorded_data = []

    def on_packets(batch):
        nonlocal rx_packet_count, crc_errors
        for item in batch:
            pkt, pre_decoded = item if isinstance(item, tuple) else (item, None)
            rx_packet_count += 1
            raw_logger.log("RX", pkt.raw)
            if not pkt.ok:
                crc_errors += 1
                continue
            
            decoded = pre_decoded if pre_decoded is not None else decode_frame(cfg, pkt.frame_id, pkt.payload)
            elapsed_ms = int((time.perf_counter() - log_start) * 1000)
            decoded_logger.log_frame(decoded, elapsed_ms)

    def on_tx(data: bytes):
        nonlocal tx_packet_count
        tx_packet_count += 1
        tx_recorded_data.append(data)
        raw_logger.log("TX", data)
        
        # Also decode and log TX commands in the decoded log
        try:
            # Reconstruct frame_id and payload
            # header 0xAA 0x55 (2 bytes) + frame_id LE (2 bytes) + len (1 byte)
            if len(data) >= 8 and data[0] == 0xAA and data[1] == 0x55:
                frame_id = int.from_bytes(data[2:4], "little")
                payload_len = data[4]
                payload = data[5:5+payload_len]
                decoded_tx = decode_frame(cfg, frame_id, payload)
                elapsed_ms = int((time.perf_counter() - log_start) * 1000)
                decoded_logger.log_frame(decoded_tx, elapsed_ms)
        except Exception as e:
            rep.note(f"Failed to decode TX packet: {e}")

    def on_error(msg: str):
        rep.note(f"Worker Error occurred: {msg}")

    # Set up serial worker
    settings = SerialSettings(port=args.port, baud_rate=115200)
    worker = PollingWorker(settings, cfg.protocol, cfg.polling_schedules)
    
    # Enable pipelining for faster and reliable communication
    worker.set_pipelining(True, depth=2, gap_ms=30)

    worker.packets_received.connect(on_packets)
    worker.tx_recorded.connect(on_tx)
    worker.error_occurred.connect(on_error)

    print(f"\n=================== STARTING 10-MINUTE TEST ON {args.port} ===================")
    print(f"Target duration: {args.seconds} seconds ({args.seconds/60:.1f} minutes)")
    print(f"Writing raw log to: {raw_path}")
    print(f"Writing decoded log to: {dec_path}")

    # Open loggers and worker
    raw_logger.open()
    decoded_logger.open()
    worker.open()
    worker.set_polling_global(False) # passive initially
    
    log_start = time.perf_counter()

    # Settle delay
    print(f"Waiting {args.boot_delay} seconds for board bootloader settle...")
    run_for(app, args.boot_delay)

    # Enable active polling
    worker.set_polling_global(True)
    print("Active polling enabled.")

    # Autonomous streaming OFF command
    # Turns off streaming so we only receive response to poll requests
    try:
        worker.enqueue_priority_tx(build_packet(cfg.protocol, 0x1005, b"\x01"))
        print("Autonomous streaming disabled on board (polling-only active).")
    except Exception as e:
        rep.note(f"Failed to send streaming OFF command: {e}")

    # Reset initial state
    try:
        worker.enqueue_priority_tx(build_tx_command(cfg, "Reset Faults", {}))
        print("Initial Reset Faults command sent.")
    except Exception as e:
        rep.note(f"Failed to send initial reset faults command: {e}")

    # Initial voltage limit set
    current_limit = 55.0
    try:
        worker.enqueue_priority_tx(build_tx_command(cfg, "Set Voltage Limit", {"Voltage Limit (V)": current_limit}))
        print(f"Initial Voltage Limit set to {current_limit} V.")
    except Exception as e:
        rep.note(f"Failed to send initial voltage limit: {e}")

    # Main run loop
    next_status_print = log_start + 10.0
    test_end_time = log_start + args.seconds

    # Commands timeline
    # {time_offset_seconds: (type, value)}
    command_schedule = {
        60:  ("Set Voltage Limit", 52.0),
        120: ("Set Voltage Limit", 58.0),
        180: ("Reset Faults", None),
        240: ("Set Voltage Limit", 45.0),
        300: ("Inject Stress Mode", True),
        360: ("Inject Stress Mode", False),
        420: ("Set Voltage Limit", 58.5),
        480: ("Reset Faults", None),
        540: ("Set Voltage Limit", 50.0)
    }
    
    commands_sent = set()

    try:
        while time.perf_counter() < test_end_time:
            run_for(app, 0.1)
            now = time.perf_counter()
            elapsed = now - log_start

            # Print status periodically
            if now >= next_status_print:
                print(f"[STATUS] {elapsed/60:.1f} mins elapsed | "
                      f"RX Packets: {rx_packet_count} | "
                      f"TX Packets: {tx_packet_count} | "
                      f"CRC Errors: {crc_errors}")
                next_status_print = now + 30.0 # print every 30 seconds

            # Check command schedule
            for sched_sec, cmd_info in list(command_schedule.items()):
                if elapsed >= sched_sec and sched_sec not in commands_sent:
                    commands_sent.add(sched_sec)
                    cmd_type, cmd_val = cmd_info
                    print(f"\n[ACTION] Elapsed {elapsed:.1f}s: Triggering {cmd_type} with value {cmd_val}")
                    
                    try:
                        if cmd_type == "Set Voltage Limit":
                            pkt = build_tx_command(cfg, "Set Voltage Limit", {"Voltage Limit (V)": cmd_val})
                            worker.enqueue_priority_tx(pkt)
                            current_limit = cmd_val
                        elif cmd_type == "Reset Faults":
                            pkt = build_tx_command(cfg, "Reset Faults", {})
                            worker.enqueue_priority_tx(pkt)
                        elif cmd_type == "Inject Stress Mode":
                            # 0x1002 is stress mode frame
                            payload = b"\x01" if cmd_val else b"\x00"
                            pkt = build_packet(cfg.protocol, 0x1002, payload)
                            worker.enqueue_priority_tx(pkt)
                    except Exception as err:
                        rep.fail(f"Sending command scheduled at {sched_sec}s", str(err))

    except KeyboardInterrupt:
        print("\nTest interrupted by user. Stopping loggers and saving data...")
    finally:
        print("\nStopping serial worker and loggers...")
        worker.close()
        raw_logger.close()
        decoded_logger.close()
        
    print("Logs closed. Awaiting writer threads to completely save files...")
    time.sleep(3.0)

    # ========================== VERIFICATION PHASE ==========================
    print("\n=================== VERIFICATION PHASE ===================")
    
    # 1. Check raw log file existence & columns
    if raw_path.exists() and raw_path.stat().st_size > 0:
        rep.ok("Raw log file created", f"{raw_path.name} ({raw_path.stat().st_size} bytes)")
        try:
            wb_raw = openpyxl.load_workbook(raw_path, read_only=True)
            if "Metadata" in wb_raw.sheetnames and "Data" in wb_raw.sheetnames:
                rep.ok("Raw log contains Metadata and Data sheets")
                
                # Check Data columns
                data_ws = wb_raw["Data"]
                headers = [cell.value for cell in next(data_ws.iter_rows(max_row=1))]
                expected_cols = ["timestamp", "direction", "hex", "delta_t_ms"]
                if headers == expected_cols:
                    rep.ok("Raw log Data sheet headers are correct")
                else:
                    rep.fail("Raw log headers mismatch", f"got {headers} expected {expected_cols}")
                
                # Check contents (RX and TX)
                directions = set()
                rows_count = 0
                timestamps = []
                for row in data_ws.iter_rows(min_row=2, max_row=200, values_only=True):
                    if row[1]:
                        directions.add(row[1])
                    if row[0]:
                        timestamps.append(row[0])
                    rows_count += 1
                
                if "RX" in directions and "TX" in directions:
                    rep.ok("Raw log contains both RX packets and TX commands")
                else:
                    rep.fail("Raw log missing either RX or TX direction data", f"found directions: {directions}")
                
                # Check timings
                if len(timestamps) > 1:
                    is_ordered = True
                    for i in range(1, len(timestamps)):
                        try:
                            t1 = datetime.strptime(timestamps[i-1], "%Y-%m-%d %H:%M:%S.%f")
                            t2 = datetime.strptime(timestamps[i], "%Y-%m-%d %H:%M:%S.%f")
                            if t2 < t1:
                                is_ordered = False
                                break
                        except ValueError:
                            pass
                    if is_ordered:
                        rep.ok("Raw log timestamps are sequential and monotonic")
                    else:
                        rep.fail("Raw log timestamps out-of-order detected")
            else:
                rep.fail("Raw log missing required sheets", f"found {wb_raw.sheetnames}")
            wb_raw.close()
        except Exception as e:
            rep.fail("Loading raw log Excel", str(e))
    else:
        rep.fail("Raw log file missing or empty")

    # 2. Check decoded log file existence & columns & command values
    if dec_path.exists() and dec_path.stat().st_size > 0:
        rep.ok("Decoded log file created", f"{dec_path.name} ({dec_path.stat().st_size} bytes)")
        try:
            wb_dec = openpyxl.load_workbook(dec_path, read_only=True)
            if "Metadata" in wb_dec.sheetnames and "Data" in wb_dec.sheetnames:
                rep.ok("Decoded log contains Metadata and Data sheets")
                
                # Check Data columns
                data_ws = wb_dec["Data"]
                headers = [cell.value for cell in next(data_ws.iter_rows(max_row=1))]
                
                # Verify that BMS Status, BMS Settings, and Status Flags are in headers
                has_status = any("BMS Status" in h for h in headers if h)
                has_settings = any("BMS Settings" in h for h in headers if h)
                has_flags = any("Status Flags" in h for h in headers if h)
                
                if has_status and has_settings and has_flags:
                    rep.ok("Decoded log contains columns for all frame groups")
                else:
                    rep.fail("Decoded log headers missing frame groups", f"status={has_status}, settings={has_settings}, flags={has_flags}")
                
                # Verify that Voltage Limit changes are reflected in the log
                # We can find the column for Voltage Limit
                v_lim_idx = None
                for idx, h in enumerate(headers):
                    if h and "Voltage Limit" in h:
                        v_lim_idx = idx
                        break
                
                if v_lim_idx is not None:
                    logged_limits = set()
                    for row in data_ws.iter_rows(min_row=2, values_only=True):
                        val = row[v_lim_idx]
                        if val is not None:
                            logged_limits.add(float(val))
                    
                    expected_set = {55.0, 52.0, 58.0, 45.0, 58.5, 50.0}
                    # Filter out any other transient limits or seed value (firmware defaults)
                    intersection = expected_set.intersection(logged_limits)
                    if len(intersection) >= 4: # allow some tolerance if command was sent at the very end
                        rep.ok(f"Voltage Limit updates correctly captured in log: {sorted(intersection)} V")
                    else:
                        rep.fail("Voltage Limit updates missing in log", f"expected {expected_set}, logged {logged_limits}")
                else:
                    rep.fail("Voltage Limit column missing in decoded log")
            else:
                rep.fail("Decoded log missing required sheets", f"found {wb_dec.sheetnames}")
            wb_dec.close()
        except Exception as e:
            rep.fail("Loading decoded log Excel", str(e))
    else:
        rep.fail("Decoded log file missing or empty")

    # 3. Generate and verify plot
    try:
        generate_plot(dec_path, plot_path)
        if plot_path.exists() and plot_path.stat().st_size > 0:
            rep.ok("Matplotlib telemetry plot generated successfully", f"{plot_path.name}")
        else:
            rep.fail("Plot file size is 0 bytes")
    except Exception as e:
        rep.fail("Generating matplotlib plot", str(e))

    # ========================== FINAL REPORT ==========================
    print(f"\n=================== TEST RESULT SUMMARY ===================")
    print(f"PASSED: {len(rep.passed)}")
    print(f"FAILED: {len(rep.failed)}")
    if rep.failed:
        for f in rep.failed:
            print(f"  FAIL: {f}")
        return len(rep.failed)
    
    print("ALL TESTS PASSED SUCCESSFULLY!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
