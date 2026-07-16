from __future__ import annotations

import time
from datetime import datetime

from openpyxl import load_workbook

from app.decoder.frame_decoder import DecodedFrame, DecodedSignal
from app.decoder.template_io import snapshot_config
from app.decoder.types import CalcGroupSpec, FrameConfig, FrameDefinition, SignalSpec
from app.serial_logging.decoded_logger import DecodedLogger
from app.serial_logging.raw_logger import RawLogger
from tests.conftest import dummy_protocol_config


def test_raw_logger_writes_xlsx_with_header(tmp_path):
    path = tmp_path / "raw.xlsx"
    with RawLogger(path) as logger:
        logger.log(
            "RX",
            bytes.fromhex("AA550010040FA00BB8BE70"),
            datetime(2026, 5, 4, 12, 37, 37, 125000),
        )
    wb = load_workbook(path, read_only=True)
    assert wb.sheetnames == ["Metadata", "Data"]
    data_rows = list(wb["Data"].iter_rows(values_only=True))
    wb.close()
    assert data_rows[0] == ("timestamp", "direction", "hex", "delta_t_ms")
    assert data_rows[1] == ("2026-05-04 12:37:37.125", "RX", "AA 55 00 10 04 0F A0 0B B8 BE 70", 0.0)


def test_raw_logger_compact_hex_format(tmp_path):
    """hex_format='compact' writes contiguous uppercase bytes (no spaces)."""
    path = tmp_path / "raw.xlsx"
    with RawLogger(path, hex_format="compact") as logger:
        logger.log(
            "RX",
            bytes.fromhex("AA550010040FA00BB8BE70"),
            datetime(2026, 5, 4, 12, 37, 37, 125000),
        )
    wb = load_workbook(path, read_only=True)
    data_rows = list(wb["Data"].iter_rows(values_only=True))
    wb.close()
    assert data_rows[1] == ("2026-05-04 12:37:37.125", "RX", "AA550010040FA00BB8BE70", 0.0)


def test_raw_logger_invalid_hex_format_rejected(tmp_path):
    import pytest
    with pytest.raises(ValueError, match="hex_format must be"):
        RawLogger(tmp_path / "raw.xlsx", hex_format="binary")


def _make_test_config(signal_b_name: str = "Pack_V") -> FrameConfig:
    protocol = dummy_protocol_config()
    frame_a = FrameDefinition(frame_id=0x0100, frame_name="FrameA")
    frame_b = FrameDefinition(frame_id=0x0200, frame_name="FrameB")

    signals_a = [
        SignalSpec(
            frame_id=0x0100,
            frame_name="FrameA",
            signal_name="Cell_V1",
            start_byte=0,
            byte_length=2,
            endianness="little",
            data_type="uint16",
            scale=0.001,
            offset=0.0,
            unit="V",
            group="Cells",
        ),
        SignalSpec(
            frame_id=0x0100,
            frame_name="FrameA",
            signal_name="Pack_I",
            start_byte=2,
            byte_length=2,
            endianness="little",
            data_type="int16",
            scale=0.1,
            offset=0.0,
            unit="",
        ),
    ]
    signals_b = [
        SignalSpec(
            frame_id=0x0200,
            frame_name="FrameB",
            signal_name=signal_b_name,
            start_byte=0,
            byte_length=2,
            endianness="little",
            data_type="uint16",
            scale=0.1,
            offset=0.0,
            unit="V",
        )
    ]

    return FrameConfig(
        protocol=protocol,
        frames={0x0100: frame_a, 0x0200: frame_b},
        signals_by_frame={0x0100: signals_a, 0x0200: signals_b},
        frame_names={0x0100: "FrameA", 0x0200: "FrameB"},
        calc_groups=[
            CalcGroupSpec(group="Cells", stat="avg", unit="V", frame_id=0x0100)
        ],
    )


def test_decoded_logger_writes_wide_rows(tmp_path):
    config = _make_test_config()
    path = tmp_path / "decoded.xlsx"

    frame_a = DecodedFrame(
        frame_id=0x0100,
        frame_name="FrameA",
        signals=[
            DecodedSignal(
                frame_id=0x0100,
                frame_name="FrameA",
                signal_name="Cell_V1",
                raw_value=3850,
                scaled_value=3.85,
                unit="V",
                status="ok",
                group="Cells",
            ),
            DecodedSignal(
                frame_id=0x0100,
                frame_name="FrameA",
                signal_name="Pack_I",
                raw_value=121,
                scaled_value=12.1,
                unit="",
                status="ok",
            ),
        ],
        calculations=[
            DecodedSignal(
                frame_id=0x0100,
                frame_name="FrameA",
                signal_name="Cells avg",
                raw_value=None,
                scaled_value=3.9,
                unit="V",
                status="ok",
                group="Cells",
                display_value="3.9",
                is_calculated=True,
            )
        ],
    )

    frame_b = DecodedFrame(
        frame_id=0x0200,
        frame_name="FrameB",
        signals=[
            DecodedSignal(
                frame_id=0x0200,
                frame_name="FrameB",
                signal_name="Pack_V",
                raw_value=482,
                scaled_value=48.2,
                unit="V",
                status="ok",
            )
        ],
    )

    # Cycle 1: both frames arrive (FrameB is the trigger as the last in config).
    # Cycle 2: FrameA arrives at 2000.
    # Cycle 3: FrameA arrives at 3000. This is a duplicate of FrameA, so the buffer (containing FrameA at 2000) is emitted as an incomplete cycle.
    # Cycle 4: FrameB arrives at 3100. This is the trigger frame, so the buffer (containing FrameA at 3000 and FrameB at 3100) is emitted.
    with DecodedLogger(path, config) as logger:
        logger.log_frame(frame_a, 1000)
        logger.log_frame(frame_b, 1100)
        logger.log_frame(frame_a, 2000)
        logger.log_frame(frame_a, 3000)
        logger.log_frame(frame_b, 3100)

    wb = load_workbook(path, read_only=True)
    assert wb.sheetnames == [DecodedLogger.METADATA_SHEET, DecodedLogger.DATA_SHEET]
    data_rows = list(wb[DecodedLogger.DATA_SHEET].iter_rows(values_only=True))
    wb.close()

    expected_header = (
        "0x100.elapsed_ms",
        "0x100.frame_id",
        "0x100.Cell_V1 (V)",
        "0x100.Pack_I",
        "Cells avg (V)",
        "0x200.elapsed_ms",
        "0x200.frame_id",
        "0x200.Pack_V (V)",
    )
    assert data_rows[0] == expected_header
    # 3 complete/incomplete cycles → 3 data rows.
    assert len(data_rows) == 4

    header = list(expected_header)
    row1 = dict(zip(header, data_rows[1], strict=False))
    assert row1["0x100.elapsed_ms"] == 1000
    assert row1["0x100.frame_id"] == "0x0100"
    assert row1["0x100.Cell_V1 (V)"] == 3.85
    assert row1["0x100.Pack_I"] == 12.1
    assert row1["Cells avg (V)"] == 3.9
    assert row1["0x200.elapsed_ms"] == 1100
    assert row1["0x200.frame_id"] == "0x0200"
    assert row1["0x200.Pack_V (V)"] == 48.2

    # Second row is the incomplete cycle emitted when duplicate FrameA arrives
    row2 = dict(zip(header, data_rows[2], strict=False))
    assert row2["0x100.elapsed_ms"] == 2000
    assert row2["0x100.frame_id"] == "0x0100"
    assert row2["0x200.elapsed_ms"] is None

    # Third row is the cycle emitted when FrameB arrives
    row3 = dict(zip(header, data_rows[3], strict=False))
    assert row3["0x100.elapsed_ms"] == 3000
    assert row3["0x200.elapsed_ms"] == 3100


def test_decoded_logger_writes_metadata_sheet(tmp_path):
    config = _make_test_config()
    path = tmp_path / "decoded.xlsx"
    metadata = {
        "app": "Bytehound",
        "app_version": "0.2.0",
        "baud_rate": "115200",
        "serial_port": "COM9",
        "logging_mode": "Raw + Decoded",
        "session_started": "2026-05-15 11:23:52",
    }

    frame = DecodedFrame(
        frame_id=0x0100,
        frame_name="FrameA",
        signals=[
            DecodedSignal(
                frame_id=0x0100,
                frame_name="FrameA",
                signal_name="Cell_V1",
                raw_value=3850,
                scaled_value=3.85,
                unit="V",
                status="ok",
                group="Cells",
            )
        ],
    )

    with DecodedLogger(path, config, metadata=metadata) as logger:
        logger.log_frame(frame, 1000, datetime(2026, 5, 4, 12, 37, 37, 125000))

    wb = load_workbook(path, read_only=True)
    meta_rows = list(wb[DecodedLogger.METADATA_SHEET].iter_rows(values_only=True))
    wb.close()

    assert meta_rows[0] == ("Key", "Value")
    meta_pairs = dict(meta_rows[1:])
    for key, value in metadata.items():
        assert meta_pairs[key] == value


def test_decoded_logger_writes_incomplete_cycles(tmp_path):
    """Only FrameA arrives, never FrameB (the trigger). It emits incomplete cycles when FrameA repeats."""
    config = _make_test_config()
    path = tmp_path / "decoded.xlsx"

    frame_a = DecodedFrame(
        frame_id=0x0100,
        frame_name="FrameA",
        signals=[
            DecodedSignal(
                frame_id=0x0100,
                frame_name="FrameA",
                signal_name="Cell_V1",
                raw_value=3850,
                scaled_value=3.85,
                unit="V",
                status="ok",
                group="Cells",
            )
        ],
    )

    with DecodedLogger(path, config) as logger:
        logger.log_frame(frame_a, 1000)
        logger.log_frame(frame_a, 2000)
        logger.log_frame(frame_a, 3000)

    wb = load_workbook(path, read_only=True)
    data_rows = list(wb[DecodedLogger.DATA_SHEET].iter_rows(values_only=True))
    wb.close()

    # The header row + 2 incomplete cycle rows (emitted when the duplicate 2000 and 3000 frames arrive)
    assert len(data_rows) == 3


def test_snapshot_config_copies_csv_templates(resources_dir, tmp_path):
    snapshot = snapshot_config(resources_dir / "config_template", tmp_path)
    assert snapshot.is_dir()
    assert (snapshot / "variables.csv").exists()


# ─── Regression: writer-thread ownership transfer on slow-disk shutdown ─────
#
# Previously, RawLogger.close() called self._fp.close() and DecodedLogger.close()
# set self._workbook = None right after the writer-thread join timed out. If the
# writer was still draining a backlog (slow USB / network share), the UI-thread
# tear-down destroyed the resources the writer was using — raw_logger raised
# ValueError on its next writerow() and decoded_logger silently skipped wb.save(),
# producing partial or empty files.
#
# The fix captures `fp` / `wb` / `ws` as locals at the top of _writer_loop so
# the writer thread keeps its own strong references. Even if close() returns
# after a join timeout, the writer continues, drains the queue, and persists
# everything before exiting cleanly.
#
# Both tests simulate the slow-disk scenario by:
#   1. Patching the per-row write to sleep, so the queue cannot drain in time.
#   2. Patching the writer thread's join to use a near-zero timeout, so close()
#      definitely times out and returns while the writer is still alive.
#   3. After close() returns, joining the writer thread directly with a
#      generous timeout and verifying every row landed on disk.


def _shorten_join(thread, timeout_sec: float = 0.05):
    """Replace thread.join with a wrapper that ignores the caller's timeout
    and uses a near-zero one instead. Returns the original join so the test
    can wait for the writer to truly finish.
    """
    orig = thread.join
    thread.join = lambda *_, **__: orig(timeout=timeout_sec)
    return orig


def test_raw_logger_persists_all_rows_when_close_join_times_out(tmp_path):
    path = tmp_path / "slow_drain.xlsx"
    rl = RawLogger(path)
    rl.open()

    # Inject a per-row delay big enough that even our short forced-timeout
    # can't drain N rows in time.
    orig_write_one = rl._write_one
    def slow_write_one(writer, item):
        time.sleep(0.05)
        return orig_write_one(writer, item)
    rl._write_one = slow_write_one

    N = 5
    ts = datetime(2026, 5, 19, 12, 0, 0)
    for i in range(N):
        rl.log("RX", bytes([i & 0xFF, (i + 1) & 0xFF]), ts)

    orig_join = _shorten_join(rl._writer_thread, timeout_sec=0.05)
    writer_thread = rl._writer_thread  # save ref before close() nils it
    rl.close()

    # Without the fix: rl.close() would have closed self._fp during cleanup
    # and the writer thread would have died with rows still in the queue.
    # With the fix: writer thread holds local fp ref, keeps draining, then
    # flushes + closes the file in its finally block.
    orig_join(timeout=5.0)
    assert not writer_thread.is_alive(), (
        "Writer thread should have completed its drain + close cleanly"
    )

    wb = load_workbook(path, read_only=True)
    data_rows = list(wb["Data"].iter_rows(values_only=True))
    wb.close()

    # 1 header row + N data rows.
    assert len(data_rows) == N + 1, (
        f"Expected {N + 1} lines after slow-drain shutdown, got {len(data_rows)}"
    )
    # Each data row's hex column reflects one of the bytes we logged.
    written_hex = {row[2] for row in data_rows[1:]}
    expected_hex = {f"{i & 0xFF:02X} {(i + 1) & 0xFF:02X}" for i in range(N)}
    assert written_hex == expected_hex


def test_raw_logger_await_drain_blocks_until_writer_finishes(tmp_path):
    """After close() returns with the writer still draining, await_drain()
    must block the caller until the writer has finished and the file is
    on disk. This is what MainWindow.closeEvent uses to honour the
    zero-data-loss guarantee — without it, the daemon writer thread
    would be killed at interpreter exit mid-flush.
    """
    path = tmp_path / "await_drain.xlsx"
    rl = RawLogger(path)
    rl.open()

    # Slow writer: 50ms/row. Forced 50ms join in close() guarantees the
    # writer is still alive when close() returns.
    orig_write_one = rl._write_one
    def slow_write_one(writer, item):
        time.sleep(0.05)
        return orig_write_one(writer, item)
    rl._write_one = slow_write_one

    N = 4
    ts = datetime(2026, 5, 19, 12, 0, 0)
    for i in range(N):
        rl.log("RX", bytes([i]), ts)

    writer_thread = rl._writer_thread
    orig_join = _shorten_join(writer_thread, timeout_sec=0.05)
    rl.close()

    # close() returned but the writer is still running — observable via
    # the new is_draining() API.
    assert rl.is_draining(), "Writer should still be alive after timed-out close()"
    assert rl.pending_rows() >= 0  # smoke check

    # Restore the real join now that close() has used the shortened one;
    # await_drain needs the unpatched join to actually wait for completion.
    writer_thread.join = orig_join

    # await_drain blocks until the writer truly finishes (or times out).
    finished = rl.await_drain(timeout=5.0)
    assert finished, "Writer thread should have completed within 5s"
    assert not rl.is_draining(), "is_draining() must be False once writer exits"

    # File on disk has every row.
    wb = load_workbook(path, read_only=True)
    data_rows = list(wb["Data"].iter_rows(values_only=True))
    wb.close()
    assert len(data_rows) == N + 1, (
        f"Expected {N + 1} lines on disk after await_drain, got {len(data_rows)}"
    )


def test_raw_logger_await_drain_returns_true_when_no_writer(tmp_path):
    """await_drain() on a never-opened (or already-closed-and-finished)
    logger must return True immediately — the contract is "writer is done"
    not "writer existed and finished"."""
    path = tmp_path / "noop_drain.csv"
    rl = RawLogger(path)
    # Never opened.
    assert rl.await_drain(timeout=0.1) is True
    assert rl.is_draining() is False


def test_raw_logger_open_after_timed_out_close_reuses_dead_thread_slot(tmp_path):
    """If a prior close() left a dead writer-thread reference (because
    await_drain wasn't called before the thread finished naturally), the
    next open() must clear the stale reference and start a fresh writer.
    """
    path = tmp_path / "reuse.csv"
    rl = RawLogger(path)
    rl.open()
    rl.log("RX", b"\x01", datetime(2026, 5, 19, 12, 0, 0))
    rl.close()  # writer finishes within the 5s join → ref cleared
    # Simulate a leftover dead thread reference (the case open() must handle).
    import threading
    dead = threading.Thread(target=lambda: None, daemon=True)
    dead.start()
    dead.join()
    rl._writer_thread = dead
    # open() must notice the dead thread and reset, then start a fresh writer.
    rl.open()
    assert rl._writer_thread is not None
    assert rl._writer_thread is not dead
    assert rl._writer_thread.is_alive()
    rl.close()


def test_decoded_logger_persists_workbook_when_close_join_times_out(tmp_path):
    config = _make_test_config()
    path = tmp_path / "slow_drain.xlsx"
    dl = DecodedLogger(path, config)
    dl.open()

    # Inject a per-row delay on ws.append.
    orig_write_one = dl._write_one
    def slow_write_one(ws, item):
        time.sleep(0.05)
        return orig_write_one(ws, item)
    dl._write_one = slow_write_one

    # Build minimal frame_a + frame_b objects (frame_b is the trigger as the
    # last frame in config). Each (a, b) pair produces one row.
    def _frame_a(elapsed_idx: int) -> DecodedFrame:
        return DecodedFrame(
            frame_id=0x0100, frame_name="FrameA",
            signals=[
                DecodedSignal(frame_id=0x0100, frame_name="FrameA",
                              signal_name="Cell_V1", raw_value=3850 + elapsed_idx,
                              scaled_value=3.85 + elapsed_idx * 0.001, unit="V",
                              status="ok", group="Cells"),
                DecodedSignal(frame_id=0x0100, frame_name="FrameA",
                              signal_name="Pack_I", raw_value=121, scaled_value=12.1,
                              unit="", status="ok"),
            ],
        )

    def _frame_b() -> DecodedFrame:
        return DecodedFrame(
            frame_id=0x0200, frame_name="FrameB",
            signals=[
                DecodedSignal(frame_id=0x0200, frame_name="FrameB",
                              signal_name="Pack_V", raw_value=482, scaled_value=48.2,
                              unit="V", status="ok"),
            ],
        )

    N = 5
    for i in range(N):
        dl.log_frame(_frame_a(i), 1000 * (i + 1))
        dl.log_frame(_frame_b(), 1000 * (i + 1) + 100)

    orig_join = _shorten_join(dl._writer_thread, timeout_sec=0.05)
    writer_thread = dl._writer_thread
    dl.close()

    # Without the fix: close() nullified self._workbook before the writer
    # thread reached its finally, so _save_and_close_workbook saw None and
    # skipped wb.save(). The .xlsx on disk would be unreadable / empty.
    # With the fix: writer captured `wb` locally and saves cleanly.
    orig_join(timeout=10.0)
    assert not writer_thread.is_alive(), (
        "DecodedLogger writer thread should have saved + closed the workbook"
    )

    wb = load_workbook(path, read_only=True)
    data_rows = list(wb[DecodedLogger.DATA_SHEET].iter_rows(values_only=True))
    wb.close()
    # 1 header row + N data rows for N complete cycles.
    assert len(data_rows) == N + 1, (
        f"Expected {N + 1} rows on disk after slow-drain shutdown, got {len(data_rows)}"
    )


def test_decoded_logger_temporary_file_recovery(tmp_path):
    config = _make_test_config()
    path = tmp_path / "crash_test.xlsx"

    metadata = {"app": "Bytehound", "test": "recovery"}
    dl = DecodedLogger(path, config, metadata=metadata)
    dl.open()

    def _frame_a(elapsed_idx: int) -> DecodedFrame:
        return DecodedFrame(
            frame_id=0x0100, frame_name="FrameA",
            signals=[
                DecodedSignal(frame_id=0x0100, frame_name="FrameA",
                              signal_name="Cell_V1", raw_value=3850 + elapsed_idx,
                              scaled_value=3.85 + elapsed_idx * 0.001, unit="V",
                              status="ok", group="Cells"),
                DecodedSignal(frame_id=0x0100, frame_name="FrameA",
                              signal_name="Pack_I", raw_value=121, scaled_value=12.1,
                              unit="", status="ok"),
            ],
        )

    def _frame_b() -> DecodedFrame:
        return DecodedFrame(
            frame_id=0x0200, frame_name="FrameB",
            signals=[
                DecodedSignal(frame_id=0x0200, frame_name="FrameB",
                              signal_name="Pack_V", raw_value=482, scaled_value=48.2,
                              unit="V", status="ok"),
            ],
        )

    # Log one cycle
    dl.log_frame(_frame_a(1), 1000)
    dl.log_frame(_frame_b(), 1100)

    # Wait for the item to be processed from the queue
    import time
    for _ in range(50):
        if dl._queue.qsize() == 0:
            break
        time.sleep(0.05)

    # Patch _compile_workbook to be a no-op to simulate a crash/interruption
    dl._compile_workbook = lambda: None
    dl.close()

    tmp_data = path.with_suffix(path.suffix + ".tmp_data")
    tmp_meta = path.with_suffix(path.suffix + ".tmp_meta")

    # Check that temp files exist but target .xlsx does not
    assert tmp_data.exists()
    assert tmp_meta.exists()
    assert not path.exists()

    # Perform recovery
    DecodedLogger.recover_temp_files(tmp_data, tmp_meta, path)

    # Verify temp files deleted
    assert not tmp_data.exists()
    assert not tmp_meta.exists()
    assert path.exists()

    # Verify data content
    wb = load_workbook(path, read_only=True)
    assert wb.sheetnames == [DecodedLogger.METADATA_SHEET, DecodedLogger.DATA_SHEET]

    meta_rows = list(wb[DecodedLogger.METADATA_SHEET].iter_rows(values_only=True))
    assert meta_rows[0] == ("Key", "Value")
    meta_dict = dict(meta_rows[1:])
    assert meta_dict["app"] == "Bytehound"
    assert meta_dict["test"] == "recovery"

    data_rows = list(wb[DecodedLogger.DATA_SHEET].iter_rows(values_only=True))
    wb.close()

    assert len(data_rows) == 2  # header + 1 cycle row


def test_raw_logger_queue_saturation_warning(tmp_path):
    import queue
    log_path = tmp_path / "raw_sat.csv"
    warnings = []

    def on_warn(msg):
        warnings.append(msg)

    logger = RawLogger(log_path, on_warning=on_warn)
    logger.open()

    # Restrict queue size to 1 to force overflow easily
    logger._queue = queue.Queue(maxsize=1)

    # Call log multiple times to trigger overflow and rate-limited warning
    logger.log("TX", b"\x11\x22")
    logger.log("TX", b"\x33\x44")
    logger.log("TX", b"\x55\x66")

    assert logger._dropped_count > 0
    assert len(warnings) >= 1
    assert "RawLogger queue is full" in warnings[0]

    logger.close()
    assert any("dropped" in w for w in warnings)


def test_decoded_logger_queue_saturation_warning(tmp_path):
    import queue
    from app.decoder.frame_decoder import DecodedFrame

    log_path = tmp_path / "decoded_sat.xlsx"
    config = _make_test_config()
    warnings = []

    def on_warn(msg):
        warnings.append(msg)

    logger = DecodedLogger(log_path, config, on_warning=on_warn)
    logger.open()

    # Restrict queue size to 1 to force overflow
    logger._queue = queue.Queue(maxsize=1)
    logger._cycle_frame_ids = [0x0100]
    logger._trigger_id = 0x0100
    logger._cycle_buffer[0x0100] = {0: "data"}

    frame = DecodedFrame(frame_id=0x0100, frame_name="FrameA", signals=[])
    logger.log_frame(frame, 1.0)
    logger.log_frame(frame, 1.0)
    logger.log_frame(frame, 1.0)

    assert logger._dropped_count > 0
    assert len(warnings) >= 1
    assert "DecodedLogger queue is full" in warnings[0]

    logger.close()
    assert any("dropped" in w for w in warnings)


def test_decoded_logger_retains_incomplete_cycles_in_polling_mode(tmp_path):
    """With polling_mode=True, incomplete cycles are NOT dropped. They are written to disk."""
    config = _make_test_config()
    path = tmp_path / "decoded_polling.xlsx"

    frame_a = DecodedFrame(
        frame_id=0x0100,
        frame_name="FrameA",
        signals=[
            DecodedSignal(
                frame_id=0x0100,
                frame_name="FrameA",
                signal_name="Cell_V1",
                raw_value=3850,
                scaled_value=3.85,
                unit="V",
                status="ok",
                group="Cells",
            )
        ],
    )

    with DecodedLogger(path, config, polling_mode=True) as logger:
        logger.log_frame(frame_a, 1000)
        logger.log_frame(frame_a, 2000)
        logger.log_frame(frame_a, 3000)

    wb = load_workbook(path, read_only=True)
    data_rows = list(wb[DecodedLogger.DATA_SHEET].iter_rows(values_only=True))
    wb.close()

    # Header row + 3 cycle rows (1000, 2000, 3000)
    assert len(data_rows) == 4


def test_decoded_logger_handles_tx_only_frames(tmp_path):
    """TX-only frames are excluded from required cycle frame IDs, and do not block logging."""
    protocol = dummy_protocol_config()
    frame_rx = FrameDefinition(frame_id=0x0100, frame_name="FrameRX", direction="rx")
    frame_tx = FrameDefinition(frame_id=0x0200, frame_name="FrameTX", direction="tx")

    signals_rx = [
        SignalSpec(
            frame_id=0x0100,
            frame_name="FrameRX",
            signal_name="SigRX",
            start_byte=0,
            byte_length=2,
            endianness="little",
            data_type="uint16",
            scale=1.0,
            offset=0.0,
            unit="",
        )
    ]
    signals_tx = [
        SignalSpec(
            frame_id=0x0200,
            frame_name="FrameTX",
            signal_name="SigTX",
            start_byte=0,
            byte_length=2,
            endianness="little",
            data_type="uint16",
            scale=1.0,
            offset=0.0,
            unit="",
        )
    ]

    config = FrameConfig(
        protocol=protocol,
        frames={0x0100: frame_rx, 0x0200: frame_tx},
        signals_by_frame={0x0100: signals_rx, 0x0200: signals_tx},
        frame_names={0x0100: "FrameRX", 0x0200: "FrameTX"},
    )

    path = tmp_path / "decoded_tx.xlsx"

    # FrameRX is the trigger because it's the last (and only) RX-capable frame.
    logger = DecodedLogger(path, config)
    assert logger._required_cycle_frame_ids == [0x0100]
    assert logger._trigger_id == 0x0100

    frame_rx_dec = DecodedFrame(
        frame_id=0x0100,
        frame_name="FrameRX",
        signals=[
            DecodedSignal(
                frame_id=0x0100,
                frame_name="FrameRX",
                signal_name="SigRX",
                raw_value=42,
                scaled_value=42.0,
                unit="",
                status="ok",
            )
        ]
    )

    frame_tx_dec = DecodedFrame(
        frame_id=0x0200,
        frame_name="FrameTX",
        signals=[
            DecodedSignal(
                frame_id=0x0200,
                frame_name="FrameTX",
                signal_name="SigTX",
                raw_value=100,
                scaled_value=100.0,
                unit="",
                status="ok",
            )
        ]
    )

    with logger:
        # Cycle 1: Only RX arrives. Trigger fires immediately because only RX is required.
        logger.log_frame(frame_rx_dec, 1000)

        # Cycle 2: TX arrives, then RX. Trigger fires and both are written.
        logger.log_frame(frame_tx_dec, 1900)
        logger.log_frame(frame_rx_dec, 2000)

    wb = load_workbook(path, read_only=True)
    data_rows = list(wb[DecodedLogger.DATA_SHEET].iter_rows(values_only=True))
    wb.close()

    # Header + 2 data rows
    assert len(data_rows) == 3

    # Headers check
    headers = data_rows[0]
    assert "0x100.SigRX" in headers
    assert "0x200.SigTX" in headers

    # Row 1 (Cycle 1): RX present, TX empty/None
    row1 = dict(zip(headers, data_rows[1], strict=False))
    assert row1["0x100.SigRX"] == 42.0
    assert row1["0x200.SigTX"] in (None, "", " ")

    # Row 2 (Cycle 2): Both present
    row2 = dict(zip(headers, data_rows[2], strict=False))
    assert row2["0x100.SigRX"] == 42.0
    assert row2["0x200.SigTX"] == 100.0


def test_decoded_logger_rxtx_with_command_does_not_block_cycle(tmp_path):
    """An RXTX frame that has an associated TX command is excluded from required cycle frame IDs."""
    protocol = dummy_protocol_config()
    frame_rx = FrameDefinition(frame_id=0x0100, frame_name="FrameRX", direction="rx")
    frame_cmd = FrameDefinition(frame_id=0x0200, frame_name="FrameCMD", direction="rxtx")

    signals_rx = [
        SignalSpec(
            frame_id=0x0100,
            frame_name="FrameRX",
            signal_name="SigRX",
            start_byte=0,
            byte_length=2,
            endianness="little",
            data_type="uint16",
            scale=1.0,
            offset=0.0,
            unit="",
        )
    ]
    signals_cmd = [
        SignalSpec(
            frame_id=0x0200,
            frame_name="FrameCMD",
            signal_name="SigCMD",
            start_byte=0,
            byte_length=2,
            endianness="little",
            data_type="uint16",
            scale=1.0,
            offset=0.0,
            unit="",
        )
    ]

    # Create a TX command on FrameCMD (0x0200)
    from app.decoder.types import TxCommandSpec
    tx_commands = {
        "TestCommand": TxCommandSpec(command_name="TestCommand", frame_id=0x0200)
    }

    config = FrameConfig(
        protocol=protocol,
        frames={0x0100: frame_rx, 0x0200: frame_cmd},
        signals_by_frame={0x0100: signals_rx, 0x0200: signals_cmd},
        frame_names={0x0100: "FrameRX", 0x0200: "FrameCMD"},
        tx_commands=tx_commands,
    )

    path = tmp_path / "decoded_rxtx_cmd.xlsx"

    # FrameRX is strictly RX, FrameCMD is RXTX but has a TX command, so it is excluded.
    logger = DecodedLogger(path, config)
    assert logger._required_cycle_frame_ids == [0x0100]
    assert logger._trigger_id == 0x0100
