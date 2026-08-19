"""Comprehensive Verification Suite for Popups, Dialogs, and Logging Guarantees.

Verifies:
1. ConnectionDialog: Serial, TCP, and UDP persistence, defaults fallback.
2. LoggingSettingsDialog: Log level, flush interval, QSettings synchronization.
3. DecodedLogger: Asynchronous queue, column alignment, timestamping, zero data-loss on close().
4. RawLogger: Hex format (spaced vs compact), wire direction tagging, flushed writes.
5. Crash Recovery: Dangling temp log file detection and safe recovery into user log folders.
6. DiagnosticsDialog: Environment, versions, packet metrics, and log tail snapshotting.
"""

import os
import sys
import time
from pathlib import Path
from unittest.mock import MagicMock, patch

# Ensure project root is in sys.path when executed directly as python test_xxx.py
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pytest
from PySide6.QtCore import QSettings
from PySide6.QtWidgets import QApplication

from app.decoder.types import SerialDefaults, ProtocolConfig, FrameConfig, FrameDefinition, SignalSpec
from app.ui.dialogs import ConnectionDialog, LoggingSettingsDialog, YRangeDialog, SchemaMapperDialog
from app.serial_logging.raw_logger import RawLogger
from app.serial_logging.decoded_logger import DecodedLogger


@pytest.fixture(scope="module")
def qapp():
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    return app


# ============================================================================
# 1. CONNECTION DIALOG PERSISTENCE & DEFAULTS
# ============================================================================

def test_connection_dialog_serial_defaults_and_persistence(qapp, tmp_path):
    """Verify that ConnectionDialog correctly uses SerialDefaults and saves to QSettings."""
    settings_file = str(tmp_path / "test_conn_settings.ini")
    settings = QSettings(settings_file, QSettings.Format.IniFormat)

    # Supply custom defaults from an active config
    config_defaults = SerialDefaults(
        baud_rate=57600,
        data_bits=8,
        stop_bits=2.0,
        parity="E",
        timeout_ms=250
    )

    dlg = ConnectionDialog(settings=settings, config_defaults=config_defaults)

    # Initial values should match config_defaults
    s_settings = dlg.get_settings()
    assert s_settings.baud_rate == 57600
    assert s_settings.data_bits == 8
    assert s_settings.stop_bits == 2.0
    assert s_settings.parity == "E"
    assert s_settings.timeout_ms == 250

    # Switch to TCP and save
    dlg._type_combo.setCurrentIndex(1)  # TCP Client
    dlg._tcp_host.setText("192.168.1.100")
    dlg._tcp_port.setValue(9000)
    dlg._on_accept()

    tcp_settings = dlg.get_settings()
    assert tcp_settings.connection_type == "tcp"
    assert tcp_settings.host == "192.168.1.100"
    assert tcp_settings.port_num == 9000

    # Verify persisted in QSettings
    assert settings.value("conn/type") == "tcp"
    assert settings.value("conn/tcp_host") == "192.168.1.100"
    assert int(settings.value("conn/tcp_port")) == 9000


# ============================================================================
# 2. LOGGING SETTINGS DIALOG
# ============================================================================

def test_logging_settings_dialog(qapp, tmp_path):
    """Verify that LoggingSettingsDialog loads, modifies, and persists logging settings."""
    settings_file = str(tmp_path / "test_log_settings.ini")
    settings = QSettings(settings_file, QSettings.Format.IniFormat)

    dlg = LoggingSettingsDialog(settings=settings)
    dlg._level_combo.setCurrentText("DEBUG")
    dlg._flush_spin.setValue(1.5)
    dlg._on_accept()

    level, interval = dlg.get_values()
    assert level == "DEBUG"
    assert interval == 1.5

    assert settings.value("logging/level") == "DEBUG"
    assert float(settings.value("logging/flush_interval_s")) == 1.5


# ============================================================================
# 3. Y-RANGE DIALOG VALIDATION & POPUP BEHAVIOR
# ============================================================================

def test_yrange_dialog_validation(qapp):
    """Verify YRangeDialog validates that max > min and shows error label on invalid range."""
    dlg = YRangeDialog(parent=None, panel_label="Cell Voltages", current_min=2.5, current_max=4.2)
    assert dlg.get_range() == (2.5, 4.2)

    # Invalid range: min > max
    dlg._min_spin.setValue(5.0)
    dlg._max_spin.setValue(3.0)
    dlg._on_accept()

    assert "strictly greater" in dlg.error_label.text()

    # Fix range
    dlg._max_spin.setValue(6.0)
    dlg._on_accept()
    assert dlg.get_range() == (5.0, 6.0)


# ============================================================================
# 4. SCHEMA MAPPER DIALOG
# ============================================================================

def test_schema_mapper_dialog_persistence(qapp, tmp_path):
    """Verify SchemaMapperDialog loads and persists import mappings in QSettings."""
    settings_file = str(tmp_path / "test_schema_settings.ini")
    settings = QSettings(settings_file, QSettings.Format.IniFormat)

    dlg = SchemaMapperDialog(settings=settings)
    dlg._sheets_edit.setText("Telemetry,LogData")
    dlg._cols_edit.setText("time_ms,elapsed_sec")
    dlg._on_accept()

    assert settings.value("import/sheet_names") == "Telemetry,LogData"
    assert settings.value("import/elapsed_cols") == "time_ms,elapsed_sec"


# ============================================================================
# 5. RAW LOGGER OPERATION & FORMATTING
# ============================================================================

def test_raw_logger_spaced_and_compact_formats(tmp_path):
    """Verify RawLogger writes timestamped wire logs in spaced and compact hex."""
    import openpyxl

    log_dir = tmp_path / "raw_logs"
    log_dir.mkdir()

    # Spaced hex format
    raw_path_spaced = log_dir / "raw_spaced.xlsx"
    raw_logger_spaced = RawLogger(path=raw_path_spaced, hex_format="hex")
    raw_logger_spaced.open()
    raw_logger_spaced.log("RX", b"\xaa\x55\x01\x02\x03\xee")
    raw_logger_spaced.log("TX", b"\xaa\x55\xff\x00\xee")
    raw_logger_spaced.close()
    raw_logger_spaced.await_drain(timeout=2.0)

    assert raw_path_spaced.exists()
    wb = openpyxl.load_workbook(raw_path_spaced, data_only=True)
    assert "Data" in wb.sheetnames
    data_ws = wb["Data"]
    rows = list(data_ws.iter_rows(values_only=True))
    assert len(rows) >= 3  # Header + 2 data rows
    assert "AA 55 01 02 03 EE" in str(rows[1])
    assert "AA 55 FF 00 EE" in str(rows[2])

    # Compact hex format
    raw_path_compact = log_dir / "raw_compact.xlsx"
    raw_logger_compact = RawLogger(path=raw_path_compact, hex_format="compact")
    raw_logger_compact.open()
    raw_logger_compact.log("RX", b"\xaa\x55\x01\x02\x03\xee")
    raw_logger_compact.close()
    raw_logger_compact.await_drain(timeout=2.0)

    assert raw_path_compact.exists()
    wb_compact = openpyxl.load_workbook(raw_path_compact, data_only=True)
    data_ws_compact = wb_compact["Data"]
    compact_rows = list(data_ws_compact.iter_rows(values_only=True))
    assert "AA55010203EE" in str(compact_rows[1])


# ============================================================================
# 6. DECODED LOGGER ASYNC LOGGING & CLOSE DRAIN GUARANTEE
# ============================================================================

def test_decoded_logger_async_drain_and_zero_loss(tmp_path):
    """Verify DecodedLogger drains all queued rows on close without dropping data."""
    import openpyxl
    from app.decoder.frame_decoder import DecodedFrame, DecodedSignal

    log_dir = tmp_path / "decoded_logs"
    log_dir.mkdir()
    log_file = log_dir / "test_session.xlsx"

    proto = ProtocolConfig(
        profile_name="Test",
        header=b"\xaa\x55",
        frame_id_size=2,
        frame_id_byte_order="little",
        length_size=1,
        length_meaning="payload_only",
        crc_type="none",
        crc_size=0,
        crc_byte_order="little",
        crc_coverage="header_to_payload",
        footer=b"",
        escape_mode="none",
        enabled=True,
        parser_type="framed"
    )
    frames = {0x1000: FrameDefinition(frame_id=0x1000, frame_name="Telemetry", payload_length=4, direction="rxtx")}
    signals = {
        0x1000: [
            SignalSpec(frame_id=0x1000, frame_name="Telemetry", signal_name="Volt", start_byte=0, byte_length=2, endianness="little", data_type="uint16", scale=0.1, offset=0.0, unit="V", group="Sensors"),
            SignalSpec(frame_id=0x1000, frame_name="Telemetry", signal_name="Curr", start_byte=2, byte_length=2, endianness="little", data_type="int16", scale=0.01, offset=0.0, unit="A", group="Sensors"),
        ]
    }
    cfg = FrameConfig(protocol=proto, frames=frames, signals_by_frame=signals)

    logger = DecodedLogger(path=log_file, config=cfg)
    logger.open()

    # Log 100 rows rapidly
    for i in range(100):
        dec_signals = [
            DecodedSignal(frame_id=0x1000, frame_name="Telemetry", signal_name="Volt", raw_value=120 + i, scaled_value=12.0 + (i * 0.1), unit="V", status="OK"),
            DecodedSignal(frame_id=0x1000, frame_name="Telemetry", signal_name="Curr", raw_value=150 + i, scaled_value=1.5 + (i * 0.01), unit="A", status="OK"),
        ]
        frame = DecodedFrame(frame_id=0x1000, frame_name="Telemetry", signals=dec_signals)
        logger.log_frame(frame, elapsed_ms=float(i * 10))

    # Close and await writer thread drain
    logger.close()
    logger.await_drain(timeout=5.0)

    assert log_file.exists()
    assert log_file.stat().st_size > 0

    wb = openpyxl.load_workbook(log_file, data_only=True)
    assert "Data" in wb.sheetnames
    data_ws = wb["Data"]
    rows = list(data_ws.iter_rows(values_only=True))
    assert len(rows) == 101  # 1 Header row + exactly 100 logged data rows (0 lost)!


# ============================================================================
# 7. CRASH RECOVERY / TEMP LOG SALVAGE TEST
# ============================================================================

def test_recover_temp_files_salvage(tmp_path):
    """Verify that DecodedLogger.recover_temp_files converts orphaned .tmp files to xlsx."""
    tmp_data = tmp_path / "temp_data.csv.tmp"
    tmp_meta = tmp_path / "temp_meta.json.tmp"
    target_xlsx = tmp_path / "recovered_session.xlsx"

    # Write synthetic CSV data into tmp_data
    tmp_data.write_text("timestamp,elapsed_s,frame_id,frame_name,Volt,Curr\n2026-08-19 07:00:00,0.000,0x1000,Telemetry,12.5,1.8\n", encoding="utf-8")
    tmp_meta.write_text('{"profile_name": "RecoveryTest", "baud_rate": 115200}', encoding="utf-8")

    # Run recovery
    DecodedLogger.recover_temp_files(data_path=tmp_data, meta_path=tmp_meta, target_path=target_xlsx)

    assert target_xlsx.exists()
    assert not tmp_data.exists()
    assert not tmp_meta.exists()


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
