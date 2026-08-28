"""Export config templates and session snapshots."""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import pandas as pd


CONFIG_CSV_FILES = (
    "protocol.csv",
    "frames.csv",      # per-frame names and payload_length (replaces protocol-level payload_length)
    "variables.csv",
    "frame_config.csv",
    "bitfields.csv",
    "enums.csv",
    "calc_groups.csv",
    "tx_commands.csv",
    "tx_command_fields.csv",
    "serial_defaults.csv",
    "polling_schedule.csv",
)


def get_bundled_template_dir() -> Path:
    """Return the absolute path to the bundled configuration CSV templates directory."""
    return Path(__file__).resolve().parents[1] / "resources" / "config_template"


def load_template_tables(source_dir: str | Path | None = None) -> dict[str, pd.DataFrame]:
    """Read all configuration CSV files into a dictionary of DataFrames keyed by Excel sheet name."""
    import pandas as pd

    source = Path(source_dir) if source_dir else get_bundled_template_dir()
    tables: dict[str, pd.DataFrame] = {}
    for name in CONFIG_CSV_FILES:
        src = source / name
        if not src.exists():
            continue
        sheet_name = _sheet_name_from_csv(name)
        tables[sheet_name] = pd.read_csv(src, dtype=str).fillna("")
    return tables


def export_csv_template(source_dir: str | Path, target_dir: str | Path) -> list[Path]:
    """Copy all known CSV template files to ``target_dir``."""

    source = Path(source_dir)
    target = Path(target_dir)
    target.mkdir(parents=True, exist_ok=True)
    copied: list[Path] = []
    for name in CONFIG_CSV_FILES:
        src = source / name
        if src.exists():
            dst = target / name
            shutil.copy2(src, dst)
            copied.append(dst)
    return copied


def write_workbook_from_tables(tables: dict[str, pd.DataFrame], target_path: str | Path) -> Path:
    """Create a fully styled, validated Excel workbook from a dictionary of sheet DataFrames.

    Includes the formatted ReadMe guide, table header styling, auto-width columns,
    and interactive Excel DataValidation dropdowns for all constrained columns.
    """
    try:
        import pandas as pd
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Excel export requires pandas and openpyxl") from exc

    target = Path(target_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    from .types import SUPPORTED_CRC_TYPES, SUPPORTED_FMT_TYPES, ByteOrder, ParserType, ReadWrite

    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        workbook = writer.book
        readme = workbook.create_sheet("ReadMe", 0)
        readme.append(["🐾 Bytehound Configuration Template & User Guide"])
        readme.append([])
        readme.append(["SHEET OVERVIEW:"])
        readme.append(["• Protocol: Defines serial packet framing (header, CRC, footer) or Waveshare CAN adapter mode."])
        readme.append(["• Frames: Defines message frame IDs, friendly frame names, expected payload lengths, and direction (rx/tx)."])
        readme.append(["• Variables: Defines signal decoding rules, offsets, data types, scales, units, and array expansion."])
        readme.append(["• Bitfields: Decodes individual status/fault bits from integer variables."])
        readme.append(["• Enums: Maps integer states to human-readable strings (e.g. 0=IDLE, 1=CHARGING, 2=FAULT)."])
        readme.append(["• CalcGroups: Computes real-time math metrics (min, max, delta, avg) across grouped signals (e.g. Cell Voltages)."])
        readme.append(["• TxCommands: Configures transmit commands that can be sent interactively from the UI."])
        readme.append(["• TxCommandFields: Dynamic parameters and input boxes attached to transmit commands."])
        readme.append(["• SerialDefaults: Default connection parameters (Baud rate, data bits, stop bits, parity)."])
        readme.append(["• PollingSchedule: Periodic command transmission intervals for query-response protocols."])
        readme.append([])
        readme.append(["COMMON PROTOCOL CONFIGURATIONS:"])
        readme.append(["1. Generic Microcontroller Serial Framing (STM32, NXP, ESP32, Arduino):"])
        readme.append(["   - In Protocol sheet: Set parser_type = 'framed', specify header_hex (e.g. AA 55), frame_id_size (e.g. 2), crc_type (e.g. crc16_modbus)."])
        readme.append(["2. Waveshare USB-CAN Adapters (USB-CAN-A / USB-CAN-B):"])
        readme.append(["   - Variable-Length CAN Mode: Set parser_type = 'waveshare_can_variable_length' (Header 0xAA, DLC, CAN ID, payload, footer 0x55)."])
        readme.append(["   - Fixed 20-Byte CAN Mode: Set parser_type = 'waveshare_can_20_bytes' or 'fixed_20_bytes' (Fixed 20-byte frames starting with 0xAA 0x55)."])
        readme.append(["   - In Variables sheet: Put CAN IDs in id_or_address (e.g. 0x0123 for standard 11-bit ID or 0x18FF50E5 for extended 29-bit ID)."])
        readme.append([])
        readme.append(["ARRAY EXPANSION (count & start_index):"])
        readme.append(["• If a signal is an array (e.g. 16 cell voltages), set count = 16. Bytehound will expand it into 16 signals."])
        readme.append(["• start_index controls the starting index number (default = 1: Signal_1..Signal_16; or 0: Signal_0..Signal_15)."])
        readme.append([])
        readme.append(["TIPS:"])
        readme.append(["• Hexadecimal numbers can be written as 0x1000 or 1000."])
        readme.append(["• Click dropdown arrows in constrained cells to pick valid options."])
        readme["A1"].font = Font(bold=True, size=14, color="1E88E5")
        for r_idx in (3, 14, 21, 26):
            readme[f"A{r_idx}"].font = Font(bold=True)
        readme.column_dimensions["A"].width = 110

        for sheet_name, df in tables.items():
            if df is None:
                continue
            df_clean = df.fillna("")
            df_clean.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for col_num, col_name in enumerate(df_clean.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = header_fill

                max_len = max((len(str(v)) for v in df_clean[col_name]), default=0)
                max_len = max(max_len, len(str(col_name)))
                ws.column_dimensions[get_column_letter(col_num)].width = min(max_len + 2, 50)

                col_letter = get_column_letter(col_num)
                val_range = f"{col_letter}2:{col_letter}1048576"
                dv = None
                if col_name == "parser_type":
                    dv = DataValidation(type="list", formula1=f'"{",".join([m.value for m in ParserType])}"', allow_blank=True)
                elif col_name == "data_type":
                    dv = DataValidation(type="list", formula1=f'"{",".join(sorted(SUPPORTED_FMT_TYPES))}"', allow_blank=True)
                elif col_name == "crc_type":
                    dv = DataValidation(type="list", formula1=f'"{",".join(sorted(SUPPORTED_CRC_TYPES))}"', allow_blank=True)
                elif col_name in ("endianness", "byte_order", "frame_id_byte_order", "crc_byte_order", "length_byte_order"):
                    dv = DataValidation(type="list", formula1=f'"{",".join([m.value for m in ByteOrder])}"', allow_blank=True)
                elif col_name == "read_write":
                    dv = DataValidation(type="list", formula1=f'"{",".join([m.value for m in ReadWrite])}"', allow_blank=True)
                elif col_name == "direction":
                    dv = DataValidation(type="list", formula1='"rx,tx,rxtx"', allow_blank=True)
                elif col_name in ("enabled", "waveshare_fixed_20_bytes", "waveshare_fixed"):
                    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
                elif col_name == "length_meaning":
                    dv = DataValidation(type="list", formula1='"payload_only,frame_total,header_to_crc,payload_plus_crc"', allow_blank=True)
                elif col_name == "crc_coverage":
                    dv = DataValidation(type="list", formula1='"header_to_payload,frame_id_to_payload,payload_only,full_frame"', allow_blank=True)
                elif col_name == "escape_mode":
                    dv = DataValidation(type="list", formula1='"none,slip,hdlc,cobs"', allow_blank=True)
                elif col_name == "raw_log_format":
                    dv = DataValidation(type="list", formula1='"hex,compact"', allow_blank=True)
                elif col_name == "operations":
                    dv = DataValidation(type="list", formula1='"min,max,sum,diff,avg,min|max,min|max|avg,min|max|diff|avg,sum|avg"', allow_blank=True)
                elif col_name == "frame_id_size":
                    dv = DataValidation(type="list", formula1='"1,2,4"', allow_blank=True)
                elif col_name == "length_size":
                    dv = DataValidation(type="list", formula1='"0,1,2,4"', allow_blank=True)
                elif col_name == "crc_size":
                    dv = DataValidation(type="list", formula1='"0,2,4"', allow_blank=True)
                elif col_name == "parity":
                    dv = DataValidation(type="list", formula1='"N,E,O,M,S"', allow_blank=True)
                elif col_name == "baud_rate":
                    dv = DataValidation(type="list", formula1='"9600,19200,38400,57600,115200,230400,921600,2000000"', allow_blank=True)

                if dv:
                    dv.add(val_range)
                    ws.add_data_validation(dv)
    return target


def export_excel_template(source_dir: str | Path, target_path: str | Path) -> Path:
    """Create an Excel workbook from the known CSV template files.

    If ``source_dir`` is already an Excel workbook (``.xlsx``/``.xlsm``), it is
    copied to ``target_path``.
    """

    source = Path(source_dir)
    target = Path(target_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    if source.is_file() and source.suffix.lower() in {".xlsx", ".xlsm"}:
        shutil.copy2(source, target)
        return target
    if source.is_file():
        raise ValueError(f"Excel export source must be a directory or .xlsx file: {source}")

    tables = load_template_tables(source)
    return write_workbook_from_tables(tables, target)


def snapshot_config(source_path: str | Path, target_dir: str | Path) -> Path:
    """Copy the active config beside a log session."""

    source = Path(source_path)
    target = Path(target_dir)
    target.mkdir(parents=True, exist_ok=True)
    if source.is_dir():
        snapshot_dir = target / "config_snapshot"
        export_csv_template(source, snapshot_dir)
        return snapshot_dir
    snapshot_file = target / f"config_snapshot{source.suffix}"
    shutil.copy2(source, snapshot_file)
    return snapshot_file


def _sheet_name_from_csv(name: str) -> str:
    words = Path(name).stem.split("_")
    return "".join(word.capitalize() for word in words)[:31]
