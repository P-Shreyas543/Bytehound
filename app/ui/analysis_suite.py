"""Analysis Suite — Multi-log comparison tool with overlay plots.

Launched from the Data menu as a non-modal QMainWindow.  All Excel loading
runs in a QThread so the live test is never blocked.  Plots use OpenGL
acceleration via pyqtgraph for lag-free pan/zoom even with many data points.
"""
import csv
import json
import logging
import os
import re
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# Module-level logger. Routing to stderr / files is left to the host
# application's logging config — we only emit. The existing _log_activity
# calls into the main window's activity log remain in place; this is an
# additional channel so terminal users and crash reports see context.
_log = logging.getLogger("bytehound.analysis_suite")

import datetime as _dt

import numpy as np
import pyqtgraph as pg
from openpyxl import load_workbook

# ── Global pyqtgraph config (must run before any PlotWidget is created) ──
pg.setConfigOptions(antialias=True, useOpenGL=True)
from PySide6.QtCore import QThread, Qt, Signal, QPointF, QObject, QSettings, QTimer
from PySide6.QtGui import (
    QAction, QColor, QFont, QKeySequence, QPainter, QPen,
)
from PySide6.QtWidgets import (
    QAbstractItemView, QApplication, QCheckBox, QColorDialog, QComboBox,
    QDialog, QDialogButtonBox, QDoubleSpinBox, QFileDialog, QFrame,
    QGridLayout, QGroupBox, QHBoxLayout, QHeaderView, QLabel, QLineEdit,
    QListWidget, QListWidgetItem, QMainWindow, QMenu, QMessageBox,
    QPushButton, QScrollArea, QSizePolicy, QSlider, QSpinBox, QSplitter,
    QStatusBar, QTableWidget, QTableWidgetItem, QTabWidget, QToolBar,
    QTreeWidget, QTreeWidgetItem, QVBoxLayout, QWidget,
)

# ═══════════════════════════════════════════════════════════════════════
# Stubs to replace missing dependencies
# ═══════════════════════════════════════════════════════════════════════
APP_NAME = "Bytehound"
APP_ORG = "Bytehound"

def get_datalogs_dir() -> str:
    path = Path.home() / "Documents" / APP_NAME / "Logs"
    path.mkdir(parents=True, exist_ok=True)
    return str(path)

def get_analysis_dir() -> str:
    path = Path.home() / "Documents" / APP_NAME / "Analysis"
    path.mkdir(parents=True, exist_ok=True)
    return str(path)

# Theme-aware color provider so plots/cursors follow the app's dark/light
# setting. Mirrors _apply_plot_theme() in main_window.py — same Slate (#1E293B)
# for dark plot bg and #FFFFFF for light, so the Analysis Suite plots match
# the Live Plot canvas exactly.
_DARK_PLOT_COLORS = {
    'plot_bg': '#1E293B',
    'plot_fg': '#CBD5E1',
    'crosshair': '#94A3B8',
    'cursor_label_bg': '#0F172A',
    'border': '#334155',
    'legend_bg': (0, 0, 0, 100),
}
_LIGHT_PLOT_COLORS = {
    'plot_bg': '#FFFFFF',
    'plot_fg': '#475569',
    'crosshair': '#64748B',
    'cursor_label_bg': '#F3F4F6',
    'border': '#E5E7EB',
    'legend_bg': (255, 255, 255, 160),
}

class _AppTheme(QObject):
    theme_changed = Signal(str)

    def theme(self) -> str:
        return str(QSettings(APP_ORG, APP_NAME).value("ui/theme", "dark"))

    def c(self, key: str) -> str:
        palette = _LIGHT_PLOT_COLORS if self.theme() == "light" else _DARK_PLOT_COLORS
        return palette.get(key, '#FFFFFF')

    def plot_grid_alpha(self) -> float:
        return 0.15

THEME = _AppTheme()

# ═══════════════════════════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════════════════════════
LOG_COLORS = [
    '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
    '#8c564b', '#e377c2', '#7f7f7f', '#17becf', '#bcbd22',
    '#ff9896', '#98df8a', '#c5b0d5', '#393b79',
]

CURSOR_COLORS = ['#e63946', '#457b9d', '#2a9d8f', '#f4a261', '#6a4c93']
SELECTED_CURSOR_COLOR = '#ff0000'

DEFAULT_PARAMS = [
    "Vehicle Speed (Kmph)", "Dyno Act Torque (Nm)",
    "Vehicle Power (W)", "Roller Speed (RPM)",
]

SESSION_VERSION = 4
MIN_PLOT_HEIGHT = 80
MAX_PLOT_HEIGHT = 600
DEFAULT_PLOT_HEIGHT = 200


# ─────────────────────────────────────────────────────────────────────
# Custom time-axis — renders ticks as mm:ss or H:MM:SS
# ─────────────────────────────────────────────────────────────────────
class TimeAxisItem(pg.AxisItem):
    """X-axis that renders elapsed seconds as mm:ss or H:MM:SS.

    Set *epoch_offset* to a POSIX timestamp to switch to wall-clock
    mode (HH:MM:SS).  Reset to ``None`` to return to elapsed mode.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._epoch_offset: float | None = None

    @property
    def epoch_offset(self) -> float | None:
        return self._epoch_offset

    @epoch_offset.setter
    def epoch_offset(self, value: float | None):
        self._epoch_offset = value
        # Force label refresh
        self.picture = None
        self.update()

    # -- formatting helpers --------------------------------------------------
    @staticmethod
    def _fmt_elapsed(seconds: float) -> str:
        """Format elapsed seconds as mm:ss or H:MM:SS."""
        neg = seconds < 0
        s = abs(seconds)
        h = int(s // 3600)
        m = int((s % 3600) // 60)
        sec = int(s % 60)
        if h > 0:
            txt = f"{h}:{m:02d}:{sec:02d}"
        else:
            txt = f"{m}:{sec:02d}"
        return f"-{txt}" if neg else txt

    @staticmethod
    def _fmt_wallclock(posix: float) -> str:
        """Format a POSIX timestamp as HH:MM:SS."""
        try:
            dt = _dt.datetime.fromtimestamp(posix)
            return dt.strftime("%H:%M:%S")
        except (OSError, OverflowError, ValueError):
            return "?"

    def tickStrings(self, values, scale, spacing):
        if self._epoch_offset is not None:
            return [self._fmt_wallclock(v + self._epoch_offset) for v in values]
        return [self._fmt_elapsed(v) for v in values]

# ─────────────────────────────────────────────────────────────────────
# Multi-param subplot visual encoding
# ─────────────────────────────────────────────────────────────────────
# When several parameters share one subplot we need TWO independent visual
# channels — one for the parameter and one for the log. Lightness shifts of
# the same hue are not enough (a "lighter blue" vs "darker blue" still reads
# as blue in a small legend swatch), so we use distinct channels:
#
#     color       = parameter slot   (matplotlib tab10 — high-contrast hues)
#     line style  = log slot         (solid / dashed / dotted / dash-dot …)
#
# In a single-parameter subplot we revert to the original "color = log"
# convention so the curve color still matches the log swatch in the sidebar.
_PARAM_COLORS = [
    '#1f77b4',  # blue
    '#ff7f0e',  # orange
    '#2ca02c',  # green
    '#d62728',  # red
    '#9467bd',  # purple
    '#8c564b',  # brown
    '#e377c2',  # pink
    '#17becf',  # cyan
    '#bcbd22',  # olive
    '#7f7f7f',  # gray
]

# Line styles cycled per log within a multi-param subplot.
_LOG_LINE_STYLES = [Qt.SolidLine, Qt.DashLine, Qt.DotLine, Qt.DashDotLine, Qt.DashDotDotLine]


def _curve_visuals(group: list[str], param: str, log_index: int, log_color: str
                    ) -> tuple[str, Qt.PenStyle]:
    """Return (color, line_style) for a single (param, log) curve.

    Single-param subplot  → (log_color, solid).
    Multi-param subplot   → (param-slot color, log-slot line style).
    """
    if len(group) <= 1:
        return log_color, Qt.SolidLine
    try:
        slot = group.index(param)
    except ValueError:
        slot = 0
    color = _PARAM_COLORS[slot % len(_PARAM_COLORS)]
    style = _LOG_LINE_STYLES[log_index % len(_LOG_LINE_STYLES)]
    return color, style


# ═══════════════════════════════════════════════════════════════════════
# Data model
# ═══════════════════════════════════════════════════════════════════════
@dataclass
class LogEntry:
    """One loaded test-log file."""
    id: str = ""
    path: str = ""
    name: str = ""
    color: str = "#1f77b4"
    visible: bool = True
    time_offset: float = 0.0
    start_timestamp: Optional[float] = None  # POSIX epoch of first sample
    elapsed: np.ndarray = field(default_factory=lambda: np.zeros(0))
    columns: dict[str, np.ndarray] = field(default_factory=dict)

    def available_params(self) -> list[str]:
        return [name for name in self.columns.keys() if not _is_time_like_param(name)]


def _is_time_like_param(name: str) -> bool:
    """Return True for columns that represent time axes rather than data."""
    norm = re.sub(r'\s+', ' ', (name or '').strip().lower())
    if norm in {
        '', 'time', 'timestamp', 'elapsed', 'elapsed (s)', 'time (s)',
        'elapsed time', 'elapsed time (s)',
    }:
        return True
    return norm.startswith('timestamp ') or norm.startswith('elapsed ')


def _test_name_from_path(path: str) -> str:
    """Extract a human-readable test name from the log filename."""
    base = os.path.splitext(os.path.basename(path))[0]
    # Strip DynoLog_ prefix and _YYYYMMDD_HHMMSS suffix
    base = re.sub(r'^DynoLog_', '', base)
    base = re.sub(r'_\d{8}_\d{6}$', '', base)
    return base or os.path.basename(path)


def _parse_unit(param_name: str) -> str | None:
    """Extract unit string from a column header.

    Recognises both parenthesised and bracketed conventions:
        ``Voltage [V]``  → ``V``
        ``Speed (Kmph)`` → ``Kmph``
        ``Temperature``  → None
    """
    m = re.search(r'[\[\(]([^\]\)]+)[\]\)]\s*$', param_name)
    return m.group(1).strip() if m else None


# File-path → LogEntry cache so re-toggling a parameter does not re-read disk.
_CSV_CACHE: dict[str, LogEntry] = {}


# ═══════════════════════════════════════════════════════════════════════
# Background loader thread
# ═══════════════════════════════════════════════════════════════════════
class LogLoaderThread(QThread):
    """Loads one .xlsx (Bytehound decoded or Dyno) or _decoded.csv log in a
    background thread."""
    log_loaded = Signal(str, str, object)   # (log_id, path, LogEntry or None)
    error = Signal(str, str)                # (path, error_message)

    def __init__(self, path: str, log_id: str, color: str, parent=None):
        super().__init__(parent)
        self._path = path
        self._log_id = log_id
        self._color = color

    def run(self):
        try:
            ext = os.path.splitext(self._path)[1].lower()
            if ext == ".csv":
                self._load_csv()
            else:
                self._load_xlsx()
        except Exception as exc:
            self.error.emit(self._path, str(exc))

    def _load_csv(self):
        """Parse a legacy _decoded.csv file (pre-xlsx Bytehound versions)."""
        with open(self._path, newline='', encoding='utf-8') as f:
            def _iter_rows():
                for line in f:
                    stripped = line.strip()
                    if not stripped or stripped.startswith("#"):
                        continue
                    yield line
            reader = csv.DictReader(_iter_rows())
            fieldnames = reader.fieldnames or []
            rows = list(reader)

        if not fieldnames:
            self.error.emit(self._path, "CSV header row is missing.")
            return
        if not rows:
            self.error.emit(self._path, "No data rows found in CSV.")
            return

        data_columns = [
            name for name in fieldnames
            if name and name not in {"timestamp", "elapsed_ms"}
        ]
        if not data_columns:
            self.error.emit(self._path, "No data columns found in CSV.")
            return

        import datetime as dt
        first_ts = None
        first_ts_posix: float | None = None  # for start_timestamp
        col_data: dict[str, list[tuple[float, float]]] = {name: [] for name in data_columns}

        for row in rows:
            elapsed_s: float | None = None
            elapsed_raw = row.get("elapsed_ms", "")
            if elapsed_raw is not None and str(elapsed_raw).strip() != "":
                try:
                    elapsed_s = float(elapsed_raw) / 1000.0
                except ValueError:
                    elapsed_s = None

            if elapsed_s is None:
                ts_str = row.get("timestamp", "")
                try:
                    ts = dt.datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S.%f")
                except ValueError:
                    try:
                        ts = dt.datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        ts = None
                if first_ts is None and ts is not None:
                    first_ts = ts
                    try:
                        first_ts_posix = first_ts.timestamp()
                    except Exception:
                        pass
                elapsed_s = (ts - first_ts).total_seconds() if (ts and first_ts) else 0.0

            for label in data_columns:
                cell = row.get(label, "")
                if cell is None or str(cell).strip() == "":
                    continue
                try:
                    value = float(cell)
                except (ValueError, TypeError):
                    continue
                col_data[label].append((elapsed_s, value))

        col_data = {label: pairs for label, pairs in col_data.items() if pairs}
        if not col_data:
            self.error.emit(self._path, "No numeric columns found in CSV.")
            return

        # Build unified time axis (union of all timestamps, sorted)
        all_times = sorted({t for times in col_data.values() for t, _ in times})
        n = len(all_times)
        time_idx = {t: i for i, t in enumerate(all_times)}
        elapsed_arr = np.array(all_times)

        columns: dict[str, np.ndarray] = {}
        for label, pairs in col_data.items():
            arr = np.full(n, np.nan)
            for t, v in pairs:
                arr[time_idx[t]] = v
            columns[label] = arr

        entry = LogEntry(
            id=self._log_id,
            path=self._path,
            name=_test_name_from_path(self._path),
            color=self._color,
            start_timestamp=first_ts_posix,
            elapsed=elapsed_arr,
            columns=columns,
        )
        _CSV_CACHE[self._path] = entry
        self.log_loaded.emit(self._log_id, self._path, entry)

    def _load_xlsx(self):
        """Parse a .xlsx file. Supports three schemas:

        * **Bytehound decoded (cycle-buffered)** — sheet ``Data`` with
          per-frame blocks like ``<frame>.elapsed_ms``, ``<frame>.frame_id``,
          ``<frame>.<signal>``. The trigger frame's elapsed_ms (the LAST
          ``*.elapsed_ms`` column) is used as the time axis.
        * **Bytehound decoded (legacy flat)** — sheet ``Data`` with a single
          ``elapsed_ms`` column.
        * **Dyno test rig** — sheet ``Record`` (or active), column
          ``Elapsed (s)`` then signals.
        """
        from openpyxl import load_workbook
        wb = load_workbook(self._path, read_only=True, data_only=True)
        if 'Data' in wb.sheetnames:
            ws = wb['Data']
        elif 'Record' in wb.sheetnames:
            ws = wb['Record']
        else:
            ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            self.error.emit(self._path, "No data rows found.")
            return

        headers = [str(c) if c else "" for c in rows[0]]
        data_rows = rows[1:]

        # Pick the elapsed-time column.
        #   Dyno      → "Elapsed (s)"             (seconds)
        #   Bytehound → "<frame>.elapsed_ms"      (ms, prefer the trigger
        #                                          frame = LAST occurrence)
        #              or legacy flat "elapsed_ms"
        elapsed_idx = -1
        elapsed_scale = 1.0
        frame_block_elapsed_indices: set[int] = set()
        frame_block_id_indices: set[int] = set()
        for i, h in enumerate(headers):
            if h.endswith(".elapsed_ms"):
                frame_block_elapsed_indices.add(i)
                elapsed_idx = i  # keep updating → LAST one wins (trigger frame)
                elapsed_scale = 1e-3
            elif h.endswith(".frame_id"):
                frame_block_id_indices.add(i)
        if elapsed_idx < 0:
            for i, h in enumerate(headers):
                if h == "Elapsed (s)":
                    elapsed_idx = i
                    elapsed_scale = 1.0
                    break
                if h == "elapsed_ms":
                    elapsed_idx = i
                    elapsed_scale = 1e-3
                    break

        n = len(data_rows)
        elapsed = np.zeros(n)
        columns: dict[str, np.ndarray] = {}

        skip_indices = frame_block_elapsed_indices | frame_block_id_indices | {elapsed_idx}
        col_map: dict[int, str] = {}
        for i, h in enumerate(headers):
            if i in skip_indices:
                continue
            if not _is_time_like_param(h):
                col_map[i] = h
                columns[h] = np.full(n, np.nan)

        for r, row in enumerate(data_rows):
            if elapsed_idx >= 0 and elapsed_idx < len(row):
                try:
                    elapsed[r] = float(row[elapsed_idx]) * elapsed_scale
                except (ValueError, TypeError):
                    elapsed[r] = elapsed[r - 1] if r > 0 else 0.0
            for ci, param in col_map.items():
                if ci < len(row) and row[ci] is not None:
                    try:
                        columns[param][r] = float(row[ci])
                    except (ValueError, TypeError):
                        pass

        entry = LogEntry(
            id=self._log_id,
            path=self._path,
            name=_test_name_from_path(self._path),
            color=self._color,
            elapsed=elapsed,
            columns=columns,
        )
        _CSV_CACHE[self._path] = entry
        self.log_loaded.emit(self._log_id, self._path, entry)



# ═══════════════════════════════════════════════════════════════════════
# Cursor readout widget
# ═══════════════════════════════════════════════════════════════════════
class CursorReadoutPanel(QGroupBox):
    """Displays interpolated values at vertical cursor positions.

    Features:
    - Monospace numbers, right-aligned for scan-ability.
    - Colored swatch ● next to each log name.
    - ΔX / ΔY readouts when two cursors are active.
    """

    _MONO = QFont("Consolas", 9)
    _MONO_BOLD = QFont("Consolas", 9, QFont.Bold)
    _LABEL_FONT = QFont("PT Sans", 8, QFont.Bold)

    def __init__(self, parent=None):
        super().__init__("Cursor Readout", parent)
        self.setMinimumWidth(200)

        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setFrameShape(QFrame.NoFrame)

        self._inner = QWidget()
        self._layout = QVBoxLayout(self._inner)
        self._layout.setContentsMargins(4, 4, 4, 4)
        self._layout.setSpacing(1)
        self._scroll.setWidget(self._inner)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 12, 0, 0)
        outer.addWidget(self._scroll)

        self._delta_label = QLabel("")
        self._delta_label.setFont(self._MONO_BOLD)
        self._layout.addWidget(self._delta_label)
        self._info_label = QLabel("Add cursors with V / Shift+V / H keys.")
        self._info_label.setFont(QFont("PT Sans", 8))
        self._info_label.setWordWrap(True)
        self._layout.addWidget(self._info_label)
        self._layout.addStretch()

    # -- helpers -------------------------------------------------------------
    @staticmethod
    def _fmt_val(v: float) -> str:
        """Format a value with 3 significant figures."""
        if not np.isfinite(v):
            return "\u2014"
        if v == 0:
            return "0"
        return f"{v:.4g}"

    def _make_row(self, left: str, right: str, *,
                  color: str = "", bold: bool = False) -> QWidget:
        """Build a compact horizontal row: left-aligned label + right-aligned value."""
        row = QWidget()
        hl = QHBoxLayout(row)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.setSpacing(4)
        ll = QLabel(left)
        ll.setFont(self._MONO_BOLD if bold else self._MONO)
        if color:
            ll.setStyleSheet(f"color: {color};")
        rl = QLabel(right)
        rl.setFont(self._MONO)
        rl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        hl.addWidget(ll, 1)
        hl.addWidget(rl)
        return row

    def update_readout(self, cursors: list[dict],
                       logs: list['LogEntry'],
                       active_params: list[str]):
        vbar = self._scroll.verticalScrollBar()
        old_scroll = vbar.value()

        # Clear dynamic labels (keep delta + info + stretch)
        while self._layout.count() > 3:
            item = self._layout.takeAt(2)
            w = item.widget() if item else None
            if w:
                w.deleteLater()

        if not cursors:
            self._delta_label.setText("")
            self._info_label.setText("Add cursors with V / Shift+V / H keys.")
            vbar.setValue(min(old_scroll, vbar.maximum()))
            return

        self._info_label.setText("")

        # ΔX header when two cursors active
        if len(cursors) >= 2:
            dt_val = abs(cursors[1]['time'] - cursors[0]['time'])
            self._delta_label.setText(
                f"\u0394X = {TimeAxisItem._fmt_elapsed(dt_val)}  ({dt_val:.3f} s)")
        else:
            t0 = cursors[0]['time']
            self._delta_label.setText(
                f"t = {TimeAxisItem._fmt_elapsed(t0)}  ({t0:.3f} s)")

        # Per-cursor readout
        # Collect values for ΔY computation (cursor_idx -> {(log_id,param): value})
        cursor_vals: list[dict[tuple[str, str], float]] = []

        for ci, cursor in enumerate(cursors):
            t = cursor['time']
            label_num = cursor.get('label', 0)
            scope = cursor.get('scope', 'all')
            plot_param = cursor.get('plot_param')
            hdr_txt = f"C{label_num}  {TimeAxisItem._fmt_elapsed(t)}"
            if scope == 'plot' and plot_param:
                hdr_txt += f"  [{plot_param}]"
            hdr = QLabel(hdr_txt)
            hdr.setFont(self._LABEL_FONT)
            self._layout.insertWidget(self._layout.count() - 1, hdr)

            if scope == 'plot' and plot_param:
                params_for_cursor = [plot_param]
            else:
                params_for_cursor = active_params

            vals: dict[tuple[str, str], float] = {}
            for log in logs:
                if not log.visible or len(log.elapsed) == 0:
                    continue
                x = log.elapsed + log.time_offset
                # Swatch + log name
                row_w = self._make_row(f"\u25cf {log.name}", "",
                                       color=log.color, bold=True)
                self._layout.insertWidget(self._layout.count() - 1, row_w)
                for param in params_for_cursor:
                    if param not in log.columns:
                        continue
                    idx = int(np.clip(np.searchsorted(x, t), 0, len(x) - 1))
                    v = log.columns[param][idx]
                    short = re.sub(r'\s*[\[\(][^\]\)]*[\]\)]\s*$', '', param).strip()
                    unit = _parse_unit(param) or ""
                    if np.isfinite(v):
                        val_str = f"{self._fmt_val(v)} {unit}".strip()
                        vals[(log.id, param)] = v
                    else:
                        val_str = "\u2014"
                    row_w = self._make_row(f"  {short}", val_str)
                    self._layout.insertWidget(self._layout.count() - 1, row_w)
            cursor_vals.append(vals)

        # ΔY readout between two cursors
        if len(cursor_vals) >= 2:
            sep = QLabel("\u2500\u2500 \u0394Y \u2500\u2500")
            sep.setFont(self._LABEL_FONT)
            self._layout.insertWidget(self._layout.count() - 1, sep)
            keys = set(cursor_vals[0].keys()) & set(cursor_vals[1].keys())
            for key in sorted(keys, key=lambda k: k[1]):
                v0 = cursor_vals[0][key]
                v1 = cursor_vals[1][key]
                delta = v1 - v0
                _, param = key
                short = re.sub(r'\s*[\[\(][^\]\)]*[\]\)]\s*$', '', param).strip()
                unit = _parse_unit(param) or ""
                row_w = self._make_row(
                    f"  {short}",
                    f"\u0394 {self._fmt_val(delta)} {unit}".strip())
                self._layout.insertWidget(self._layout.count() - 1, row_w)

        vbar.setValue(min(old_scroll, vbar.maximum()))


# ═══════════════════════════════════════════════════════════════════════
# X-Y Scatter Plot Window
# ═══════════════════════════════════════════════════════════════════════
# Symbol name → pyqtgraph symbol string
_XY_SYMBOLS = [
    ('Circle (outline)', 'o'),
    ('Circle (filled)', 'o'),
    ('Square', 's'),
    ('Triangle', 't'),
    ('Diamond', 'd'),
    ('Cross', '+'),
]


# ═══════════════════════════════════════════════════════════════════════
# Statistics panel — visible-range descriptive stats per curve
# ═══════════════════════════════════════════════════════════════════════
class StatisticsPanel(QWidget):
    """Per-curve descriptive statistics (min/max/mean/median/std/percentiles)
    computed over the currently visible x-range of each subplot.

    Updates are debounced so panning/zooming doesn't trigger continuous
    recomputation on large logs.
    """

    STATS_COLUMNS = ["Curve", "Min", "P5", "Mean", "Median", "P95", "Max", "Std", "n"]

    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 6, 0, 0)
        layout.setSpacing(4)

        self._info = QLabel("Load logs and check parameters to see statistics.")
        self._info.setWordWrap(True)
        self._info.setStyleSheet("color: palette(mid); font-size: 11px; padding: 2px;")
        layout.addWidget(self._info)

        self._table = QTableWidget(0, len(self.STATS_COLUMNS))
        self._table.setHorizontalHeaderLabels(self.STATS_COLUMNS)
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.setFont(QFont("Consolas", 9))
        hh = self._table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        for col in range(1, len(self.STATS_COLUMNS)):
            hh.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        layout.addWidget(self._table, 1)

    @staticmethod
    def compute_stats(x: np.ndarray, y: np.ndarray,
                       x_range: tuple[float, float] | None
                       ) -> dict[str, float] | None:
        """Return a stats dict for the slice of y where x is within
        ``x_range`` (or the full series if x_range is None). NaNs are
        ignored. Returns None if there are no finite samples in range."""
        if x.size == 0 or y.size == 0:
            return None
        if x_range is not None:
            lo, hi = x_range
            mask = (x >= lo) & (x <= hi)
        else:
            mask = np.ones_like(x, dtype=bool)
        yv = y[mask]
        yv = yv[~np.isnan(yv)]
        if yv.size == 0:
            return None
        return {
            "min":    float(np.min(yv)),
            "p5":     float(np.percentile(yv, 5)),
            "max":    float(np.max(yv)),
            "mean":   float(np.mean(yv)),
            "median": float(np.median(yv)),
            "p95":    float(np.percentile(yv, 95)),
            "std":    float(np.std(yv)),
            "n":      int(yv.size),
        }

    @staticmethod
    def _fmt(v: float) -> str:
        """Format with 3 significant figures."""
        if not np.isfinite(v):
            return "\u2014"
        if v == 0:
            return "0"
        return f"{v:.3g}"

    def update_stats(self, rows: list[dict]):
        """``rows`` is a list of {curve, log_id, param, color, x, y, x_range}.
        We compute stats and re-render the table."""
        self._table.setRowCount(0)
        if not rows:
            self._info.setText("No visible curves. Check parameters to populate.")
            return
        self._info.setText(f"Stats over visible x-range  \u00b7  {len(rows)} curve(s)")
        self._table.setRowCount(len(rows))
        for r, info in enumerate(rows):
            stats = self.compute_stats(info["x"], info["y"], info.get("x_range"))
            label = info["curve"]
            color = info.get("color", "")

            name_item = QTableWidgetItem(label)
            if color:
                name_item.setForeground(QColor(color))
            self._table.setItem(r, 0, name_item)

            if stats is None:
                placeholders = ["\u2014"] * (len(self.STATS_COLUMNS) - 2) + ["0"]
                for c, txt in enumerate(placeholders, start=1):
                    cell = QTableWidgetItem(txt)
                    cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self._table.setItem(r, c, cell)
                continue

            for c, key in enumerate(["min", "p5", "mean", "median", "p95", "max", "std", "n"], start=1):
                if key == "n":
                    txt = str(stats[key])
                else:
                    txt = self._fmt(stats[key])
                cell = QTableWidgetItem(txt)
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self._table.setItem(r, c, cell)


class XYPlotWindow(QDialog):
    """Scatter / X-Y plot window for cross-parameter analysis."""

    def __init__(self, logs: dict[str, LogEntry], parent=None):
        super().__init__(parent)
        # Qt.Window: independent z-order (not an owned always-on-top window)
        self.setWindowFlag(Qt.Window)
        self.setWindowTitle(f"X-Y Plotter — {APP_NAME}")
        self.resize(900, 700)
        self.setMinimumSize(600, 400)
        self.setAttribute(Qt.WA_DeleteOnClose, False)

        self._logs = logs
        self._curves: list = []

        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)

        # ── Row 1: axis selectors ─────────────────────────────────
        ctrl = QHBoxLayout()
        ctrl.addWidget(QLabel("X-Axis:"))
        self._x_combo = QComboBox()
        self._x_combo.setMinimumWidth(160)
        ctrl.addWidget(self._x_combo)
        ctrl.addWidget(QLabel("Y-Axis:"))
        self._y_combo = QComboBox()
        self._y_combo.setMinimumWidth(160)
        ctrl.addWidget(self._y_combo)
        ctrl.addSpacing(12)

        # ── Symbol style ─────────────────────────────────────────
        ctrl.addWidget(QLabel("Symbol:"))
        self._sym_combo = QComboBox()
        for name, _ in _XY_SYMBOLS:
            self._sym_combo.addItem(name)
        self._sym_combo.setCurrentIndex(1)  # filled circle default
        ctrl.addWidget(self._sym_combo)
        ctrl.addSpacing(8)

        # ── Symbol size ──────────────────────────────────────────
        ctrl.addWidget(QLabel("Size:"))
        self._size_spin = QSpinBox()
        self._size_spin.setRange(2, 30)
        self._size_spin.setValue(6)
        self._size_spin.setFixedWidth(55)
        ctrl.addWidget(self._size_spin)
        ctrl.addSpacing(12)

        # ── Regression checkbox ──────────────────────────────────
        self._regress_cb = QCheckBox("Regression")
        self._regress_cb.setToolTip("Show linear regression line with R²")
        ctrl.addWidget(self._regress_cb)
        ctrl.addSpacing(8)

        btn_plot = QPushButton("Plot")
        btn_plot.clicked.connect(self._do_plot)
        ctrl.addWidget(btn_plot)

        btn_swap = QPushButton("⇄ Swap")
        btn_swap.setToolTip("Swap X and Y axis parameters")
        btn_swap.clicked.connect(self._swap_axes)
        ctrl.addWidget(btn_swap)

        btn_clear = QPushButton("Clear")
        btn_clear.clicked.connect(self._clear_plot)
        ctrl.addWidget(btn_clear)
        ctrl.addStretch()
        layout.addLayout(ctrl)

        # ── Populate combos ───────────────────────────────────────
        all_params: list[str] = []
        seen = set()
        for entry in self._logs.values():
            for p in entry.available_params():
                if p not in seen:
                    seen.add(p)
                    all_params.append(p)
        self._x_combo.addItems(all_params)
        self._y_combo.addItems(all_params)
        if len(all_params) >= 2:
            self._x_combo.setCurrentIndex(0)
            self._y_combo.setCurrentIndex(1)

        # ── Plot widget ───────────────────────────────────────────
        self._plot = pg.PlotWidget()
        bg = THEME.c('plot_bg')
        fg = THEME.c('plot_fg')
        self._plot.setBackground(bg)
        self._plot.showGrid(x=True, y=True, alpha=THEME.plot_grid_alpha())
        for ax_name in ('left', 'bottom'):
            ax = self._plot.getAxis(ax_name)
            pen = pg.mkPen(fg)
            ax.setPen(pen)
            ax.setTextPen(pen)
        self._legend = self._plot.addLegend(offset=(10, 10))
        layout.addWidget(self._plot)

    def _do_plot(self):
        x_param = self._x_combo.currentText()
        y_param = self._y_combo.currentText()
        if not x_param or not y_param:
            return
        self._plot.setLabel('bottom', x_param)
        self._plot.setLabel('left', y_param)

        sym_idx = self._sym_combo.currentIndex()
        sym_str = _XY_SYMBOLS[sym_idx][1]
        size = self._size_spin.value()
        # Outline vs filled: outline uses mkPen with color, filled uses no pen
        filled = (self._sym_combo.currentText().lower().find('filled') >= 0
                  or self._sym_combo.currentIndex() == 1)

        for entry in self._logs.values():
            if not entry.visible:
                continue
            if x_param not in entry.columns or y_param not in entry.columns:
                continue
            x = entry.columns[x_param]
            y = entry.columns[y_param]
            mask = ~(np.isnan(x) | np.isnan(y))
            pen = pg.mkPen(None) if filled else pg.mkPen(entry.color, width=1)
            brush = pg.mkBrush(entry.color) if filled else pg.mkBrush(None)
            scatter = pg.ScatterPlotItem(
                x=x[mask], y=y[mask],
                pen=pen, brush=brush,
                symbol=sym_str, size=size,
                name=entry.name)
            self._plot.addItem(scatter)
            self._curves.append(scatter)

    def _clear_plot(self):
        for c in self._curves:
            self._plot.removeItem(c)
        self._curves.clear()


# ═══════════════════════════════════════════════════════════════════════
# Main analysis window
# ═══════════════════════════════════════════════════════════════════════
class AnalysisSuiteWindow(QMainWindow):
    """Multi-log comparison analysis tool."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlag(Qt.Window)
        self.setWindowTitle(f"Analysis Suite — {APP_NAME}")
        self.resize(1500, 900)
        self.setMinimumSize(900, 600)

        self._logs: dict[str, LogEntry] = {}
        self._color_index = 0
        self._loader_threads: list[LogLoaderThread] = []

        # Plot management.
        # _plot_groups[i] holds the list of parameter names rendered on
        # _plot_widgets[i]. A single-param subplot has a list of length 1.
        self._plot_widgets: list[pg.PlotWidget] = []
        self._plot_groups: list[list[str]] = []
        # Persists the user-defined subplot layout (list of param lists) even
        # when params are unchecked or new logs are loaded.
        self._subplot_layout: list[list[str]] = []
        # Per-subplot toggles keyed by ``frozenset(params)`` so they survive
        # subplot reorder, splitting, and merging.
        self._normalized_subplots: set[frozenset[str]] = set()
        # Per-subplot smoothing — value is a moving-average window size; missing
        # entries mean "raw, no smoothing".
        self._smoothed_subplots: set[frozenset[str]] = set()
        self._smoothing_window: int = 5
        # Auto-fit Y when the x-range changes (toolbar toggle).
        self._auto_fit_y: bool = False
        # Recent files (Phase 2) — populated lazily from QSettings.
        self._recent_files: list[str] = []
        # Settings (window geometry, recent files, theme prefs, etc.)
        self._qsettings = QSettings(APP_ORG, APP_NAME)
        # Guard: True while we're programmatically populating the tree so the
        # model's rowsInserted signal (used to detect drag-drop) doesn't
        # mistake our own inserts for user drops.
        self._populating_tree: bool = False
        self._curves: dict[str, dict[str, pg.PlotDataItem]] = {}
        self._v_cursors: list[dict] = []
        self._h_cursors: list[dict] = []
        self._crosshair_lines: dict[pg.PlotWidget, tuple] = {}
        self._selected_v_cursor: str = ''   # cursor ID string, '' = none
        self._selected_h_cursor: int = -1
        self._cursor_dots: dict[str, list] = {}   # cursor_id → [{'pw', 'item'}]
        self._v_cursor_counter: int = 0          # ever-increasing label counter
        self._xy_window = None                   # keep reference to non-modal XY window
        self._plot_height = int(
            self._qsettings.value("analysis/plot_height", DEFAULT_PLOT_HEIGHT))
        self._last_mouse_pw: Optional[pg.PlotWidget] = None
        self._last_mouse_pos: Optional[QPointF] = None
        self._wall_clock_mode: bool = False       # X-axis: elapsed vs wall-clock
        self._persisted_x_range = self._qsettings.value("analysis/x_range")
        self._persisted_cursors = self._qsettings.value(
            "analysis/cursor_positions", [])

        # Log sidebar UI elements
        self._log_entries_ui: dict[str, dict] = {}  # log_id → {checkbox, spin, container, ...}

        self._build_ui()
        self._build_menus()
        # Apply the shared app QSS (cards, docks, tables, tabs) so this window
        # visually matches the MainWindow. qdarktheme's palette is already
        # applied app-wide, but the custom QSS is per-top-level-window.
        self.apply_theme(str(QSettings(APP_ORG, APP_NAME).value("ui/theme", "dark")))
        THEME.theme_changed.connect(self._apply_theme)

    def _log_activity(self, text: str) -> None:
        """Forward Activity Log events to the main app window if available."""
        parent = self.parent()
        # Walk up the Qt parent chain looking for MainWindow._log_activity.
        while parent is not None:
            if hasattr(parent, "_log_activity"):
                try:
                    parent._log_activity(text)  # type: ignore[attr-defined]
                except Exception:
                    pass
                return
            parent = parent.parent() if hasattr(parent, "parent") else None

    def _log_popup(self, kind: str, title: str, message: str) -> None:
        message_text = "" if message is None else str(message)
        lines = message_text.splitlines()
        tag = f"Analysis Suite/{kind}"
        if not lines:
            self._log_activity(f"[{tag}] {title}")
            return
        if len(lines) == 1:
            self._log_activity(f"[{tag}] {title}: {lines[0]}")
            return
        self._log_activity(f"[{tag}] {title}:")
        for line in lines:
            self._log_activity(f"    {line}")

    def _popup_information(self, title: str, message: str) -> None:
        self._log_popup("INFO", title, message)
        _log.info("[%s] %s", title, message.replace("\n", " · "))
        QMessageBox.information(self, title, message)

    def _popup_warning(self, title: str, message: str) -> None:
        self._log_popup("WARN", title, message)
        _log.warning("[%s] %s", title, message.replace("\n", " · "))
        QMessageBox.warning(self, title, message)

    # ──────────────────────────────────────────────────────────────────
    # UI construction
    # ──────────────────────────────────────────────────────────────────
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QHBoxLayout(central)
        root_layout.setContentsMargins(4, 4, 4, 4)

        # ── Left sidebar ─────────────────────────────────────────────
        sidebar = QWidget()
        sidebar.setMaximumWidth(320)
        sidebar.setMinimumWidth(240)
        side_layout = QVBoxLayout(sidebar)
        side_layout.setContentsMargins(4, 4, 4, 4)
        side_layout.setSpacing(4)

        # Loaded logs group (with inline time offsets)
        log_group = QGroupBox("Loaded Logs")
        log_layout = QVBoxLayout(log_group)
        log_layout.setContentsMargins(4, 12, 4, 4)
        log_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)

        btn_row = QHBoxLayout()
        self._btn_load = QPushButton("Load Logs")
        self._btn_load.clicked.connect(self._on_load_logs)
        btn_row.addWidget(self._btn_load)
        self._btn_remove = QPushButton("Remove")
        self._btn_remove.clicked.connect(self._on_remove_selected)
        btn_row.addWidget(self._btn_remove)
        btn_row.addStretch(1)
        log_layout.addLayout(btn_row)

        # Log list — each row: checkbox + color swatch + name + offset spin
        self._log_list_widget = QWidget()
        self._log_list_layout = QVBoxLayout(self._log_list_widget)
        self._log_list_layout.setContentsMargins(0, 0, 0, 0)
        self._log_list_layout.setSpacing(2)
        self._log_list_layout.addStretch()

        log_scroll = QScrollArea()
        log_scroll.setWidgetResizable(True)
        log_scroll.setFrameShape(QFrame.NoFrame)
        log_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        log_scroll.setWidget(self._log_list_widget)
        log_layout.addWidget(log_scroll, 1)
        side_layout.addWidget(log_group, 1)

        # Plot height slider
        height_row = QHBoxLayout()
        height_row.addWidget(QLabel("Plot Height:"))
        self._height_slider = QSlider(Qt.Horizontal)
        self._height_slider.setRange(MIN_PLOT_HEIGHT, MAX_PLOT_HEIGHT)
        self._height_slider.setValue(self._plot_height)
        self._height_slider.valueChanged.connect(self._on_plot_height_changed)
        self._height_slider.mouseDoubleClickEvent = lambda e: (
            self._height_slider.setValue(DEFAULT_PLOT_HEIGHT))
        self._height_slider.setToolTip(
            f"Drag to resize subplots ({MIN_PLOT_HEIGHT}–{MAX_PLOT_HEIGHT}px).\n"
            f"Double-click to reset to {DEFAULT_PLOT_HEIGHT}px.")
        height_row.addWidget(self._height_slider)
        self._height_label = QLabel(f"{DEFAULT_PLOT_HEIGHT}px")
        self._height_label.setMinimumWidth(40)
        height_row.addWidget(self._height_label)
        side_layout.addLayout(height_row)

        # Parameter selector group — pushed to bottom
        param_group = QGroupBox("Parameters (Subplots)")
        param_layout = QVBoxLayout(param_group)
        param_layout.setContentsMargins(4, 12, 4, 4)

        # Search row: filter box + live hit count.
        search_row = QHBoxLayout()
        search_row.setSpacing(6)
        self._param_search = QLineEdit()
        self._param_search.setPlaceholderText("Search parameters…")
        self._param_search.setClearButtonEnabled(True)
        self._param_search.textChanged.connect(self._on_param_search_changed)
        search_row.addWidget(self._param_search, 1)
        self._param_hit_count = QLabel("")
        self._param_hit_count.setStyleSheet("color: palette(mid); padding: 0 4px;")
        self._param_hit_count.setMinimumWidth(48)
        search_row.addWidget(self._param_hit_count)
        param_layout.addLayout(search_row)

        # Primary actions — grouping is the main thing users want to do here.
        p_btn_row = QHBoxLayout()
        p_btn_row.setSpacing(4)
        self._btn_group = QPushButton("Group")
        self._btn_group.setToolTip(
            "Put the selected parameters on the same subplot.\n"
            "Tip: Ctrl/Shift-click parameters to multi-select, then click Group."
        )
        self._btn_group.clicked.connect(self._group_selected_params)
        self._btn_group.setEnabled(False)
        p_btn_row.addWidget(self._btn_group)
        self._btn_ungroup = QPushButton("Ungroup")
        self._btn_ungroup.setToolTip("Move the selected parameters to their own subplots.")
        self._btn_ungroup.clicked.connect(self._ungroup_selected_params)
        self._btn_ungroup.setEnabled(False)
        p_btn_row.addWidget(self._btn_ungroup)
        btn_add_subplot = QPushButton("+ Subplot")
        btn_add_subplot.setToolTip("Add a new empty subplot — drag parameters into it.")
        btn_add_subplot.clicked.connect(self._add_empty_subplot)
        p_btn_row.addWidget(btn_add_subplot)
        param_layout.addLayout(p_btn_row)

        # Secondary actions
        sec_btn_row = QHBoxLayout()
        sec_btn_row.setSpacing(4)
        btn_all = QPushButton("Check All")
        btn_all.clicked.connect(lambda: self._set_all_params(True))
        sec_btn_row.addWidget(btn_all)
        btn_none = QPushButton("Clear")
        btn_none.setToolTip("Uncheck all parameters")
        btn_none.clicked.connect(lambda: self._set_all_params(False))
        sec_btn_row.addWidget(btn_none)
        sec_btn_row.addStretch(1)
        param_layout.addLayout(sec_btn_row)

        # Tree: top-level items are subplots, children are parameters.
        # Drag-and-drop lets the user move a parameter onto a different
        # subplot to overlay it on the same axes.
        self._param_tree = QTreeWidget()
        self._param_tree.setFont(QFont("PT Sans", 9))
        self._param_tree.setHeaderHidden(True)
        self._param_tree.setRootIsDecorated(True)
        self._param_tree.setDragEnabled(True)
        self._param_tree.setAcceptDrops(True)
        self._param_tree.setDropIndicatorShown(True)
        self._param_tree.setDragDropMode(QAbstractItemView.InternalMove)
        # Extended selection so users can Ctrl/Shift-click multiple params and
        # operate on them at once (the primary grouping interaction).
        self._param_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._param_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self._param_tree.customContextMenuRequested.connect(self._on_param_tree_context_menu)
        self._param_tree.itemChanged.connect(self._on_param_changed)
        self._param_tree.itemSelectionChanged.connect(self._on_param_selection_changed)
        # Catch drops so we can sync the model and rebuild plots.
        self._param_tree.model().rowsInserted.connect(self._on_tree_rows_inserted)
        param_layout.addWidget(self._param_tree)

        # Inline tip line so the gestures aren't a hidden feature.
        tip = QLabel(
            "Tip: Ctrl/Shift-click parameters → Group · Right-click for more"
        )
        tip.setWordWrap(True)
        tip.setStyleSheet("color: palette(mid); font-size: 11px; padding: 2px 0;")
        param_layout.addWidget(tip)

        side_layout.addWidget(param_group, 2)

        root_layout.addWidget(sidebar)

        # ── Middle: plot area ────────────────────────────────────────
        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._plot_container = QWidget()
        self._plot_layout = QVBoxLayout(self._plot_container)
        # Generous outer padding + spacing so axis tick labels at the edges
        # don't collide with the sidebar/scrollbar or the next subplot below.
        self._plot_layout.setContentsMargins(8, 8, 8, 8)
        self._plot_layout.setSpacing(10)
        self._plot_layout.addStretch()
        self._scroll.setWidget(self._plot_container)
        root_layout.addWidget(self._scroll, stretch=1)

        # ── Right: tabbed analyst panels (cursor readout + statistics) ──
        right_tabs = QTabWidget()
        right_tabs.setMaximumWidth(360)
        right_tabs.setMinimumWidth(240)
        self._cursor_readout = CursorReadoutPanel()
        right_tabs.addTab(self._cursor_readout, "Cursors")
        self._stats_panel = StatisticsPanel()
        right_tabs.addTab(self._stats_panel, "Statistics")
        self._right_tabs = right_tabs
        root_layout.addWidget(right_tabs)

        # Debounced stats refresh — sigRangeChanged fires continuously while
        # the user pans/zooms; 150ms is short enough to feel live but stops
        # us from recomputing on every mouse-move tick.
        self._stats_timer = QTimer(self)
        self._stats_timer.setSingleShot(True)
        self._stats_timer.setInterval(150)
        self._stats_timer.timeout.connect(self._refresh_stats_panel)

        # Status bar
        self._status = QStatusBar()
        self.setStatusBar(self._status)
        self._status.showMessage("Ready — load test logs to begin analysis.")

    def _build_menus(self):
        mb = self.menuBar()

        # ── File ─────────────────────────────────────────────────────
        file_menu = mb.addMenu("File")
        file_menu.addAction("Load Logs...", self._on_load_logs,
                            QKeySequence("Ctrl+L"))
        # Recent-files submenu — refreshed on aboutToShow so it always
        # reflects the latest list (writes happen in _on_log_loaded).
        self._recent_menu = file_menu.addMenu("Recent Logs")
        self._recent_menu.aboutToShow.connect(self._rebuild_recent_menu)
        file_menu.addSeparator()
        file_menu.addAction("Save Session", self._save_session,
                            QKeySequence.Save)
        file_menu.addAction("Load Session...", self._load_session,
                            QKeySequence.Open)
        file_menu.addSeparator()
        file_menu.addAction("Export Visible Data as CSV...",
                            self._export_visible_csv, QKeySequence("Ctrl+E"))
        file_menu.addAction("Export Plots as Image...", self._export_image)
        file_menu.addAction("Export Plots as PDF...", self._export_pdf)
        file_menu.addSeparator()
        file_menu.addAction("Open Data Logs Folder", self._open_datalogs_folder)
        file_menu.addAction("Open Analysis Folder", self._open_analysis_folder)
        file_menu.addSeparator()
        file_menu.addAction("Close", self.close, QKeySequence("Ctrl+W"))

        # ── View ─────────────────────────────────────────────────────
        view_menu = mb.addMenu("View")
        self._auto_fit_action = QAction("Auto-Fit Y to Visible Range", self)
        self._auto_fit_action.setCheckable(True)
        self._auto_fit_action.setShortcut(QKeySequence("F"))
        self._auto_fit_action.setToolTip(
            "When on, panning/zooming the X axis automatically rescales the "
            "Y axis to fit the visible data on every subplot."
        )
        self._auto_fit_action.toggled.connect(self._on_auto_fit_y_toggled)
        view_menu.addAction(self._auto_fit_action)
        view_menu.addAction("Reset Zoom", self._reset_zoom, QKeySequence("Ctrl+0"))
        view_menu.addSeparator()
        view_menu.addAction("Show Statistics Panel",
                            lambda: self._right_tabs.setCurrentIndex(1))
        view_menu.addAction("Show Cursors Panel",
                            lambda: self._right_tabs.setCurrentIndex(0))
        view_menu.addSeparator()
        view_menu.addAction("Smoothing Window…",
                            self._prompt_smoothing_window)
        view_menu.addSeparator()
        self._wallclock_action = QAction("Wall Clock Time (X-Axis)", self)
        self._wallclock_action.setCheckable(True)
        self._wallclock_action.setToolTip(
            "Display X-axis as wall-clock time (HH:MM:SS) using the log's "
            "start timestamp.  Uncheck to revert to elapsed mm:ss."
        )
        self._wallclock_action.toggled.connect(self._on_wallclock_toggled)
        view_menu.addAction(self._wallclock_action)

        # ── Layout ───────────────────────────────────────────────────
        # The grouping operations are first-class: keyboard shortcuts make
        # them practical for daily use. The buttons in the sidebar do the
        # exact same thing — these just put hands on keys.
        layout_menu = mb.addMenu("Layout")
        layout_menu.addAction("Group Selected Parameters",
                              self._group_selected_params, QKeySequence("G"))
        layout_menu.addAction("Ungroup Selected Parameters",
                              self._ungroup_selected_params, QKeySequence("U"))
        layout_menu.addAction("New Empty Subplot",
                              self._add_empty_subplot, QKeySequence("Ctrl+N"))

        # ── Tools ────────────────────────────────────────────────────
        tools_menu = mb.addMenu("Tools")
        tools_menu.addAction("Add Common Vertical Cursor At Mouse",
                             self._add_vertical_cursor, QKeySequence("V"))
        tools_menu.addAction("Add Plot Vertical Cursor At Mouse",
                     self._add_plot_vertical_cursor, QKeySequence("Shift+V"))
        tools_menu.addAction("Add Horizontal Cursor at Mouse",
                             self._add_horizontal_cursor_at_mouse, QKeySequence("H"))
        tools_menu.addSeparator()
        tools_menu.addAction("Clear All Cursors", self._clear_all_cursors)

        # ── Scatter ──────────────────────────────────────────────────
        scatter_menu = mb.addMenu("Scatter")
        scatter_menu.addAction("X-Y Plotter...", self._open_xy_plotter)

    # ──────────────────────────────────────────────────────────────────
    # Key events (Delete to remove selected cursor)
    # ──────────────────────────────────────────────────────────────────
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Delete:
            if self._selected_v_cursor:
                self._delete_v_cursor(self._selected_v_cursor)
                return
            if self._selected_h_cursor >= 0:
                self._delete_h_cursor(self._selected_h_cursor)
                return
        super().keyPressEvent(event)

    # ──────────────────────────────────────────────────────────────────
    # Theme
    # ──────────────────────────────────────────────────────────────────
    def apply_theme(self, theme: str) -> None:
        """Public entry point — called by MainWindow when the app theme changes.

        Installs the shared card/dock/table/tab QSS on this window and repaints
        the pyqtgraph plots to match the new theme.
        """
        try:
            from .main_window import build_card_qss
            self.setStyleSheet(build_card_qss(theme))
        except Exception:
            pass
        self._apply_theme(theme)

    def _apply_theme(self, _mode: str = ""):
        bg = THEME.c('plot_bg')
        fg = THEME.c('plot_fg')
        alpha = THEME.plot_grid_alpha()
        border_color = THEME.c('border')

        for pw in self._plot_widgets:
            pw.setBackground(bg)
            pw.setStyleSheet(f"border: 1px solid {border_color}; border-radius: 4px;")
            for axis_name in ('left', 'bottom'):
                ax = pw.getAxis(axis_name)
                pen = pg.mkPen(fg)
                ax.setPen(pen)
                ax.setTextPen(pen)
            pw.showGrid(x=True, y=True, alpha=alpha)
            # Update legend styling
            try:
                legend = pw.getPlotItem().legend
                if legend is not None:
                    legend.setLabelTextColor(fg)
                    legend_bg = THEME.c('legend_bg')
                    if isinstance(legend_bg, tuple):
                        legend.setBrush(pg.mkBrush(*legend_bg))
                    else:
                        legend.setBrush(pg.mkBrush(legend_bg))
            except Exception:
                pass

    # ──────────────────────────────────────────────────────────────────
    # Log loading
    # ──────────────────────────────────────────────────────────────────
    def _next_color(self) -> str:
        c = LOG_COLORS[self._color_index % len(LOG_COLORS)]
        self._color_index += 1
        return c

    def _on_load_logs(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Load Test Log Files", get_datalogs_dir(),
            "Log Files (*.xlsx *.csv);;Excel Files (*.xlsx);;CSV Files (*.csv)")
        if not paths:
            return
        for path in paths:
            if any(l.path == path for l in self._logs.values()):
                continue
            log_id = uuid.uuid4().hex[:8]
            color = self._next_color()
            thread = LogLoaderThread(path, log_id, color, self)
            thread.log_loaded.connect(self._on_log_loaded)
            thread.error.connect(self._on_load_error)
            thread.finished.connect(lambda t=thread: self._cleanup_thread(t))
            self._loader_threads.append(thread)
            thread.start()

        n = len(self._loader_threads)
        if n:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            self._status.showMessage(f"Loading {n} file(s)... please wait")

    def _cleanup_thread(self, t):
        if t in self._loader_threads:
            self._loader_threads.remove(t)
        if not self._loader_threads:
            QApplication.restoreOverrideCursor()

    def _on_log_loaded(self, log_id: str, path: str, entry: LogEntry):
        self._logs[log_id] = entry
        self._push_recent_file(path)
        self._add_log_to_sidebar(entry)
        self._rebuild_param_list()
        self._rebuild_plots()
        self._status.showMessage(
            f"Loaded: {entry.name}  ({len(entry.elapsed)} rows)", 5000)

    def _on_load_error(self, path: str, msg: str):
        self._status.showMessage(f"Error loading {os.path.basename(path)}: {msg}", 8000)
        QApplication.restoreOverrideCursor()  # restore on error too
        _log.error("Log load failed: path=%s err=%s", path, msg)
        self._popup_warning("Load Error", f"Could not load:\n{path}\n\n{msg}")

    def _add_log_to_sidebar(self, entry: LogEntry):
        """Add a log entry as an inline row: [checkbox] [color swatch] name [offset spin]"""
        row = QHBoxLayout()
        row.setContentsMargins(2, 1, 2, 1)
        row.setSpacing(4)

        cb = QCheckBox()
        cb.setChecked(entry.visible)
        cb.stateChanged.connect(lambda state, lid=entry.id:
                                self._on_log_visibility_toggled(lid, state == 2))
        row.addWidget(cb)

        swatch = QLabel("\u25cf")
        swatch.setFont(QFont("PT Sans", 12))
        swatch.setStyleSheet(f"color: {entry.color};")
        swatch.setFixedWidth(16)
        swatch.setCursor(Qt.PointingHandCursor)
        swatch.mousePressEvent = lambda e, lid=entry.id: self._change_log_color(lid)
        row.addWidget(swatch)

        name_lbl = QLabel(entry.name)
        name_lbl.setFont(QFont("PT Sans", 8))
        name_lbl.setToolTip(entry.path)
        name_lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        row.addWidget(name_lbl, stretch=1)

        spin = QDoubleSpinBox()
        spin.setRange(-999.0, 999.0)
        spin.setSingleStep(0.1)
        spin.setDecimals(2)
        spin.setSuffix(" s")
        spin.setValue(entry.time_offset)
        spin.setFont(QFont("PT Sans", 8))
        spin.setFixedWidth(80)
        spin.setToolTip("Time offset")
        spin.valueChanged.connect(lambda v, lid=entry.id: self._on_offset_changed(lid, v))
        row.addWidget(spin)

        container = QWidget()
        container.setLayout(row)
        container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        # Insert before the stretch
        idx = self._log_list_layout.count() - 1
        self._log_list_layout.insertWidget(idx, container)

        self._log_entries_ui[entry.id] = {
            'checkbox': cb, 'spin': spin, 'container': container,
            'swatch': swatch, 'name_lbl': name_lbl,
        }

    def _on_remove_selected(self):
        if not self._logs:
            return
        last_id = list(self._logs.keys())[-1]
        self._remove_log(last_id)

    def _remove_log(self, log_id: str):
        entry = self._logs.pop(log_id, None)
        # Free cached parsed data so memory is released
        if entry is not None:
            _CSV_CACHE.pop(entry.path, None)
            entry.elapsed = np.zeros(0)
            entry.columns.clear()
        ui = self._log_entries_ui.pop(log_id, None)
        if ui:
            ui['container'].deleteLater()
        self._curves.pop(log_id, None)
        self._rebuild_param_list()
        self._rebuild_plots()

    def _change_log_color(self, log_id: str):
        entry = self._logs.get(log_id)
        if not entry:
            return
        c = QColorDialog.getColor(QColor(entry.color), self, "Log Color")
        if c.isValid():
            entry.color = c.name()
            ui = self._log_entries_ui.get(log_id)
            if ui:
                ui['swatch'].setStyleSheet(f"color: {c.name()};")
            self._rebuild_plots()

    def _on_log_visibility_toggled(self, log_id: str, visible: bool):
        if log_id not in self._logs:
            return
        self._logs[log_id].visible = visible
        self._update_curve_visibility(log_id, visible)

    def _update_curve_visibility(self, log_id: str, visible: bool):
        if log_id in self._curves:
            for curve in self._curves[log_id].values():
                curve.setVisible(visible)
        self._update_cursor_dots()
        self._update_cursor_readout()

    # ──────────────────────────────────────────────────────────────────
    # Time offset
    # ──────────────────────────────────────────────────────────────────
    def _on_offset_changed(self, log_id: str, value: float):
        entry = self._logs.get(log_id)
        if not entry:
            return
        entry.time_offset = value
        self._update_curves_for_log(log_id)
        self._update_cursor_dots()
        self._update_cursor_readout()

    def _update_curves_for_log(self, log_id: str):
        entry = self._logs.get(log_id)
        if not entry:
            return
        x = entry.elapsed + entry.time_offset
        if log_id in self._curves:
            for param, curve in self._curves[log_id].items():
                if param in entry.columns:
                    y = entry.columns[param]
                    mask = ~np.isnan(y)
                    curve.setData(x[mask], y[mask])

    # ──────────────────────────────────────────────────────────────────
    # Plot height
    # ──────────────────────────────────────────────────────────────────
    def _on_plot_height_changed(self, value: int):
        self._plot_height = value
        self._height_label.setText(f"{value}px")
        for pw in self._plot_widgets:
            pw.setMinimumHeight(value)
            pw.setMaximumHeight(value)

    # ──────────────────────────────────────────────────────────────────
    # Parameter selector
    # ──────────────────────────────────────────────────────────────────
    # ------------------------------------------------------------------
    # Subplot/parameter tree helpers
    # ------------------------------------------------------------------
    def _collect_available_params(self) -> list[str]:
        """Union of parameter names across all loaded logs (order preserved)."""
        out: list[str] = []
        seen = set()
        for entry in self._logs.values():
            for p in entry.available_params():
                if p not in seen:
                    seen.add(p)
                    out.append(p)
        return out

    def _rebuild_param_list(self):
        """Rebuild the subplot tree, preserving the user's layout & checks
        across log add/remove."""
        tree = self._param_tree
        tree.blockSignals(True)
        self._populating_tree = True

        # Snapshot prior check state by param name
        prev_checked: set[str] = set()
        for layout_group in self._iter_tree_groups():
            for p, checked in layout_group:
                if checked:
                    prev_checked.add(p)

        all_params = self._collect_available_params()

        # Build the working layout: keep any existing user layout, append new
        # params (those not yet placed) as their own subplots.
        layout: list[list[str]] = []
        placed: set[str] = set()
        for grp in self._subplot_layout:
            kept = [p for p in grp if p in all_params]
            if kept:
                layout.append(kept)
                placed.update(kept)
        for p in all_params:
            if p not in placed:
                layout.append([p])

        # If this is the very first build (no prior layout, no prior checks),
        # pre-check DEFAULT_PARAMS just like the old behaviour.
        first_build = not self._subplot_layout and not prev_checked
        self._subplot_layout = layout

        tree.clear()
        for gi, group in enumerate(layout):
            top = QTreeWidgetItem(tree, [f"Subplot {gi + 1}"])
            top.setFlags(
                (top.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsDropEnabled)
                & ~Qt.ItemIsDragEnabled
            )
            top.setCheckState(0, Qt.Unchecked)
            top.setData(0, Qt.UserRole, "subplot")
            for p in group:
                child = QTreeWidgetItem(top, [p])
                child.setFlags(
                    (child.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsDragEnabled)
                    & ~Qt.ItemIsDropEnabled
                )
                child.setData(0, Qt.UserRole, "param")
                if first_build:
                    state = Qt.Checked if p in DEFAULT_PARAMS else Qt.Unchecked
                else:
                    state = Qt.Checked if p in prev_checked else Qt.Unchecked
                child.setCheckState(0, state)
            top.setExpanded(True)

        self._renumber_subplots()
        # Sync each parent header to mirror its children's check state.
        for i in range(tree.topLevelItemCount()):
            self._refresh_parent_tristate(tree.topLevelItem(i))
        tree.blockSignals(False)
        self._populating_tree = False
        # Re-apply the active filter to newly added rows
        self._apply_param_filter(self._param_search.text())

    def _iter_tree_groups(self):
        """Yield list[(param_name, checked: bool)] for each subplot row."""
        tree = self._param_tree
        for i in range(tree.topLevelItemCount()):
            top = tree.topLevelItem(i)
            group = []
            for j in range(top.childCount()):
                child = top.child(j)
                group.append((child.text(0), child.checkState(0) == Qt.Checked))
            yield group

    def _sync_subplot_layout_from_tree(self):
        """Mirror the current tree state into self._subplot_layout."""
        self._subplot_layout = [
            [p for p, _ in grp] for grp in self._iter_tree_groups()
        ]

    @staticmethod
    def _strip_units(name: str) -> str:
        """Strip a trailing ' (Unit)' from a param name so the short label fits.
        Keeps interior parens intact (rare). ``Vehicle Speed (Kmph)`` → ``Vehicle Speed``."""
        return re.sub(r'\s*\([^()]*\)\s*$', '', name).strip() or name

    def _axis_title_for_group(self, group: list[str]) -> str:
        """Compact label suitable for a rotated Y-axis title — strips units
        and tops out at two params + '+N' so it never overshoots vertically."""
        if not group:
            return ""
        names = [self._strip_units(p) for p in group]
        head = ", ".join(names[:2])
        if len(names) > 2:
            head += f"  +{len(names) - 2}"
        return head

    def _subplot_header_text(self, top: QTreeWidgetItem, index: int) -> str:
        """Build a descriptive header like 'Subplot 2 · Speed, Power · +1'."""
        params = [top.child(j).text(0) for j in range(top.childCount())]
        if not params:
            return f"Subplot {index + 1}  ·  (empty — drag a param in)"
        preview = ", ".join(params[:2])
        if len(params) > 2:
            preview += f"  +{len(params) - 2}"
        return f"Subplot {index + 1}  ·  {preview}"

    def _renumber_subplots(self):
        """Relabel top-level rows after a drop/add/delete with previews of
        their parameters so the row is informative at a glance."""
        tree = self._param_tree
        for i in range(tree.topLevelItemCount()):
            top = tree.topLevelItem(i)
            top.setText(0, self._subplot_header_text(top, i))

    def _get_subplot_groups(self) -> list[list[str]]:
        """Return list of param lists for currently checked items, skipping
        any subplot whose params are all unchecked."""
        groups: list[list[str]] = []
        for grp in self._iter_tree_groups():
            checked_params = [p for p, c in grp if c]
            if checked_params:
                groups.append(checked_params)
        return groups

    def _set_all_params(self, checked: bool):
        tree = self._param_tree
        tree.blockSignals(True)
        state = Qt.Checked if checked else Qt.Unchecked
        for i in range(tree.topLevelItemCount()):
            top = tree.topLevelItem(i)
            for j in range(top.childCount()):
                child = top.child(j)
                # Respect the active search filter — only toggle visible rows.
                if not child.isHidden():
                    child.setCheckState(0, state)
            self._refresh_parent_tristate(top)
        tree.blockSignals(False)
        self._rebuild_plots()

    def _on_param_changed(self, item: QTreeWidgetItem, column: int = 0):
        role = item.data(0, Qt.UserRole)
        if role == "subplot":
            # User clicked the subplot header's checkbox — cascade the new
            # state down to every (visible) parameter child, then rebuild.
            # AutoTristate would otherwise make the parent state a passive
            # mirror; this turns the header check into an actual toggle-all.
            tree = self._param_tree
            tree.blockSignals(True)
            new_state = item.checkState(0)
            if new_state == Qt.PartiallyChecked:
                # Treat a partial state click as "check all" so the user can
                # use the header as a quick on/off switch.
                new_state = Qt.Checked
                item.setCheckState(0, Qt.Checked)
            for j in range(item.childCount()):
                child = item.child(j)
                if not child.isHidden():
                    child.setCheckState(0, new_state)
            tree.blockSignals(False)
            self._rebuild_plots()
            return
        if role == "param":
            # Reflect the new aggregate state on the parent header without
            # firing another itemChanged loop.
            parent = item.parent()
            if parent is not None:
                self._refresh_parent_tristate(parent)
            self._rebuild_plots()

    def _refresh_parent_tristate(self, parent: QTreeWidgetItem):
        """Manually mirror children's check states onto the parent header.
        We can't use Qt.ItemIsAutoTristate because that makes the header
        checkbox read-only for users."""
        tree = self._param_tree
        n = parent.childCount()
        checked = sum(1 for j in range(n) if parent.child(j).checkState(0) == Qt.Checked)
        tree.blockSignals(True)
        if n == 0:
            parent.setCheckState(0, Qt.Unchecked)
        elif checked == 0:
            parent.setCheckState(0, Qt.Unchecked)
        elif checked == n:
            parent.setCheckState(0, Qt.Checked)
        else:
            parent.setCheckState(0, Qt.PartiallyChecked)
        tree.blockSignals(False)

    def _on_param_search_changed(self, text: str):
        self._apply_param_filter(text)

    def _apply_param_filter(self, text: str):
        query = (text or "").strip().lower()
        tree = self._param_tree
        total = 0
        matches = 0
        for i in range(tree.topLevelItemCount()):
            top = tree.topLevelItem(i)
            top.setHidden(False)
            subplot_has_match = False
            for j in range(top.childCount()):
                child = top.child(j)
                total += 1
                is_hit = (not query) or (query in child.text(0).lower())
                child.setHidden(not is_hit)
                if is_hit:
                    matches += 1
                    if query:
                        subplot_has_match = True
            # When searching, expand subplots that contain matches so the
            # user sees them without an extra click. When not searching,
            # leave the user's expand state alone.
            if query and subplot_has_match:
                top.setExpanded(True)
            if query and not subplot_has_match:
                top.setHidden(True)

        if not query:
            self._param_hit_count.setText("")
        else:
            self._param_hit_count.setText(f"{matches}/{total}")

    def _add_empty_subplot(self):
        """Append a new empty subplot row at the bottom of the tree."""
        tree = self._param_tree
        tree.blockSignals(True)
        self._populating_tree = True
        top = QTreeWidgetItem(tree, [""])
        top.setFlags(
            (top.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsDropEnabled)
            & ~Qt.ItemIsDragEnabled
        )
        top.setCheckState(0, Qt.Unchecked)
        top.setData(0, Qt.UserRole, "subplot")
        top.setExpanded(True)
        self._renumber_subplots()
        self._populating_tree = False
        tree.blockSignals(False)
        self._sync_subplot_layout_from_tree()
        # No checked params → no plot change yet, but the row is now a drop
        # target for the user to drag parameters into.

    def _on_tree_rows_inserted(self, parent, first, last):
        """Drag-drop landed — sync model and rebuild plots."""
        # Fires for every row inserted, including during our own programmatic
        # populates. Skip those; we sync the model manually after.
        if self._populating_tree:
            return
        # Defer with a singleShot so the move completes before we re-read.
        from PySide6.QtCore import QTimer
        QTimer.singleShot(0, self._after_tree_drop)

    def _after_tree_drop(self):
        self._sync_subplot_layout_from_tree()
        # Drop any newly-emptied subplots? Keep them — user may want to
        # drag more params in. They can use the context menu to delete.
        self._renumber_subplots()
        # Refresh parent tristate on every subplot — a dropped child may have
        # changed several rows' aggregate state.
        tree = self._param_tree
        tree.blockSignals(True)
        for i in range(tree.topLevelItemCount()):
            self._refresh_parent_tristate(tree.topLevelItem(i))
        tree.blockSignals(False)
        self._rebuild_plots()

    def _on_param_tree_context_menu(self, pos):
        tree = self._param_tree
        item = tree.itemAt(pos)
        if item is None:
            return
        menu = QMenu(self)
        role = item.data(0, Qt.UserRole)
        selected_params = self._selected_param_items()
        if role == "param":
            # Grouping actions take priority when multiple params are selected.
            if len(selected_params) >= 2:
                act_group = menu.addAction(f"Group {len(selected_params)} selected on one subplot")
                act_group.triggered.connect(self._group_selected_params)
                act_ungroup = menu.addAction("Ungroup (each on own subplot)")
                act_ungroup.triggered.connect(self._ungroup_selected_params)
                menu.addSeparator()

            move_menu = menu.addMenu("Move to subplot")
            parent_top = item.parent()
            for i in range(tree.topLevelItemCount()):
                top = tree.topLevelItem(i)
                if top is parent_top:
                    continue
                act = move_menu.addAction(top.text(0))
                act.triggered.connect(
                    lambda _c=False, _child=item, _dest=top: self._move_param_to(_child, _dest))
            move_menu.addSeparator()
            act_new = move_menu.addAction("New subplot")
            act_new.triggered.connect(lambda _c=False, _child=item: self._move_param_to_new(_child))
        elif role == "subplot":
            idx = tree.indexOfTopLevelItem(item)
            n = tree.topLevelItemCount()
            act_up = menu.addAction("Move subplot up")
            act_up.setEnabled(idx > 0)
            act_up.triggered.connect(lambda _c=False, _top=item: self._move_subplot(_top, -1))
            act_down = menu.addAction("Move subplot down")
            act_down.setEnabled(idx < n - 1)
            act_down.triggered.connect(lambda _c=False, _top=item: self._move_subplot(_top, +1))
            menu.addSeparator()
            # Normalize toggle — recommended for multi-param subplots where
            # each curve has a different unit / scale (e.g. speed in kmph vs
            # power in W). The flag survives subplot reorder via frozenset key.
            is_norm = self._group_key(item) in self._normalized_subplots
            act_norm = menu.addAction(
                "Normalize curves to 0–1  " + ("✓" if is_norm else ""))
            act_norm.setToolTip(
                "Min-max scale each curve to [0,1] for the same subplot — "
                "useful when params share an axis but have very different ranges."
            )
            act_norm.triggered.connect(
                lambda _c=False, _top=item: self._toggle_subplot_normalize(_top))
            is_smooth = self._group_key(item) in self._smoothed_subplots
            act_smooth = menu.addAction(
                f"Smooth (rolling avg N={self._smoothing_window})  "
                + ("✓" if is_smooth else ""))
            act_smooth.setToolTip(
                "Overlay a moving-average smoothed version of each curve. "
                "Change the window size in View → Smoothing Window…"
            )
            act_smooth.triggered.connect(
                lambda _c=False, _top=item: self._toggle_subplot_smoothing(_top))
            menu.addSeparator()
            act_check_all = menu.addAction("Check all params in subplot")
            act_check_all.setEnabled(item.childCount() > 0)
            act_check_all.triggered.connect(
                lambda _c=False, _top=item: self._set_subplot_children_checked(_top, True))
            act_uncheck_all = menu.addAction("Uncheck all params in subplot")
            act_uncheck_all.setEnabled(item.childCount() > 0)
            act_uncheck_all.triggered.connect(
                lambda _c=False, _top=item: self._set_subplot_children_checked(_top, False))
            menu.addSeparator()
            act_del = menu.addAction("Delete subplot (split params out)")
            act_del.triggered.connect(lambda _c=False, _top=item: self._delete_subplot(_top))
        if menu.actions():
            menu.exec(tree.viewport().mapToGlobal(pos))

    # ------------------------------------------------------------------
    # Multi-select grouping actions
    # ------------------------------------------------------------------
    def _selected_param_items(self) -> list[QTreeWidgetItem]:
        return [it for it in self._param_tree.selectedItems()
                if it.data(0, Qt.UserRole) == "param"]

    def _on_param_selection_changed(self):
        n = len(self._selected_param_items())
        self._btn_group.setEnabled(n >= 2)
        self._btn_ungroup.setEnabled(n >= 1)

    def _group_selected_params(self):
        """Pull every selected parameter onto a single subplot. If all selected
        params already belong to one subplot, this is a no-op. Otherwise we
        merge them into the subplot of the first selection (most expected)
        and prune any subplot that ends up empty."""
        items = self._selected_param_items()
        if len(items) < 2:
            return
        tree = self._param_tree
        tree.blockSignals(True)
        self._populating_tree = True

        dest = items[0].parent()
        for it in items[1:]:
            old_parent = it.parent()
            if old_parent is dest:
                continue
            old_parent.removeChild(it)
            dest.addChild(it)

        # Drop any subplot that's now empty (cleanup after a merge).
        i = 0
        while i < tree.topLevelItemCount():
            top = tree.topLevelItem(i)
            if top is not dest and top.childCount() == 0:
                tree.takeTopLevelItem(i)
                continue
            i += 1

        dest.setExpanded(True)
        self._renumber_subplots()
        self._populating_tree = False
        tree.blockSignals(False)
        self._sync_subplot_layout_from_tree()
        self._rebuild_plots()

    def _ungroup_selected_params(self):
        """Move every selected param into its own dedicated subplot."""
        items = self._selected_param_items()
        if not items:
            return
        tree = self._param_tree
        tree.blockSignals(True)
        self._populating_tree = True

        for it in items:
            old_parent = it.parent()
            # Skip if it's already alone in its subplot.
            if old_parent is not None and old_parent.childCount() == 1:
                continue
            if old_parent is not None:
                old_parent.removeChild(it)
            new_top = QTreeWidgetItem(tree, [""])
            new_top.setFlags(
                (new_top.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsDropEnabled)
                & ~Qt.ItemIsDragEnabled
            )
            new_top.setCheckState(0, Qt.Unchecked)
            new_top.setData(0, Qt.UserRole, "subplot")
            new_top.addChild(it)
            new_top.setExpanded(True)

        self._renumber_subplots()
        self._populating_tree = False
        tree.blockSignals(False)
        self._sync_subplot_layout_from_tree()
        self._rebuild_plots()

    def _move_subplot(self, top: QTreeWidgetItem, delta: int):
        tree = self._param_tree
        idx = tree.indexOfTopLevelItem(top)
        new_idx = idx + delta
        if not (0 <= new_idx < tree.topLevelItemCount()):
            return
        tree.blockSignals(True)
        self._populating_tree = True
        was_expanded = top.isExpanded()
        tree.takeTopLevelItem(idx)
        tree.insertTopLevelItem(new_idx, top)
        top.setExpanded(was_expanded)
        self._renumber_subplots()
        self._populating_tree = False
        tree.blockSignals(False)
        self._sync_subplot_layout_from_tree()
        self._rebuild_plots()

    def _set_subplot_children_checked(self, top: QTreeWidgetItem, checked: bool):
        tree = self._param_tree
        tree.blockSignals(True)
        state = Qt.Checked if checked else Qt.Unchecked
        for j in range(top.childCount()):
            child = top.child(j)
            if not child.isHidden():
                child.setCheckState(0, state)
        self._refresh_parent_tristate(top)
        tree.blockSignals(False)
        self._rebuild_plots()

    # ------------------------------------------------------------------
    # Per-subplot normalize
    # ------------------------------------------------------------------
    @staticmethod
    def _group_key(top: QTreeWidgetItem) -> frozenset[str]:
        return frozenset(top.child(j).text(0) for j in range(top.childCount()))

    def _is_subplot_normalized(self, group: list[str]) -> bool:
        return frozenset(group) in self._normalized_subplots

    def _toggle_subplot_normalize(self, top: QTreeWidgetItem):
        """Right-click → Normalize toggle. State is keyed by frozenset of
        param names so it survives reorder / split-and-rejoin."""
        key = self._group_key(top)
        if key in self._normalized_subplots:
            self._normalized_subplots.discard(key)
        else:
            self._normalized_subplots.add(key)
        self._rebuild_plots()

    @staticmethod
    def _normalize_series(y: np.ndarray) -> np.ndarray:
        """Min-max scale a 1-D array to [0,1]. Constant or all-NaN series
        return zeros (so they plot flat instead of crashing)."""
        finite = y[np.isfinite(y)]
        if finite.size == 0:
            return np.zeros_like(y)
        lo = float(np.min(finite))
        hi = float(np.max(finite))
        if hi - lo < 1e-12:
            return np.zeros_like(y)
        return (y - lo) / (hi - lo)

    # ------------------------------------------------------------------
    # Smoothing (moving-average)
    # ------------------------------------------------------------------
    def _is_subplot_smoothed(self, group: list[str]) -> bool:
        return frozenset(group) in self._smoothed_subplots

    def _toggle_subplot_smoothing(self, top: QTreeWidgetItem):
        key = self._group_key(top)
        if key in self._smoothed_subplots:
            self._smoothed_subplots.discard(key)
        else:
            self._smoothed_subplots.add(key)
        self._rebuild_plots()

    def _prompt_smoothing_window(self):
        from PySide6.QtWidgets import QInputDialog
        new_n, ok = QInputDialog.getInt(
            self, "Smoothing Window",
            "Rolling-average window size (samples).\n"
            "Applies to subplots with smoothing enabled.",
            self._smoothing_window, 1, 999, 1)
        if not ok:
            return
        self._smoothing_window = int(new_n)
        if self._smoothed_subplots:
            self._rebuild_plots()
        else:
            self._status.showMessage(
                f"Smoothing window set to {new_n}. Right-click a subplot → "
                "Smooth to enable it.", 6000)

    @staticmethod
    def _moving_average(y: np.ndarray, window: int) -> np.ndarray:
        """NaN-aware centered rolling mean. Window < 2 returns the input."""
        if window is None or window < 2:
            return y
        y = np.asarray(y, dtype=float)
        if y.size == 0:
            return y
        finite = np.isfinite(y)
        y_clean = np.where(finite, y, 0.0)
        weights = np.ones(int(window))
        summed = np.convolve(y_clean, weights, mode='same')
        counted = np.convolve(finite.astype(float), weights, mode='same')
        with np.errstate(invalid='ignore', divide='ignore'):
            out = np.where(counted > 0, summed / counted, np.nan)
        return out

    # ------------------------------------------------------------------
    # Auto-fit Y to visible X range
    # ------------------------------------------------------------------
    def _on_auto_fit_y_toggled(self, on: bool):
        self._auto_fit_y = bool(on)
        if on:
            self._apply_auto_fit_y()

    def _apply_auto_fit_y(self):
        """For each subplot, compute the Y range over the currently visible
        X window across all visible curves, then set the y view-range with
        a small padding. No-op when there are no plots."""
        if not self._plot_widgets:
            return
        for pi, pw in enumerate(self._plot_widgets):
            if pi >= len(self._plot_groups):
                continue
            group = self._plot_groups[pi]
            try:
                x_range = tuple(pw.getPlotItem().vb.viewRange()[0])
            except Exception:
                continue
            normalized = self._is_subplot_normalized(group)
            ymin = np.inf
            ymax = -np.inf
            for param in group:
                for entry in self._logs.values():
                    if not entry.visible or param not in entry.columns:
                        continue
                    x = entry.elapsed + entry.time_offset
                    y = entry.columns[param]
                    if normalized:
                        y = self._normalize_series(y)
                    if x.size == 0:
                        continue
                    mask = (x >= x_range[0]) & (x <= x_range[1]) & np.isfinite(y)
                    if not np.any(mask):
                        continue
                    ymin = min(ymin, float(np.min(y[mask])))
                    ymax = max(ymax, float(np.max(y[mask])))
            if ymin == np.inf or ymax == -np.inf:
                continue
            if ymax - ymin < 1e-12:
                pad = max(abs(ymax) * 0.05, 0.5)
                ymin -= pad
                ymax += pad
            else:
                pad = (ymax - ymin) * 0.05
                ymin -= pad
                ymax += pad
            pw.getPlotItem().vb.setYRange(ymin, ymax, padding=0)

    # ------------------------------------------------------------------
    # CSV export of visible data
    # ------------------------------------------------------------------
    def _export_visible_csv(self):
        """Wide-format CSV export of every visible (log, param) curve, sliced
        to the current x view range. All series are interpolated onto a
        common time grid (the union of visible-range timestamps of the first
        log) so the file opens cleanly in Excel/pandas."""
        rows = self._visible_curve_rows()
        if not rows:
            self._popup_information(
                "Nothing to Export",
                "Load logs and check parameters before exporting.")
            return

        x_range = rows[0].get("x_range")
        # Build a master time grid from the union of all visible timestamps
        # within the x-range. Sorted and de-duplicated → monotonic so np.interp
        # is well-defined for each curve.
        masters: list[np.ndarray] = []
        for r in rows:
            x = r["x"]
            if x_range is not None:
                mask = (x >= x_range[0]) & (x <= x_range[1])
                masters.append(x[mask])
            else:
                masters.append(x)
        if not masters:
            self._popup_information("Nothing to Export", "No samples in range.")
            return
        merged = np.unique(np.concatenate(masters))
        if merged.size == 0:
            self._popup_information("Nothing to Export", "No samples in range.")
            return

        # Build the data matrix: cols = ['time', curve1, curve2, …]
        headers = ["time"]
        data_cols: list[np.ndarray] = []
        for r in rows:
            x = r["x"]
            y = r["y"]
            order = np.argsort(x)
            x_sorted = x[order]
            y_sorted = y[order]
            interp = np.interp(merged, x_sorted, y_sorted,
                               left=np.nan, right=np.nan)
            headers.append(f"{r['log_name']} · {r['param']}")
            data_cols.append(interp)

        suggested = os.path.join(
            get_analysis_dir(),
            f"{APP_NAME}_visible_data.csv")
        path, _filter = QFileDialog.getSaveFileName(
            self, "Export Visible Data", suggested, "CSV Files (*.csv)")
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"

        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(headers)
                for i in range(merged.size):
                    row = [f"{merged[i]:.6f}"]
                    for col in data_cols:
                        v = col[i]
                        row.append("" if not np.isfinite(v) else f"{v:.6g}")
                    w.writerow(row)
        except Exception as exc:
            self._popup_warning("Export Failed", str(exc))
            return
        self._status.showMessage(
            f"Exported {len(rows)} curve(s) × {merged.size} rows to "
            f"{os.path.basename(path)}", 6000)

    # ------------------------------------------------------------------
    # Recent files (Phase 2)
    # ------------------------------------------------------------------
    _RECENT_KEY = "analysis/recent_files"
    _RECENT_MAX = 10

    def _load_recent_files(self) -> list[str]:
        raw = self._qsettings.value(self._RECENT_KEY, [])
        if isinstance(raw, str):
            raw = [raw]
        if not isinstance(raw, list):
            raw = []
        return [p for p in raw if p and os.path.isfile(p)]

    def _save_recent_files(self):
        self._qsettings.setValue(self._RECENT_KEY, self._recent_files)

    def _push_recent_file(self, path: str):
        if not path:
            return
        path = os.path.abspath(path)
        if not self._recent_files:
            self._recent_files = self._load_recent_files()
        self._recent_files = [p for p in self._recent_files if p != path]
        self._recent_files.insert(0, path)
        self._recent_files = self._recent_files[:self._RECENT_MAX]
        self._save_recent_files()

    def _rebuild_recent_menu(self):
        self._recent_menu.clear()
        recent = self._load_recent_files()
        self._recent_files = recent
        if not recent:
            act = self._recent_menu.addAction("(no recent files)")
            act.setEnabled(False)
            return
        for path in recent:
            act = self._recent_menu.addAction(os.path.basename(path))
            act.setToolTip(path)
            act.triggered.connect(lambda _checked=False, p=path:
                                   self._load_single_path(p))
        self._recent_menu.addSeparator()
        clear_act = self._recent_menu.addAction("Clear Recent Files")
        clear_act.triggered.connect(self._clear_recent_files)

    def _clear_recent_files(self):
        self._recent_files = []
        self._save_recent_files()

    def _load_single_path(self, path: str):
        """Load a single log path via the same LogLoaderThread used by the
        file picker. Skips files already loaded."""
        if any(l.path == path for l in self._logs.values()):
            self._status.showMessage(f"Already loaded: {os.path.basename(path)}", 4000)
            return
        if not os.path.isfile(path):
            self._popup_warning("File Not Found", path)
            return
        log_id = uuid.uuid4().hex[:8]
        color = self._next_color()
        thread = LogLoaderThread(path, log_id, color, self)
        thread.log_loaded.connect(self._on_log_loaded)
        thread.error.connect(self._on_load_error)
        thread.finished.connect(lambda t=thread: self._cleanup_thread(t))
        self._loader_threads.append(thread)
        thread.start()

    def _move_param_to(self, child: QTreeWidgetItem, dest_top: QTreeWidgetItem):
        old_parent = child.parent()
        if old_parent is None:
            return
        old_parent.removeChild(child)
        dest_top.addChild(child)
        dest_top.setExpanded(True)
        self._sync_subplot_layout_from_tree()
        self._rebuild_plots()

    def _move_param_to_new(self, child: QTreeWidgetItem):
        self._add_empty_subplot()
        dest_top = self._param_tree.topLevelItem(self._param_tree.topLevelItemCount() - 1)
        self._move_param_to(child, dest_top)

    def _delete_subplot(self, top: QTreeWidgetItem):
        """Remove a subplot row, re-homing its params into their own subplots."""
        tree = self._param_tree
        tree.blockSignals(True)
        self._populating_tree = True
        # Move children out to new top-level subplots so they survive.
        children = [top.takeChild(0) for _ in range(top.childCount())]
        idx = tree.indexOfTopLevelItem(top)
        tree.takeTopLevelItem(idx)
        for child in children:
            new_top = QTreeWidgetItem(tree, ["Subplot"])
            new_top.setFlags(
                (new_top.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsDropEnabled)
                & ~Qt.ItemIsDragEnabled
            )
            new_top.setCheckState(0, Qt.Unchecked)
            new_top.setData(0, Qt.UserRole, "subplot")
            new_top.addChild(child)
            new_top.setExpanded(True)
        self._renumber_subplots()
        self._populating_tree = False
        tree.blockSignals(False)
        self._sync_subplot_layout_from_tree()
        self._rebuild_plots()

    # ──────────────────────────────────────────────────────────────────
    # Plot management
    # ──────────────────────────────────────────────────────────────────
    def _rebuild_plots(self):
        """Tear down and rebuild all subplots based on current state."""
        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            self._do_rebuild_plots()
        finally:
            QApplication.restoreOverrideCursor()

    def _do_rebuild_plots(self):
        for pw in self._plot_widgets:
            self._plot_layout.removeWidget(pw)
            pw.deleteLater()
        self._plot_widgets.clear()
        self._plot_groups.clear()
        self._curves.clear()
        self._crosshair_lines.clear()
        self._cursor_dots.clear()

        while self._plot_layout.count():
            item = self._plot_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        groups = self._get_subplot_groups()
        if not groups or not self._logs:
            self._plot_layout.addStretch()
            return

        bg = THEME.c('plot_bg')
        fg = THEME.c('plot_fg')
        alpha = THEME.plot_grid_alpha()

        first_pw = None
        for gi, group in enumerate(groups):
            pw = pg.PlotWidget(
                axisItems={'bottom': TimeAxisItem(orientation='bottom')})
            pw.setBackground(bg)
            pw.showGrid(x=True, y=True, alpha=alpha)
            normalized = self._is_subplot_normalized(group)
            smoothed = self._is_subplot_smoothed(group)
            # Y-axis title:
            #   1 param   → full name with units (matches sidebar swatch).
            #   ≥2 params → compact summary ("Speed, Power +1"). Full names
            #               live in the legend; rotated long titles
            #               otherwise overshoot the plot's vertical extent.
            if len(group) == 1:
                p = group[0]
                unit = _parse_unit(p)
                short = re.sub(r'\s*[\[\(][^\]\)]*[\]\)]\s*$', '', p).strip()
                axis_title = f"{short} ({unit})" if unit else p
            else:
                # Multi-param: show unit only if all params share one
                units = [_parse_unit(p) for p in group]
                shared_unit = units[0] if units and all(u == units[0] for u in units) and units[0] else None
                axis_title = self._axis_title_for_group(group)
                if shared_unit:
                    axis_title = f"{axis_title} ({shared_unit})"
            if normalized:
                axis_title = f"{axis_title}  (normalized 0–1)"
            if smoothed:
                axis_title = f"{axis_title}  · smoothed N={self._smoothing_window}"
            pw.setLabel('left', axis_title)
            pw.setMinimumHeight(self._plot_height)
            pw.setMaximumHeight(self._plot_height)
            pw.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            # Visible card-style border around each subplot so the boundary
            # is unambiguous, and internal padding so axis tick labels never
            # collide with the border or the next subplot.
            border_color = THEME.c('border')
            pw.setStyleSheet(
                f"border: 1px solid {border_color}; border-radius: 4px;"
            )
            plot_item = pw.getPlotItem()
            # (left, top, right, bottom) inside the plot — extra left padding
            # so wide tick labels (e.g. 60000) plus the rotated axis title
            # have room without colliding with the border.
            plot_item.setContentsMargins(14, 8, 14, 8)
            # Reserve a minimum width for the y-axis so tick labels and title
            # don't fight for the same pixels when values are big or names
            # are long. Width auto-grows beyond this if needed.
            plot_item.getAxis('left').setWidth(64)
            # Cap the rotated y-title at the plot's drawing height so it
            # cannot grow taller than the panel and trigger the overshoot.
            try:
                plot_item.getAxis('left').label.setTextWidth(self._plot_height - 24)
            except Exception:
                pass

            # Performance: clip and downsample
            pw.setClipToView(True)
            pw.setDownsampling(auto=True, mode='peak')

            # Y-axis padding so curves don't touch the frame
            pw.getPlotItem().vb.setDefaultPadding(0.05)

            for axis_name in ('left', 'bottom'):
                ax = pw.getAxis(axis_name)
                pen = pg.mkPen(fg)
                ax.setPen(pen)
                ax.setTextPen(pen)

            if gi == len(groups) - 1:
                pw.setLabel('bottom', 'Elapsed time (mm:ss)')
            else:
                pw.setLabel('bottom', '')

            # Apply wall-clock mode if active
            time_axis = pw.getAxis('bottom')
            if isinstance(time_axis, TimeAxisItem) and self._wall_clock_mode:
                epoch = self._best_epoch_offset()
                time_axis.epoch_offset = epoch

            if first_pw is not None:
                pw.setXLink(first_pw)
            else:
                first_pw = pw

            legend = pw.addLegend(offset=(10, 10))
            legend.setLabelTextSize('7pt')
            legend_bg = THEME.c('legend_bg')
            if isinstance(legend_bg, tuple):
                legend.setBrush(pg.mkBrush(*legend_bg))
            else:
                legend.setBrush(pg.mkBrush(legend_bg))
            legend.setLabelTextColor(fg)

            multi = len(group) > 1
            log_list = list(self._logs.values())
            for param in group:
                for li, entry in enumerate(log_list):
                    if param not in entry.columns:
                        continue
                    x = entry.elapsed + entry.time_offset
                    y = entry.columns[param]
                    if smoothed:
                        y = self._moving_average(y, self._smoothing_window)
                    if normalized:
                        y = self._normalize_series(y)
                    mask = ~np.isnan(y)
                    # Multi-param: color = parameter, style = log.
                    # Single-param: color = log, style = solid.
                    curve_color, style = _curve_visuals(group, param, li, entry.color)
                    pen = pg.mkPen(color=curve_color, width=1.6, style=style)
                    legend_name = f"{entry.name} · {param}" if multi else entry.name
                    curve = pw.plot(x[mask], y[mask], pen=pen, name=legend_name)
                    curve.setVisible(entry.visible)
                    curve.setClipToView(True)
                    self._curves.setdefault(entry.id, {})[param] = curve

            # Crosshair
            vline = pg.InfiniteLine(angle=90, movable=False,
                                    pen=pg.mkPen(THEME.c('crosshair'),
                                                 width=1, style=Qt.DashLine))
            hline = pg.InfiniteLine(angle=0, movable=False,
                                    pen=pg.mkPen(THEME.c('crosshair'),
                                                 width=1, style=Qt.DashLine))
            vline.setVisible(False)
            hline.setVisible(False)
            pw.addItem(vline, ignoreBounds=True)
            pw.addItem(hline, ignoreBounds=True)
            self._crosshair_lines[pw] = (vline, hline)

            # Mouse tracking — rate-limited at 60fps for smooth crosshair
            proxy = pg.SignalProxy(pw.scene().sigMouseMoved,
                                   rateLimit=60,
                                   slot=lambda evt, _pw=pw: self._on_mouse_moved(evt, _pw))
            pw.setProperty('_mouse_proxy', proxy)

            # Visible-range change → debounced stats refresh + (optional) auto-fit Y.
            pw.getPlotItem().vb.sigRangeChanged.connect(self._on_view_range_changed)

            self._plot_widgets.append(pw)
            self._plot_groups.append(list(group))
            self._plot_layout.addWidget(pw)

            # Per-subplot context menu
            pw.setContextMenuPolicy(Qt.CustomContextMenu)
            pw.customContextMenuRequested.connect(
                lambda pos, _pw=pw: self._on_subplot_context_menu(pos, _pw))

        self._plot_layout.addStretch()
        self._restore_v_cursors()
        self._update_cursor_dots()
        # Trigger an initial stats compute on the freshly rendered subplots.
        self._stats_timer.start()
        # (wait cursor released by _rebuild_plots wrapper)

    # ------------------------------------------------------------------
    # Wall-clock / elapsed time toggle
    # ------------------------------------------------------------------
    def _best_epoch_offset(self) -> float | None:
        """Return the earliest start_timestamp among all loaded logs."""
        ts = [e.start_timestamp for e in self._logs.values()
              if e.start_timestamp is not None]
        return min(ts) if ts else None

    def _on_wallclock_toggled(self, on: bool):
        self._wall_clock_mode = bool(on)
        epoch = self._best_epoch_offset() if on else None
        for pw in self._plot_widgets:
            time_axis = pw.getAxis('bottom')
            if isinstance(time_axis, TimeAxisItem):
                time_axis.epoch_offset = epoch
        if self._plot_widgets:
            last_pw = self._plot_widgets[-1]
            if on and epoch is not None:
                last_pw.setLabel('bottom', 'Wall clock (HH:MM:SS)')
            else:
                last_pw.setLabel('bottom', 'Elapsed time (mm:ss)')

    # ------------------------------------------------------------------
    # Per-subplot context menu
    # ------------------------------------------------------------------
    def _on_subplot_context_menu(self, pos, pw: pg.PlotWidget):
        menu = QMenu(self)
        act_csv = menu.addAction("Export visible range as CSV\u2026")
        act_csv.triggered.connect(lambda: self._export_subplot_csv(pw))
        act_img = menu.addAction("Copy plot as image")
        act_img.triggered.connect(lambda: self._copy_subplot_image(pw))
        menu.exec(pw.mapToGlobal(pos))

    def _export_subplot_csv(self, pw: pg.PlotWidget):
        """Export this subplot's visible curves to CSV."""
        pi = self._plot_widgets.index(pw) if pw in self._plot_widgets else -1
        if pi < 0:
            return
        group = self._plot_groups[pi]
        try:
            x_range = tuple(pw.getPlotItem().vb.viewRange()[0])
        except Exception:
            x_range = None
        rows = []
        for param in group:
            for entry in self._logs.values():
                if param not in entry.columns or not entry.visible:
                    continue
                rows.append({
                    "curve": f"{entry.name} \u00b7 {param}",
                    "log_name": entry.name,
                    "param": param,
                    "x": entry.elapsed + entry.time_offset,
                    "y": entry.columns[param],
                    "x_range": x_range,
                })
        if not rows:
            self._popup_information("Nothing to Export",
                                    "No visible curves on this subplot.")
            return
        # Build CSV (same wide-format logic as _export_visible_csv)
        masters = []
        for r in rows:
            x = r["x"]
            if x_range is not None:
                mask = (x >= x_range[0]) & (x <= x_range[1])
                masters.append(x[mask])
            else:
                masters.append(x)
        merged = np.unique(np.concatenate(masters))
        if merged.size == 0:
            self._popup_information("Nothing to Export", "No samples in range.")
            return
        headers = ["time"]
        data_cols = []
        for r in rows:
            x, y = r["x"], r["y"]
            order = np.argsort(x)
            interp = np.interp(merged, x[order], y[order], left=np.nan, right=np.nan)
            headers.append(r["curve"])
            data_cols.append(interp)
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Subplot Data", get_analysis_dir(), "CSV Files (*.csv)")
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(headers)
                for i in range(merged.size):
                    row = [f"{merged[i]:.6f}"]
                    for col in data_cols:
                        v = col[i]
                        row.append("" if not np.isfinite(v) else f"{v:.6g}")
                    w.writerow(row)
            self._status.showMessage(
                f"Exported subplot to {os.path.basename(path)}", 5000)
        except Exception as exc:
            self._popup_warning("Export Failed", str(exc))

    def _copy_subplot_image(self, pw: pg.PlotWidget):
        """Render the subplot to a QImage and copy to clipboard."""
        try:
            from pyqtgraph.exporters import ImageExporter
            exporter = ImageExporter(pw.getPlotItem())
            img = exporter.export(toBytes=True)
            from PySide6.QtGui import QPixmap
            QApplication.clipboard().setPixmap(QPixmap.fromImage(img))
            self._status.showMessage("Plot image copied to clipboard.", 4000)
        except Exception as exc:
            self._popup_warning("Copy Failed", str(exc))

    # ------------------------------------------------------------------
    # Checked parameters helper
    # ------------------------------------------------------------------
    def _get_checked_params(self) -> list[str]:
        """Return a flat list of all currently checked parameter names."""
        params = []
        for grp in self._iter_tree_groups():
            for p, checked in grp:
                if checked:
                    params.append(p)
        return params

    # ------------------------------------------------------------------
    # Statistics + view-range hooks
    # ------------------------------------------------------------------
    def _on_view_range_changed(self, *args):
        # Debounce — sigRangeChanged fires continuously while panning/zooming.
        self._stats_timer.start()
        # If auto-fit Y is on, recompute y-range for every subplot from the
        # currently visible x-range.
        if getattr(self, "_auto_fit_y", False):
            QTimer.singleShot(0, self._apply_auto_fit_y)

    def _visible_curve_rows(self) -> list[dict]:
        """Collect one entry per currently-rendered (log, param) curve.
        Each entry carries the raw arrays + the current x-range so the
        stats panel (or a CSV exporter) can slice consistently."""
        rows: list[dict] = []
        if not self._plot_widgets:
            return rows
        for pi, pw in enumerate(self._plot_widgets):
            if pi >= len(self._plot_groups):
                continue
            group = self._plot_groups[pi]
            multi = len(group) > 1
            try:
                x_range = tuple(pw.getPlotItem().vb.viewRange()[0])
            except Exception:
                x_range = None
            for param in group:
                for entry in self._logs.values():
                    if param not in entry.columns:
                        continue
                    if not entry.visible:
                        continue
                    label = f"{entry.name} · {param}" if multi else f"{entry.name} · {param}"
                    rows.append({
                        "curve": label,
                        "log_id": entry.id,
                        "log_name": entry.name,
                        "param": param,
                        "color": entry.color,
                        "x": entry.elapsed + entry.time_offset,
                        "y": entry.columns[param],
                        "x_range": x_range,
                    })
        return rows

    def _refresh_stats_panel(self):
        self._stats_panel.update_stats(self._visible_curve_rows())

    # ──────────────────────────────────────────────────────────────────
    # Crosshair / mouse tracking
    # ──────────────────────────────────────────────────────────────────
    def _on_mouse_moved(self, evt, pw: pg.PlotWidget):
        pos = evt[0]
        if not pw.sceneBoundingRect().contains(pos):
            if pw in self._crosshair_lines:
                self._crosshair_lines[pw][0].setVisible(False)
                self._crosshair_lines[pw][1].setVisible(False)
            return

        mouse_point = pw.getPlotItem().vb.mapSceneToView(pos)
        t = mouse_point.x()
        v = mouse_point.y()

        # Track mouse position for H-cursor placement
        self._last_mouse_pw = pw
        self._last_mouse_pos = mouse_point

        if pw in self._crosshair_lines:
            vl, hl = self._crosshair_lines[pw]
            vl.setPos(t)
            hl.setPos(v)
            vl.setVisible(True)
            hl.setVisible(True)

        # Build tooltip
        pi = self._plot_widgets.index(pw) if pw in self._plot_widgets else -1
        if pi < 0:
            return
        group = self._plot_groups[pi]
        multi = len(group) > 1
        parts = [f"t={t:.3f}s"]
        for param in group:
            for entry in self._logs.values():
                if not entry.visible or param not in entry.columns:
                    continue
                x = entry.elapsed + entry.time_offset
                if len(x) == 0:
                    continue
                idx = np.searchsorted(x, t)
                idx = min(idx, len(x) - 1)
                yv = entry.columns[param][idx]
                if not np.isnan(yv):
                    label = f"{entry.name} · {param}" if multi else entry.name
                    parts.append(f"{label}: {yv:.2f}")

        self._status.showMessage("  |  ".join(parts))

    # ──────────────────────────────────────────────────────────────────
    # Vertical cursors (global or single-plot)
    # ──────────────────────────────────────────────────────────────────
    def _resolve_vertical_cursor_anchor(self) -> tuple[Optional[pg.PlotWidget], float]:
        """Find the plot under the mouse and its X anchor position."""
        anchor_pw = None
        for pw in self._plot_widgets:
            if pw.underMouse():
                anchor_pw = pw
                break
        if anchor_pw is None and self._last_mouse_pw in self._plot_widgets:
            anchor_pw = self._last_mouse_pw
        if anchor_pw is None:
            anchor_pw = self._plot_widgets[0] if self._plot_widgets else None
        if anchor_pw is None:
            return None, 0.0

        if self._last_mouse_pos is not None and self._last_mouse_pw is anchor_pw:
            return anchor_pw, self._last_mouse_pos.x()

        vb = anchor_pw.getPlotItem().vb
        x_range = vb.viewRange()[0]
        return anchor_pw, (x_range[0] + x_range[1]) / 2.0

    def _plot_widget_for_param(self, plot_param: str) -> Optional[pg.PlotWidget]:
        """Return the current plot widget that contains the given parameter
        (may be a multi-param subplot)."""
        for idx, group in enumerate(self._plot_groups):
            if plot_param in group and idx < len(self._plot_widgets):
                return self._plot_widgets[idx]
        return None

    def _add_vertical_cursor(self, scope: str = 'all', time_value: Optional[float] = None,
                             plot_param: Optional[str] = None,
                             color: Optional[str] = None,
                             label_num: Optional[int] = None,
                             cursor_id: Optional[str] = None):
        if not self._plot_widgets:
            return

        anchor_pw, default_t = self._resolve_vertical_cursor_anchor()
        if anchor_pw is None:
            return
        t = default_t if time_value is None else time_value

        if label_num is None:
            self._v_cursor_counter += 1
            label_num = self._v_cursor_counter
        else:
            self._v_cursor_counter = max(self._v_cursor_counter, label_num)

        cid = cursor_id or uuid.uuid4().hex[:8]
        cursor_color = color or CURSOR_COLORS[(len(self._v_cursors)) % len(CURSOR_COLORS)]
        if scope == 'plot' and plot_param is None and anchor_pw in self._plot_widgets:
            group = self._plot_groups[self._plot_widgets.index(anchor_pw)]
            plot_param = group[0] if group else None

        if scope == 'plot':
            target_pw = self._plot_widget_for_param(plot_param) if plot_param else anchor_pw
            if target_pw is None:
                return
            target_plots = [target_pw]
        else:
            target_plots = list(self._plot_widgets)

        lines: dict[pg.PlotWidget, pg.InfiniteLine] = {}
        for pw in target_plots:
            line = pg.InfiniteLine(
                pos=t, angle=90, movable=True,
                pen=pg.mkPen(cursor_color, width=2),
                label=f'C{label_num}: {t:.2f}s',
                labelOpts={'position': 0.95, 'color': cursor_color,
                           'fill': THEME.c('cursor_label_bg'),
                           'movable': True})
            pw.addItem(line, ignoreBounds=True)
            line.sigPositionChanged.connect(
                lambda l, _cid=cid: self._on_v_cursor_moved(_cid, l))
            line.sigClicked.connect(
                lambda *a, _cid=cid: self._select_v_cursor(_cid))
            lines[pw] = line

        cursor_data = {
            'id': cid,
            'label': label_num,
            'scope': scope,
            'plot_param': plot_param,
            'lines': lines,
            'time': t,
            'color': cursor_color,
        }
        self._v_cursors.append(cursor_data)
        self._select_v_cursor(cid)
        self._update_cursor_dots()
        self._update_cursor_readout()

    def _add_plot_vertical_cursor(self):
        self._add_vertical_cursor(scope='plot')

    def _restore_v_cursors(self):
        """Re-add existing vertical cursor lines to newly rebuilt plots."""
        for cdata in self._v_cursors:
            cid = cdata['id']
            new_lines: dict[pg.PlotWidget, pg.InfiniteLine] = {}
            if cdata.get('scope') == 'plot':
                target_pw = self._plot_widget_for_param(cdata.get('plot_param'))
                target_plots = [target_pw] if target_pw is not None else []
            else:
                target_plots = list(self._plot_widgets)

            for pw in target_plots:
                line = pg.InfiniteLine(
                    pos=cdata['time'], angle=90, movable=True,
                    pen=pg.mkPen(cdata['color'], width=2),
                    label=f'C{cdata["label"]}: {cdata["time"]:.2f}s',
                    labelOpts={'position': 0.95, 'color': cdata['color'],
                               'fill': THEME.c('cursor_label_bg'),
                               'movable': True})
                pw.addItem(line, ignoreBounds=True)
                line.sigPositionChanged.connect(
                    lambda l, _cid=cid: self._on_v_cursor_moved(_cid, l))
                line.sigClicked.connect(
                    lambda *a, _cid=cid: self._select_v_cursor(_cid))
                new_lines[pw] = line
            cdata['lines'] = new_lines
        # Re-apply selected highlight
        if self._selected_v_cursor:
            self._select_v_cursor(self._selected_v_cursor)

    def _on_v_cursor_moved(self, cursor_id: str, moved_line: pg.InfiniteLine):
        cdata = self._find_cursor_by_id(cursor_id)
        if cdata is None:
            return
        t = moved_line.value()
        cdata['time'] = t
        for pw, line in cdata['lines'].items():
            if line is not moved_line:
                line.blockSignals(True)
                line.setValue(t)
                line.blockSignals(False)
            line.label.setText(f'C{cdata["label"]}: {t:.2f}s')
        self._select_v_cursor(cursor_id)
        self._update_cursor_dots_for_id(cursor_id)
        self._update_cursor_readout()

    def _find_cursor_by_id(self, cursor_id: str):
        """Return cursor data dict by ID, or None."""
        for cdata in self._v_cursors:
            if cdata.get('id') == cursor_id:
                return cdata
        return None

    def _select_v_cursor(self, cursor_id: str):
        """Mark cursor as selected (red pen), deselect others."""
        prev_id = self._selected_v_cursor
        self._selected_v_cursor = cursor_id
        # Restore previous cursor color
        if prev_id and prev_id != cursor_id:
            prev = self._find_cursor_by_id(prev_id)
            if prev:
                pen = pg.mkPen(prev['color'], width=2)
                for line in prev['lines'].values():
                    line.setPen(pen)
        # Highlight new selected cursor
        cur = self._find_cursor_by_id(cursor_id)
        if cur:
            sel_pen = pg.mkPen(SELECTED_CURSOR_COLOR, width=3)
            for line in cur['lines'].values():
                line.setPen(sel_pen)

    def _delete_v_cursor(self, cursor_id_or_index):
        """Delete a vertical cursor by ID string or list index (int)."""
        # Support being called with an index (from keyPressEvent)
        if isinstance(cursor_id_or_index, int):
            idx = cursor_id_or_index
            if idx < 0 or idx >= len(self._v_cursors):
                return
            cid = self._v_cursors[idx]['id']
        else:
            cid = cursor_id_or_index
        cdata = self._find_cursor_by_id(cid)
        if cdata is None:
            return
        self._v_cursors.remove(cdata)
        for pw, line in cdata['lines'].items():
            pw.removeItem(line)
        # Remove tracking dots
        if cid in self._cursor_dots:
            for dot in self._cursor_dots.pop(cid):
                dot['pw'].removeItem(dot['item'])
        # Reset selection
        if self._selected_v_cursor == cid:
            self._selected_v_cursor = ''
        self._update_cursor_readout()

    def _update_cursor_readout(self):
        visible_logs = [e for e in self._logs.values() if e.visible]
        params = self._get_checked_params()
        self._cursor_readout.update_readout(self._v_cursors, visible_logs, params)

    def _update_cursor_dots(self):
        """Rebuild tracking dots for all cursors on all plots."""
        for cid, dots in self._cursor_dots.items():
            for dot in dots:
                dot['pw'].removeItem(dot['item'])
        self._cursor_dots.clear()
        for cdata in self._v_cursors:
            self._update_cursor_dots_for_id(cdata['id'])

    def _update_cursor_dots_for_id(self, cursor_id: str):
        """Place small tracking dots where cursor intersects each curve."""
        cdata = self._find_cursor_by_id(cursor_id)
        if cdata is None:
            return
        # Remove old dots for this cursor
        if cursor_id in self._cursor_dots:
            for dot in self._cursor_dots[cursor_id]:
                dot['pw'].removeItem(dot['item'])
        dots = []
        t = cdata['time']
        log_list = list(self._logs.values())
        for pw in cdata['lines'].keys():
            if pw not in self._plot_widgets:
                continue
            pi = self._plot_widgets.index(pw)
            group = self._plot_groups[pi]
            for param in group:
                for li, entry in enumerate(log_list):
                    if not entry.visible or param not in entry.columns:
                        continue
                    x = entry.elapsed + entry.time_offset
                    if len(x) == 0:
                        continue
                    idx = np.searchsorted(x, t)
                    idx = min(idx, len(x) - 1)
                    yv = entry.columns[param][idx]
                    if np.isnan(yv):
                        continue
                    # Match the dot color to the actual curve color so the
                    # dot stays visually attached to its trace in multi-
                    # param subplots.
                    dot_color, _style = _curve_visuals(group, param, li, entry.color)
                    dot_item = pg.ScatterPlotItem(
                        [t], [yv], size=8,
                        pen=pg.mkPen(THEME.c('plot_bg'), width=1),
                        brush=pg.mkBrush(dot_color))
                    pw.addItem(dot_item, ignoreBounds=True)
                    dots.append({'pw': pw, 'item': dot_item})
        self._cursor_dots[cursor_id] = dots

    # ──────────────────────────────────────────────────────────────────
    # Horizontal cursors (placed at mouse position)
    # ──────────────────────────────────────────────────────────────────
    def _add_horizontal_cursor_at_mouse(self):
        """Add h-cursor on the plot currently under the mouse, or last hovered."""
        # Check which plot the mouse is physically over right now
        pw = None
        for p in self._plot_widgets:
            if p.underMouse():
                pw = p
                break
        # Fall back to last tracked plot
        if pw is None and self._last_mouse_pw in self._plot_widgets:
            pw = self._last_mouse_pw
        if pw is None:
            if self._plot_widgets:
                pw = self._plot_widgets[0]
            else:
                return
        if self._last_mouse_pos is not None and self._last_mouse_pw is pw:
            val = self._last_mouse_pos.y()
        else:
            y_range = pw.getPlotItem().vb.viewRange()[1]
            val = (y_range[0] + y_range[1]) / 2.0

        ci = len(self._h_cursors)
        color = CURSOR_COLORS[(ci + 2) % len(CURSOR_COLORS)]
        line = pg.InfiniteLine(
            pos=val, angle=0, movable=True,
            pen=pg.mkPen(color, width=2, style=Qt.DashDotLine),
            label=f'{val:.2f}',
            labelOpts={'position': 0.05, 'color': color,
                       'fill': THEME.c('cursor_label_bg'),
                       'movable': True})
        pw.addItem(line, ignoreBounds=True)
        line.sigPositionChanged.connect(
            lambda l: l.label.setText(f'{l.value():.2f}'))
        line.sigClicked.connect(
            lambda *a, _ci=ci: self._on_h_cursor_selected(_ci))
        self._h_cursors.append({
            'line': line, 'plot_widget': pw,
            'value': val, 'color': color,
        })

    def _on_h_cursor_selected(self, cursor_index: int):
        """Mark h-cursor as selected (red pen), deselect others."""
        prev = self._selected_h_cursor
        self._selected_h_cursor = cursor_index
        if 0 <= prev < len(self._h_cursors) and prev != cursor_index:
            hc = self._h_cursors[prev]
            hc['line'].setPen(pg.mkPen(hc['color'], width=2, style=Qt.DashDotLine))
        if 0 <= cursor_index < len(self._h_cursors):
            self._h_cursors[cursor_index]['line'].setPen(
                pg.mkPen(SELECTED_CURSOR_COLOR, width=3, style=Qt.DashDotLine))

    def _delete_h_cursor(self, cursor_index: int):
        if cursor_index < 0 or cursor_index >= len(self._h_cursors):
            return
        hc = self._h_cursors.pop(cursor_index)
        hc['plot_widget'].removeItem(hc['line'])
        if self._selected_h_cursor == cursor_index:
            self._selected_h_cursor = -1
        elif self._selected_h_cursor > cursor_index:
            self._selected_h_cursor -= 1

    def _clear_all_cursors(self):
        # Vertical
        for cdata in self._v_cursors:
            for pw, line in cdata['lines'].items():
                pw.removeItem(line)
        self._v_cursors.clear()
        self._selected_v_cursor = ''

        # Cursor dots
        for cid, dots in self._cursor_dots.items():
            for dot in dots:
                dot['pw'].removeItem(dot['item'])
        self._cursor_dots.clear()

        # Horizontal
        for hc in self._h_cursors:
            hc['plot_widget'].removeItem(hc['line'])
        self._h_cursors.clear()
        self._selected_h_cursor = -1

        self._update_cursor_readout()

    def _reset_zoom(self):
        for pw in self._plot_widgets:
            pw.autoRange()

    def _set_plot_export_theme(self, pw: pg.PlotWidget, dark: bool = False):
        """Temporarily set plot to white bg + black axes for export."""
        pw.setBackground('w')
        fg = pg.mkPen('#000000')
        for ax_name in ('left', 'bottom', 'top', 'right'):
            ax = pw.getAxis(ax_name)
            ax.setPen(fg)
            ax.setTextPen(fg)

    def _restore_plot_theme(self, pw: pg.PlotWidget):
        """Restore plot to current theme colors."""
        pw.setBackground(THEME.c('plot_bg'))
        fg = pg.mkPen(THEME.c('plot_fg'))
        for ax_name in ('left', 'bottom', 'top', 'right'):
            ax = pw.getAxis(ax_name)
            ax.setPen(fg)
            ax.setTextPen(fg)

    # ──────────────────────────────────────────────────────────────────
    # Session save / load
    # ──────────────────────────────────────────────────────────────────
    def _save_session(self):
        folder = get_analysis_dir()
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Analysis Session", folder,
            "Analysis Session (*.json)")
        if not path:
            return
        if not path.endswith('.json'):
            path += '.json'

        # Flatten currently checked params for the legacy 'parameters' key so
        # older app versions can still partially read this session.
        groups = self._get_subplot_groups()
        flat_checked: list[str] = []
        for g in groups:
            flat_checked.extend(g)

        data = {
            'version': SESSION_VERSION,
            'plot_height': self._plot_height,
            'logs': [],
            'parameters': flat_checked,
            # New in v4: full subplot layout (each subplot's param list) and
            # the user's persisted layout including unchecked params.
            'subplots': groups,
            'subplot_layout': self._subplot_layout,
            'cursors': {
                'vertical': [],
                'horizontal': [],
            }
        }

        for entry in self._logs.values():
            data['logs'].append({
                'path': entry.path,
                'color': entry.color,
                'visible': entry.visible,
                'time_offset': entry.time_offset,
            })

        for cdata in self._v_cursors:
            data['cursors']['vertical'].append({
                'time': cdata['time'],
                'color': cdata['color'],
                'label': cdata.get('label'),
                'scope': cdata.get('scope', 'all'),
                'plot_param': cdata.get('plot_param'),
            })

        for hc in self._h_cursors:
            pi = -1
            for i, pw in enumerate(self._plot_widgets):
                if pw is hc['plot_widget']:
                    pi = i
                    break
            data['cursors']['horizontal'].append({
                'plot_index': pi,
                'value': hc['line'].value(),
                'color': hc['color'],
            })

        if self._plot_widgets:
            x_range = self._plot_widgets[0].getPlotItem().vb.viewRange()[0]
            data['view_range'] = {'x': x_range}

        try:
            with open(path, 'w') as f:
                json.dump(data, f, indent=2)
            self._status.showMessage(f"Session saved to {os.path.basename(path)}", 5000)
        except Exception as exc:
            self._popup_warning("Save Error", str(exc))

    def _load_session(self):
        folder = get_analysis_dir()
        path, _ = QFileDialog.getOpenFileName(
            self, "Load Analysis Session", folder,
            "Analysis Session (*.json)")
        if not path:
            return

        try:
            with open(path, 'r') as f:
                data = json.load(f)
        except Exception as exc:
            self._popup_warning("Load Error", str(exc))
            return

        ver = data.get('version', 0)
        if ver not in (1, 2, 3, 4):
            self._popup_warning("Version Mismatch", "Session file version not supported.")
            return

        # Clear current state
        self._clear_all_cursors()
        log_ids = list(self._logs.keys())
        for lid in log_ids:
            self._remove_log(lid)

        # Restore plot height
        ph = data.get('plot_height', DEFAULT_PLOT_HEIGHT)
        self._plot_height = max(MIN_PLOT_HEIGHT, min(MAX_PLOT_HEIGHT, ph))
        self._height_slider.setValue(self._plot_height)

        session_params = set(data.get('parameters', []))
        # v4+ stores the layout; older versions only stored flat 'parameters'.
        session_subplot_layout = data.get('subplot_layout')
        if session_subplot_layout is None:
            session_subplots = data.get('subplots')
            if session_subplots is not None:
                session_subplot_layout = session_subplots
            else:
                # v1-v3 fallback: each checked param gets its own subplot.
                session_subplot_layout = [[p] for p in data.get('parameters', [])]

        pending_logs = data.get('logs', [])
        self._pending_session = {
            'params': session_params,
            'subplot_layout': session_subplot_layout,
            'cursors': data.get('cursors', {}),
            'view_range': data.get('view_range'),
            'remaining': len(pending_logs),
        }

        for log_info in pending_logs:
            log_path = log_info['path']
            if not os.path.isfile(log_path):
                self._status.showMessage(f"File not found: {log_path}", 5000)
                self._pending_session['remaining'] -= 1
                continue
            log_id = uuid.uuid4().hex[:8]
            color = log_info.get('color', self._next_color())
            thread = LogLoaderThread(log_path, log_id, color, self)
            thread.log_loaded.connect(
                lambda lid, p, e, info=log_info: self._on_session_log_loaded(lid, p, e, info))
            thread.error.connect(self._on_load_error)
            thread.finished.connect(lambda t=thread: self._cleanup_thread(t))
            self._loader_threads.append(thread)
            thread.start()

    def _on_session_log_loaded(self, log_id, path, entry, log_info):
        entry.color = log_info.get('color', entry.color)
        entry.visible = log_info.get('visible', True)
        entry.time_offset = log_info.get('time_offset', 0.0)

        self._logs[log_id] = entry
        self._add_log_to_sidebar(entry)

        # Update checkbox and offset spinbox from restored UI
        ui = self._log_entries_ui.get(log_id)
        if ui:
            if not entry.visible:
                ui['checkbox'].setChecked(False)
            ui['spin'].setValue(entry.time_offset)

        self._pending_session['remaining'] -= 1
        if self._pending_session['remaining'] <= 0:
            self._finish_session_restore()

    def _finish_session_restore(self):
        session = self._pending_session

        # Restore the persisted subplot layout BEFORE rebuilding, so the tree
        # comes back with the same grouping the user saved.
        self._subplot_layout = [list(g) for g in session.get('subplot_layout', [])]
        self._rebuild_param_list()

        # Apply checked state from the flat 'parameters' list.
        self._param_tree.blockSignals(True)
        for i in range(self._param_tree.topLevelItemCount()):
            top = self._param_tree.topLevelItem(i)
            for j in range(top.childCount()):
                child = top.child(j)
                child.setCheckState(
                    0,
                    Qt.Checked if child.text(0) in session['params']
                    else Qt.Unchecked,
                )
        self._param_tree.blockSignals(False)

        self._rebuild_plots()

        # Restore cursors
        cursors = session.get('cursors', {})
        for vc in cursors.get('vertical', []):
            self._add_vertical_cursor(
                scope=vc.get('scope', 'all'),
                time_value=vc.get('time', 0),
                plot_param=vc.get('plot_param'),
                color=vc.get('color'),
                label_num=vc.get('label'),
            )

        for hc_data in cursors.get('horizontal', []):
            pi = hc_data.get('plot_index', 0)
            if 0 <= pi < len(self._plot_widgets):
                pw = self._plot_widgets[pi]
                val = hc_data.get('value', 0)
                color = hc_data.get('color', '#e63946')
                line = pg.InfiniteLine(
                    pos=val, angle=0, movable=True,
                    pen=pg.mkPen(color, width=2, style=Qt.DashDotLine),
                    label=f'{val:.2f}',
                    labelOpts={'position': 0.05, 'color': color,
                               'fill': THEME.c('cursor_label_bg'),
                               'movable': True})
                pw.addItem(line, ignoreBounds=True)
                line.sigPositionChanged.connect(
                    lambda l: l.label.setText(f'{l.value():.2f}'))
                self._h_cursors.append({
                    'line': line, 'plot_widget': pw,
                    'value': val, 'color': color,
                })

        # Restore view range
        vr = session.get('view_range')
        if vr and self._plot_widgets:
            x = vr.get('x')
            if x and len(x) == 2:
                self._plot_widgets[0].setXRange(x[0], x[1], padding=0)

        self._update_cursor_dots()
        self._update_cursor_readout()
        self._pending_session = None
        self._status.showMessage("Session restored.", 5000)

    # ──────────────────────────────────────────────────────────────────
    # Utility actions
    # ──────────────────────────────────────────────────────────────────
    def _open_datalogs_folder(self):
        os.startfile(get_datalogs_dir())

    def _open_analysis_folder(self):
        os.startfile(get_analysis_dir())

    def _open_xy_plotter(self):
        if self._xy_window is not None and self._xy_window.isVisible():
            self._xy_window.raise_()
            self._xy_window.activateWindow()
            return
        # parent=None so the window can freely go behind other windows
        self._xy_window = XYPlotWindow(self._logs)
        self._xy_window.show()

    # ──────────────────────────────────────────────────────────────────
    # Export
    # ──────────────────────────────────────────────────────────────────
    def _export_image(self):
        if not self._plot_widgets:
            self._popup_information("Export", "No plots to export.")
            return
        path, filt = QFileDialog.getSaveFileName(
            self, "Export Plots as Image", get_analysis_dir(),
            "PNG Image (*.png);;SVG Vector (*.svg)")
        if not path:
            return
        self._do_export(path, filt)

    def _export_pdf(self):
        if not self._plot_widgets:
            self._popup_information("Export", "No plots to export.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Plots as PDF", get_analysis_dir(),
            "PDF Document (*.pdf)")
        if not path:
            return
        if not path.lower().endswith('.pdf'):
            path += '.pdf'
        self._do_export(path, "PDF")

    def _do_export(self, path: str, filt: str):
        try:
            from PySide6.QtGui import QImage, QPixmap
            from PySide6.QtCore import QMarginsF
            from pyqtgraph.exporters import ImageExporter

            if filt and "SVG" in filt:
                from pyqtgraph.exporters import SVGExporter
                for i, pw in enumerate(self._plot_widgets):
                    suffix = "_".join(self._plot_groups[i]) if i < len(self._plot_groups) else str(i)
                    # Strip any path separators a parameter name might contain.
                    suffix = re.sub(r'[\\/:*?"<>|]', '_', suffix)
                    p = path if len(self._plot_widgets) == 1 else \
                        path.replace('.svg', f'_{suffix}.svg')
                    exporter = SVGExporter(pw.getPlotItem())
                    exporter.export(p)
                self._status.showMessage(f"Exported SVG to {os.path.dirname(path)}", 5000)
                return

            if filt and "PDF" in filt or path.lower().endswith('.pdf'):
                from PySide6.QtGui import QPageSize, QPageLayout
                from PySide6.QtPrintSupport import QPrinter
                # Render each plot to image with forced white background + black axes
                plot_images = []
                export_width = 1600
                for pw in self._plot_widgets:
                    self._set_plot_export_theme(pw, dark=False)
                    exporter = ImageExporter(pw.getPlotItem())
                    exporter.parameters()['width'] = export_width
                    exporter.parameters()['background'] = pg.mkColor('w')
                    img = exporter.export(toBytes=True)
                    self._restore_plot_theme(pw)
                    plot_images.append(img)

                # Stitch vertically onto one canvas with white bg
                total_h = sum(img.height() for img in plot_images) + 8 * max(len(plot_images) - 1, 0)
                canvas = QImage(export_width, total_h, QImage.Format_ARGB32)
                canvas.fill(QColor('#ffffff'))
                painter_c = QPainter(canvas)
                y = 0
                for img in plot_images:
                    painter_c.drawImage(0, y, img)
                    y += img.height() + 8
                painter_c.end()

                printer = QPrinter(QPrinter.HighResolution)
                printer.setOutputFormat(QPrinter.PdfFormat)
                printer.setOutputFileName(path)
                printer.setPageSize(QPageSize(QPageSize.A4))
                printer.setPageOrientation(QPageLayout.Landscape)
                printer.setPageMargins(QMarginsF(10, 10, 10, 10))

                painter = QPainter()
                painter.begin(printer)
                page_rect = printer.pageRect(QPrinter.DevicePixel)
                pixmap = QPixmap.fromImage(canvas)
                # Scale to fit page preserving aspect ratio
                scaled = pixmap.scaled(
                    page_rect.width(), page_rect.height(),
                    Qt.KeepAspectRatio, Qt.SmoothTransformation)
                x_off = int((page_rect.width() - scaled.width()) / 2)
                y_off = int((page_rect.height() - scaled.height()) / 2)
                painter.drawPixmap(x_off, y_off, scaled)
                painter.end()
                self._status.showMessage(f"Exported PDF to {os.path.basename(path)}", 5000)
                return

            # Default: PNG composite using ImageExporter (avoids blank OpenGL grabs)
            images = []
            total_height = 0
            width = 0
            for pw in self._plot_widgets:
                exporter = ImageExporter(pw.getPlotItem())
                img = exporter.export(toBytes=True)
                images.append(img)
                total_height += img.height() + 10
                width = max(width, img.width())

            composite = QImage(width, total_height, QImage.Format_ARGB32)
            composite.fill(QColor(THEME.c('plot_bg')))
            painter = QPainter(composite)
            y_offset = 0
            for img in images:
                painter.drawImage(0, y_offset, img)
                y_offset += img.height() + 10
            painter.end()
            composite.save(path)
            self._status.showMessage(f"Exported PNG to {os.path.basename(path)}", 5000)

        except Exception as exc:
            self._popup_warning("Export Error", str(exc))

    # ──────────────────────────────────────────────────────────────────
    # Cleanup & persistence
    # ──────────────────────────────────────────────────────────────────
    def closeEvent(self, event):
        # Persist UI state
        try:
            qs = self._qsettings
            qs.setValue("analysis/plot_height", self._plot_height)
            if self._plot_widgets:
                x_range = self._plot_widgets[0].getPlotItem().vb.viewRange()[0]
                qs.setValue("analysis/x_range", x_range)
            cursor_times = [c['time'] for c in self._v_cursors]
            qs.setValue("analysis/cursor_positions", cursor_times)
        except Exception:
            pass
        for t in self._loader_threads:
            t.quit()
            t.wait(500)
        self._loader_threads.clear()
        super().closeEvent(event)
