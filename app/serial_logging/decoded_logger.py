"""Decoded signal log writer (.xlsx, two sheets, cycle-buffered wide format).

Output workbook
---------------
* ``Metadata`` – key/value rows describing the session (app, port, baud,
  config path, file names, start time, etc.).
* ``Data`` – one wide row per **complete poll cycle**. Columns are grouped
  per frame in ``FrameConfig.frames`` order. Each frame block starts with
  ``<FrameName>.elapsed_ms`` and ``<FrameName>.frame_id`` (the latter
  carries the ``0xNNNN`` literal so the frame is identifiable even after
  a column rename), followed by that frame's signal columns prefixed
  with ``<FrameName>.``. The next frame's block follows, and so on.

  Bitfield signals expand into one ``0``/``1`` column per defined bit
  (``<signal>.<bit_name>``) — no raw integer column.
  Enum signals keep their raw integer column and gain a sibling
  ``<signal>.label`` column with the decoded text.

  If two different frames define a signal with the same name, both
  columns are emitted in config order. Their header text is identical;
  internally each cell is independently addressable by column position
  so neither overwrites the other.

Cycle Buffer pattern
--------------------
Frames arrive one at a time and each is stashed in an in-memory buffer
keyed by ``frame_id``. When the **trigger frame** (the last frame in the
config order) arrives, the buffer is checked: if every configured frame
has been seen since the previous emit, one wide row is appended to the
Data sheet. The buffer is **always cleared on trigger arrival** so a
later cycle never carries stale values forward; incomplete cycles are
dropped silently.

Persistence trade-off
---------------------
openpyxl write-only mode streams rows into the workbook in memory but
only persists to disk on :meth:`close`. A crash before Stop Logging
loses the decoded workbook; the raw CSV (streamed) is unaffected.

Threading model
---------------
``log_frame()`` runs on the caller's thread (typically the Qt event-loop
thread inside ``MainWindow._handle_packet``). It mutates the cycle buffer
synchronously and — when the trigger frame closes a cycle — queues the
assembled wide-row list to a dedicated daemon writer thread. The writer
thread is the sole owner of the openpyxl ``Workbook`` after :meth:`open`
returns: it appends each queued row to the ``Data`` sheet, then saves the
workbook to disk before the thread exits inside :meth:`close`.

Moving the per-row ``ws.append`` off the caller's thread is a large win in
practice — measured at ~5x reduction in synchronous log_frame latency at
1 kHz on the perf harness (see PERF_NOTES.md). At rates the writer cannot
sustain, the bounded queue (``_WRITER_QUEUE_SIZE``) drops rows rather than
blocking the GUI; the dropped count is logged on close().
"""

from __future__ import annotations

import csv
import json
import logging
import queue
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Mapping, Optional, Tuple

from openpyxl import Workbook

from ..decoder.frame_decoder import DecodedFrame
from ..decoder.types import BitfieldSpec, FrameConfig, SignalSpec

_LOG = logging.getLogger("bytehound.serial_logging.decoded")

ErrorCallback = Callable[[str], None]

# Internal slot keys for the per-frame buffer entry. Stored alongside the
# real column names so we can look up each frame's housekeeping column
# positions without re-formatting the prefix each time.
_SLOT_ELAPSED = "__elapsed_ms__"
_SLOT_FRAME_ID = "__frame_id__"

# Bound the writer queue. At 1 kHz with the canonical config this is ~100 s of
# buffered cycle rows — generous enough that a transient disk stall (antivirus
# scan, slow USB stick) does not drop data, but bounded so a permanently stuck
# writer cannot exhaust memory.
_WRITER_QUEUE_SIZE = 100_000

# Sentinel pushed onto the queue by close() to signal the writer thread to
# drain remaining rows and exit. Using a distinct object (not None) lets us
# tolerate accidental Nones from a buggy caller.
_WRITER_STOP = object()


def _format_number(value: float | int) -> float | int:
    return value


def _signal_label(name: str, unit: str) -> str:
    return f"{name} ({unit})" if unit else name


def _bitfield_specs_for(config: FrameConfig, spec: SignalSpec) -> List[BitfieldSpec]:
    """Look up bitfield specs for *spec* using the same key fallback the decoder uses."""
    for name in (spec.source_name, spec.signal_name):
        if not name:
            continue
        specs = config.bitfields.get((spec.frame_id, name))
        if specs:
            return specs
    return []


def _is_enum_signal(config: FrameConfig, spec: SignalSpec) -> bool:
    for name in (spec.source_name, spec.signal_name):
        if name and (spec.frame_id, name) in config.enums:
            return True
    return False


class DecodedLogger:
    METADATA_SHEET = "Metadata"
    DATA_SHEET = "Data"

    def __init__(
        self,
        path: str | Path,
        config: FrameConfig,
        *,
        flush_interval: float = 0.5,
        metadata: Mapping[str, str] | None = None,
        on_error: ErrorCallback | None = None,
        on_warning: Callable[[str], None] | None = None,
        polling_mode: bool = False,
    ) -> None:
        self.path = Path(path)
        self._config = config
        self._workbook: Optional[Workbook] = None
        self._data_ws = None
        self._metadata = dict(metadata) if metadata else {}
        self._on_error = on_error
        self._on_warning = on_warning
        self._disabled = False
        self.polling_mode = polling_mode

        (
            self._columns,
            self._column_by_key,
            self._bit_columns_by_key,
            self._enum_label_col_by_key,
            self._block_by_frame,
            self._cycle_frame_ids,
        ) = self._build_columns()
        self._trigger_id: Optional[int] = (
            self._required_cycle_frame_ids[-1] if self._required_cycle_frame_ids else None
        )
        # Per-frame slot keyed by column POSITION (int), not name. Lets two
        # frames write into independently-positioned cells even if their
        # column headers happen to share the same text.
        self._cycle_buffer: Dict[int, Dict[int, Any]] = {}

        # Kept for API compatibility; xlsx output cannot flush incrementally.
        self._flush_interval = float(flush_interval)
        self._last_flush = 0.0

        # Writer-thread plumbing. The queue is created at module-construct
        # time so callers can bind their own bound size before open(); the
        # thread itself is spawned by open() so closed-then-reopened loggers
        # get a fresh worker.
        self._queue: "queue.Queue[Any]" = queue.Queue(maxsize=_WRITER_QUEUE_SIZE)
        self._writer_thread: Optional[threading.Thread] = None
        self._stop_event = threading.Event()
        self._error_lock = threading.Lock()
        self._pending_error: Optional[str] = None
        self._dropped_count = 0  # rows dropped because the queue was full
        self._last_drop_warning_time = 0.0

    @property
    def _required_cycle_frame_ids(self) -> List[int]:
        tx_command_fids = {cmd.frame_id for cmd in self._config.tx_commands.values()}

        # If there are strictly RX frames, prioritize those
        rx_only = [
            fid for fid in self._cycle_frame_ids
            if fid in self._config.frames and self._config.frames[fid].direction == "rx"
        ]
        if rx_only:
            return rx_only

        # Otherwise, require RX-capable frames that don't have TX commands
        candidates = [
            fid for fid in self._cycle_frame_ids
            if (fid not in self._config.frames or self._config.frames[fid].is_rx_capable)
            and fid not in tx_command_fids
        ]
        if candidates:
            return candidates

        # Absolute fallback to avoid empty list
        return [
            fid for fid in self._cycle_frame_ids
            if fid not in self._config.frames or self._config.frames[fid].is_rx_capable
        ]

    def __enter__(self) -> "DecodedLogger":
        self.open()
        return self

    def __exit__(self, *exc_info) -> None:
        self.close()

    def open(self) -> None:
        if self._disabled:
            return
        if self._workbook is not None:
            return
        # A prior close() may have left a still-alive writer thread when its
        # bounded join timed out. If that thread has since finished, clear
        # the reference so we can start a fresh session.
        if self._writer_thread is not None and not self._writer_thread.is_alive():
            self._writer_thread = None
        if self._writer_thread is not None:
            return
        self.path.parent.mkdir(parents=True, exist_ok=True)

        self._tmp_data_path = self.path.parent / (self.path.name + ".tmp_data")
        self._tmp_meta_path = self.path.parent / (self.path.name + ".tmp_meta")

        # Write metadata JSON
        try:
            with self._tmp_meta_path.open("w", encoding="utf-8") as f:
                json.dump(self._metadata, f)
        except Exception as exc:
            self._handle_error("open", exc)
            return

        # Open temp CSV file
        try:
            self._fp = self._tmp_data_path.open("w", encoding="utf-8", newline="")
            self._writer = csv.writer(self._fp)
            self._writer.writerow(self._columns)
            self._fp.flush()
        except Exception as exc:
            self._handle_error("open", exc)
            return

        self._last_flush = time.monotonic()

        # Dummy references for public API state checking
        self._workbook = True
        self._data_ws = True

        # Spawn the writer thread. After this point the CSV file is owned
        # exclusively by the writer.
        self._stop_event.clear()
        self._dropped_count = 0
        self._last_drop_warning_time = 0.0
        self._writer_thread = threading.Thread(
            target=self._writer_loop,
            name=f"DecodedLoggerWriter[{self.path.name}]",
            daemon=True,
        )
        self._writer_thread.start()

    def close(self, timeout: float = 10.0) -> None:
        # Signal the writer thread to drain remaining rows, save the
        # workbook, and exit. We join before clearing references so the file
        # is on disk by the time close() returns (callers — including the
        # tests — rely on this).
        writer_still_running = False
        if self._writer_thread is not None:
            self._stop_event.set()
            try:
                self._queue.put_nowait(_WRITER_STOP)
            except queue.Full:
                # Queue is at capacity; the writer will see the stop event
                # once it has drained enough to make room.
                pass
            if timeout > 0.0:
                self._writer_thread.join(timeout=timeout)
            writer_still_running = self._writer_thread.is_alive()
            # KEEP the thread reference if it's still alive so callers can
            # observe is_draining() / await_drain() and wait for true
            # completion. Daemon threads are killed at interpreter exit;
            # without an explicit wait, slow-disk drains lose rows.
            if not writer_still_running:
                self._writer_thread = None

        # The writer thread already saved + closed the workbook (or will,
        # if it's still draining). Drop our references so the next open()
        # builds a fresh one.
        self._workbook = None
        self._data_ws = None

        # Only drain the queue ourselves if the writer thread already exited
        # (it died before pulling the sentinel, or finished cleanly). If the
        # join above timed out and the writer is still alive, it owns the
        # queue and will drain it in its own finally block — racing it here
        # with get_nowait() would silently steal rows out from under the
        # writer on slow-disk shutdowns. Caught by
        # tests/test_logging.py::test_decoded_logger_persists_workbook_when_close_join_times_out.
        if not writer_still_running:
            while True:
                try:
                    self._queue.get_nowait()
                except queue.Empty:
                    break

        if self._dropped_count:
            msg = f"DecodedLogger dropped {self._dropped_count} row(s) at {self.path.name} — writer could not keep up."
            _LOG.warning(msg)
            if self._on_warning is not None:
                self._on_warning(msg)
            self._dropped_count = 0
        # Surface any writer-thread error to the on_error callback on the
        # calling thread.
        self._pump_pending_error()

    # ------------------------------------------------------------------
    # Drain observation (used by MainWindow.closeEvent to avoid losing
    # rows when the interpreter would kill the daemon writer thread)
    # ------------------------------------------------------------------

    def is_draining(self) -> bool:
        """True if close() returned with the writer thread still alive
        (i.e. it didn't finish saving the workbook within the bounded
        join). Callers that need a strong "data on disk" guarantee
        should follow up with await_drain() before allowing the process
        to exit.
        """
        return self._writer_thread is not None and self._writer_thread.is_alive()

    def pending_rows(self) -> int:
        """Approximate number of rows the writer thread still has to drain.
        For UI status displays — exact value can change between this call
        and the next as the writer processes items.
        """
        return self._queue.qsize()

    def await_drain(self, timeout: float | None = None) -> bool:
        """Block until the writer thread exits or ``timeout`` elapses.
        Returns True if the writer completed (workbook is on disk), False
        if the timeout expired with the writer still running.
        """
        t = self._writer_thread
        if t is None:
            return True
        t.join(timeout=timeout)
        finished = not t.is_alive()
        if finished:
            self._writer_thread = None
        return finished

    def set_flush_interval(self, seconds: float) -> None:
        try:
            self._flush_interval = max(0.0, float(seconds))
        except (TypeError, ValueError):
            self._flush_interval = 0.5

    def log_frame(
        self,
        decoded: DecodedFrame,
        elapsed_ms: int,
        timestamp: datetime | None = None,
    ) -> None:
        # Surface any writer-thread error to the on_error callback before
        # accepting more data. Mirrors RawLogger's two-thread error pattern.
        self._pump_pending_error()
        if self._disabled:
            return
        try:
            if self._writer_thread is None:
                self.open()
            if self._writer_thread is None:
                return
            self._queue.put_nowait((decoded, elapsed_ms))
        except queue.Full:
            # Writer is wedged or saturated. Drop rather than block the GUI
            # thread; the dropped count is logged on close().
            self._dropped_count += 1
            now = time.monotonic()
            if now - self._last_drop_warning_time > 2.0:
                self._last_drop_warning_time = now
                msg = f"DecodedLogger queue is full — dropping log frames. Total dropped: {self._dropped_count}"
                _LOG.warning(msg)
                if self._on_warning is not None:
                    self._on_warning(msg)
        except Exception as exc:
            self._handle_error("write", exc)

    def _process_frame_on_writer(self, decoded: DecodedFrame, elapsed_ms: int) -> None:
        block = self._block_by_frame.get(decoded.frame_id)
        if block is None:
            # Frame not represented in the schema — nothing to do.
            return

        # Always emit and start a new row if we receive a frame we already have
        # in the current cycle buffer (this naturally handles continuous streams)
        if decoded.frame_id in self._cycle_buffer:
            self._maybe_emit_row_on_writer(force=True)
            self._cycle_buffer.clear()

        slot = self._cycle_buffer.setdefault(decoded.frame_id, {})
        # Each frame has its own elapsed_ms and frame_id columns.
        if _SLOT_ELAPSED in block:
            slot[block[_SLOT_ELAPSED]] = elapsed_ms
        if _SLOT_FRAME_ID in block:
            slot[block[_SLOT_FRAME_ID]] = f"0x{decoded.frame_id:04X}"
        for signal in [*decoded.signals, *decoded.calculations]:
            key = (signal.frame_id, signal.signal_name)

            # Bitfield: one 0/1 cell per defined bit. No raw column.
            bit_cols = self._bit_columns_by_key.get(key)
            if bit_cols:
                for bit_name, active in signal.bit_values.items():
                    pos = bit_cols.get(bit_name)
                    if pos is not None:
                        slot[pos] = 1 if active else 0
                continue

            # Enum: raw value in the signal column, label in the sibling
            # `.label` column. Both written when available so unknown raw
            # values still appear in the raw column even if no label matches.
            label_pos = self._enum_label_col_by_key.get(key)
            if label_pos is not None:
                raw_pos = self._column_by_key.get(key)
                if raw_pos is not None and signal.raw_value is not None:
                    slot[raw_pos] = int(signal.raw_value)
                if signal.enum_label is not None:
                    slot[label_pos] = signal.enum_label
                continue

            # Scalar signal: write the scaled value as before.
            if signal.scaled_value is None:
                continue
            pos = self._column_by_key.get(key)
            if pos is None:
                continue
            slot[pos] = _format_number(signal.scaled_value)

        # Always emit on the trigger frame, forcing incomplete rows to be written
        # rather than silently dropping them.
        if self._trigger_id is not None and decoded.frame_id == self._trigger_id:
            self._maybe_emit_row_on_writer(force=True)
            self._cycle_buffer.clear()

    def _maybe_emit_row_on_writer(self, force: bool = False) -> None:
        if not self._cycle_buffer:
            return
        if not force and not all(fid in self._cycle_buffer for fid in self._required_cycle_frame_ids):
            return
        # Position-keyed merge: each frame writes into its own column slots,
        # so even when two frames share a header text they land in distinct
        # cells of the wide row.
        flat: Dict[int, Any] = {}
        for slot in self._cycle_buffer.values():
            flat.update(slot)
        row = [flat.get(i, "") for i in range(len(self._columns))]
        self._write_one(self._writer, row)

    def _write_one(self, writer, row: list) -> bool:
        if writer is None:
            return False
        try:
            writer.writerow(row)
            return True
        except Exception as exc:
            self._record_error("write", exc)
            return False

    def _writer_loop(self) -> None:
        """Drain queued frames and write/format them to the temporary CSV file."""
        writer = self._writer
        fp = self._fp
        try:
            while not self._stop_event.is_set():
                try:
                    item = self._queue.get(timeout=0.1)
                except queue.Empty:
                    continue
                if item is _WRITER_STOP:
                    break
                decoded, elapsed_ms = item
                self._process_frame_on_writer(decoded, elapsed_ms)
                self._maybe_periodic_flush(fp)
            # Drain anything still buffered between the stop signal and the
            # join() — even if rows arrived after the sentinel they were
            # legitimate cycle emits and the user expects them on disk.
            while True:
                try:
                    item = self._queue.get_nowait()
                except queue.Empty:
                    break
                if item is _WRITER_STOP:
                    continue
                decoded, elapsed_ms = item
                self._process_frame_on_writer(decoded, elapsed_ms)
                self._maybe_periodic_flush(fp)
        finally:
            if self.polling_mode and self._cycle_buffer:
                self._maybe_emit_row_on_writer(force=True)
                self._cycle_buffer.clear()
            self._save_and_close_workbook(writer, fp)

    def _maybe_periodic_flush(self, fp) -> None:
        if fp is None:
            return
        if self._flush_interval <= 0:
            try:
                fp.flush()
                self._last_flush = time.monotonic()
            except Exception as exc:
                self._record_error("flush", exc)
            return
        now = time.monotonic()
        if now - self._last_flush >= self._flush_interval:
            try:
                fp.flush()
                self._last_flush = now
            except Exception as exc:
                self._record_error("flush", exc)

    def _save_and_close_workbook(self, writer, fp) -> None:
        if fp is not None:
            try:
                fp.flush()
            except Exception as exc:
                self._record_error("flush", exc)
            try:
                fp.close()
            except Exception:
                pass

        # Compile the final .xlsx workbook from temporary CSV/JSON files.
        self._compile_workbook()

    def _compile_workbook(self) -> None:
        if not hasattr(self, "_tmp_data_path") or not self._tmp_data_path.exists():
            return

        wb = Workbook(write_only=True)
        try:
            meta_ws = wb.create_sheet(title=self.METADATA_SHEET)
            meta_ws.append(["Key", "Value"])
            metadata = {}
            if self._tmp_meta_path.exists():
                try:
                    with self._tmp_meta_path.open("r", encoding="utf-8") as f:
                        metadata = json.load(f)
                except Exception:
                    metadata = self._metadata
            for key in sorted(metadata):
                value = str(metadata[key]).replace("\n", " ").strip()
                meta_ws.append([key, value])

            data_ws = wb.create_sheet(title=self.DATA_SHEET)
            with self._tmp_data_path.open("r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    typed_row = []
                    for cell in row:
                        if cell == "":
                            typed_row.append("")
                        else:
                            try:
                                if "." in cell or "e" in cell.lower():
                                    typed_row.append(float(cell))
                                else:
                                    typed_row.append(int(cell))
                            except ValueError:
                                typed_row.append(cell)
                    data_ws.append(typed_row)

            wb.save(self.path)
        except Exception as exc:
            self._record_error("compile", exc)
        finally:
            try:
                wb.close()
            except Exception:
                pass
            if self.path.exists() and self.path.stat().st_size > 0:
                try:
                    self._tmp_data_path.unlink(missing_ok=True)
                    self._tmp_meta_path.unlink(missing_ok=True)
                except Exception:
                    pass

    @classmethod
    def recover_temp_files(cls, data_path: Path, meta_path: Path, target_path: Path) -> None:
        """Build the final xlsx file from leftover temp files."""
        if not data_path.exists():
            return

        import json
        import csv
        from openpyxl import Workbook

        wb = Workbook(write_only=True)
        try:
            meta_ws = wb.create_sheet(title="Metadata")
            meta_ws.append(["Key", "Value"])
            metadata = {}
            if meta_path.exists():
                with meta_path.open("r", encoding="utf-8") as f:
                    metadata = json.load(f)
            for key in sorted(metadata):
                value = str(metadata[key]).replace("\n", " ").strip()
                meta_ws.append([key, value])

            data_ws = wb.create_sheet(title="Data")
            with data_path.open("r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    typed_row = []
                    for cell in row:
                        if cell == "":
                            typed_row.append("")
                        else:
                            try:
                                if "." in cell or "e" in cell.lower():
                                    typed_row.append(float(cell))
                                else:
                                    typed_row.append(int(cell))
                            except ValueError:
                                typed_row.append(cell)
                    data_ws.append(typed_row)

            target_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target_path)

            # Clean up temp files
            data_path.unlink(missing_ok=True)
            meta_path.unlink(missing_ok=True)
        except Exception as exc:
            _LOG.error("Failed to recover decoded log from %s: %s", data_path, exc, exc_info=True)
            raise exc
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _record_error(self, context: str, exc: Exception) -> None:
        _LOG.error("DecodedLogger %s error (writer thread)", context, exc_info=True)
        with self._error_lock:
            if self._pending_error is None:
                self._pending_error = f"Decoded log {context} error for {self.path.name}: {exc}"

    def _pump_pending_error(self) -> None:
        with self._error_lock:
            err = self._pending_error
            self._pending_error = None
        if err is not None:
            self._disabled = True
            if self._on_error is not None:
                self._on_error(err)

    def _build_columns(
        self,
    ) -> Tuple[
        List[str],
        Dict[Tuple[int, str], int],
        Dict[Tuple[int, str], Dict[str, int]],
        Dict[Tuple[int, str], int],
        Dict[int, Dict[str, int]],
        List[int],
    ]:
        # Cycle frames = configured frames that have at least one signal,
        # in the order they appear in FrameConfig.frames (insertion order).
        cycle_frame_ids = [
            fid
            for fid in self._config.frames.keys()
            if self._config.signals_by_frame.get(fid)
        ]

        # Map each calc group name to the frame ids that contribute signals
        # to it, used to fan calc_groups (with no explicit frame_id) into the
        # right per-frame blocks.
        frames_by_group: Dict[str, List[int]] = {}
        for spec in self._config.all_signals:
            if not spec.group:
                continue
            frames = frames_by_group.setdefault(spec.group, [])
            if spec.frame_id not in frames:
                frames.append(spec.frame_id)

        # Every column in a frame's block carries a `<FrameName>.` prefix.
        # The block starts with elapsed_ms + frame_id housekeeping, then
        # signal/bit/enum columns. Slots are keyed by column POSITION so
        # duplicate header text from cross-frame collisions still maps to
        # independent cells.
        columns: List[str] = []
        column_by_key: Dict[Tuple[int, str], int] = {}
        bit_columns_by_key: Dict[Tuple[int, str], Dict[str, int]] = {}
        enum_label_col_by_key: Dict[Tuple[int, str], int] = {}
        block_by_frame: Dict[int, Dict[str, int]] = {}

        for frame_id in cycle_frame_ids:
            prefix = f"0x{frame_id:X}."

            elapsed_pos = len(columns)
            columns.append(f"{prefix}elapsed_ms")
            frame_id_pos = len(columns)
            columns.append(f"{prefix}frame_id")
            block: Dict[str, int] = {
                _SLOT_ELAPSED: elapsed_pos,
                _SLOT_FRAME_ID: frame_id_pos,
            }

            for spec in self._config.signals_by_frame.get(frame_id, []):
                bit_specs = _bitfield_specs_for(self._config, spec)
                if bit_specs:
                    # Bitfield: emit one 0/1 column per bit; no raw column.
                    bit_cols: Dict[str, int] = {}
                    for bit in bit_specs:
                        pos = len(columns)
                        columns.append(f"{prefix}{spec.signal_name}.{bit.bit_name}")
                        bit_cols[bit.bit_name] = pos
                    bit_columns_by_key[(frame_id, spec.signal_name)] = bit_cols
                    continue

                sig_label = f"{prefix}{_signal_label(spec.signal_name, spec.unit)}"
                pos = len(columns)
                columns.append(sig_label)
                column_by_key[(frame_id, spec.signal_name)] = pos
                block[spec.signal_name] = pos

                if _is_enum_signal(self._config, spec):
                    # Enum: add a sibling `.label` column next to the raw value.
                    label_pos = len(columns)
                    columns.append(f"{prefix}{spec.signal_name}.label")
                    enum_label_col_by_key[(frame_id, spec.signal_name)] = label_pos

            for calc in self._config.calc_groups:
                if calc.frame_id is not None:
                    if calc.frame_id != frame_id:
                        continue
                elif frame_id not in frames_by_group.get(calc.group, []):
                    continue

                # Resolve unit: use calc.unit or inherit from the group's signals
                calc_unit = calc.unit
                if not calc_unit:
                    for sig_spec in self._config.all_signals:
                        if sig_spec.group == calc.group and sig_spec.unit:
                            calc_unit = sig_spec.unit
                            break

                signal_name = f"{calc.group} {calc.stat}"
                calc_label = _signal_label(signal_name, calc_unit)
                pos = len(columns)
                columns.append(calc_label)
                column_by_key[(frame_id, signal_name)] = pos
                block[signal_name] = pos

            block_by_frame[frame_id] = block

        return (
            columns,
            column_by_key,
            bit_columns_by_key,
            enum_label_col_by_key,
            block_by_frame,
            cycle_frame_ids,
        )

    def _handle_error(self, context: str, exc: Exception) -> None:
        """Handle a calling-thread error (synchronous path inside log_frame).

        The writer thread owns the openpyxl Workbook now, so we cannot close
        it here. Signal the writer to exit; it will save+close on its way out.
        """
        if self._disabled:
            return
        self._disabled = True
        self._cycle_buffer.clear()
        if self._writer_thread is not None:
            self._stop_event.set()
            try:
                self._queue.put_nowait(_WRITER_STOP)
            except queue.Full:
                pass
        _LOG.error("DecodedLogger %s error", context, exc_info=True)
        if self._on_error is not None:
            self._on_error(f"Decoded log {context} error for {self.path.name}: {exc}")
