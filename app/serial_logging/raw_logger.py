"""Raw packet log writer (CSV).

Writes a CSV with columns ``timestamp,direction,hex,delta_t_ms``. The hex
field keeps space-separated bytes for human readability (e.g.
``AA 55 00 10``).

Threading model
---------------
``log()`` is called from the UI thread on every received/transmitted
frame. The actual disk write — including `csv.writerow` and periodic
``flush()`` — runs on a dedicated daemon thread that drains an internal
``queue.Queue``. This means the Qt event loop is never blocked by a slow
disk (network share, USB stick, antivirus scan).

* ``log()`` formats the timestamp (cheap) and enqueues a tuple. Returns
  immediately even if the disk is stuck.
* The writer thread blocks on the queue and writes rows as they arrive.
* On ``close()`` the UI thread enqueues a sentinel and joins the writer,
  guaranteeing all pending rows are flushed before the file is closed.

Errors from the writer thread are captured into ``_pending_error`` and
surfaced on the **next** UI-thread call to ``log()`` (or to
``pump_errors()``), invoking the ``on_error`` callback there. The writer
thread itself never calls the callback — that would re-enter
``_stop_logging`` on the wrong thread and deadlock the join.
"""

from __future__ import annotations

import csv
import logging
import queue
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Callable, Mapping, Optional, TextIO

_FLUSH_INTERVAL = 0.5  # seconds between explicit OS-level flushes
_LOG = logging.getLogger("bytehound.serial_logging.raw")

# Bound the handoff queue so a permanently stuck disk cannot exhaust memory.
# At ~1 kHz frame rate this is 100 s of buffered data; well above any
# plausible disk-stall window we care about.
_WRITER_QUEUE_SIZE = 100_000

ErrorCallback = Callable[[str], None]


class RawLogger:
    COLUMNS = ["timestamp", "direction", "hex", "delta_t_ms"]

    def __init__(
        self,
        path: str | Path,
        *,
        flush_interval: float = _FLUSH_INTERVAL,
        metadata: Mapping[str, str] | None = None,
        on_error: ErrorCallback | None = None,
        on_warning: Callable[[str], None] | None = None,
        hex_format: str = "hex",
    ) -> None:
        self.path = Path(path)
        self._fp: TextIO | None = None
        self._writer: "csv._writer | None" = None
        self._last_flush: float = 0.0
        self._flush_interval = float(flush_interval)
        self._metadata = dict(metadata) if metadata else {}
        self._on_error = on_error
        self._on_warning = on_warning
        self._disabled = False
        if hex_format not in {"hex", "compact"}:
            raise ValueError(f"hex_format must be 'hex' or 'compact' (got {hex_format!r})")
        self._hex_format = hex_format

        self._queue: "queue.Queue[Optional[tuple]]" = queue.Queue(maxsize=_WRITER_QUEUE_SIZE)
        self._writer_thread: Optional[threading.Thread] = None
        self._stop_event = threading.Event()

        # Set by the writer thread, read+cleared by the UI thread on its next
        # log() call. Surfaces I/O errors to on_error on the right thread.
        self._error_lock = threading.Lock()
        self._pending_error: Optional[str] = None
        self._dropped_count = 0  # rows dropped because the queue was full
        self._last_drop_warning_time = 0.0

    def __enter__(self) -> "RawLogger":
        self.open()
        return self

    def __exit__(self, *exc_info) -> None:
        self.close()

    def open(self) -> None:
        if self._disabled:
            return
        # A prior close() may have left a still-alive writer thread when its
        # bounded join timed out (slow disk). If that thread has since
        # finished, clear the reference so we can start a fresh session.
        if self._writer_thread is not None and not self._writer_thread.is_alive():
            self._writer_thread = None
        if self._writer_thread is not None:
            return
        self.path.parent.mkdir(parents=True, exist_ok=True)

        self._tmp_data_path = self.path.parent / (self.path.name + ".tmp_data")
        self._tmp_meta_path = self.path.parent / (self.path.name + ".tmp_meta")

        # Write metadata JSON
        import json
        try:
            with self._tmp_meta_path.open("w", encoding="utf-8") as f:
                json.dump(self._metadata, f)
        except Exception as exc:
            self._handle_error("open", exc)
            return

        # Open temp CSV file
        try:
            self._fp = self._tmp_data_path.open("w", encoding="utf-8", newline="")
            self._writer = csv.writer(self._fp)
            self._writer.writerow(self.COLUMNS)
            self._fp.flush()
        except Exception as exc:
            self._handle_error("open", exc)
            return

        self._last_flush = time.monotonic()

        self._stop_event.clear()
        self._dropped_count = 0
        self._last_drop_warning_time = 0.0
        self._writer_thread = threading.Thread(
            target=self._writer_loop,
            name=f"RawLoggerWriter[{self.path.name}]",
            daemon=True,
        )
        self._writer_thread.start()

    def close(self, timeout: float = 5.0) -> None:
        # Signal the writer thread to finish, then wait for it to drain the
        # queue and exit before we close the file. If the join times out
        # (slow disk), KEEP the writer_thread reference so callers can
        # observe is_draining() / call await_drain() to wait for true
        # completion before the interpreter shuts down — daemon threads
        # are killed mid-flush at process exit, which would lose data.
        if self._writer_thread is not None:
            self._stop_event.set()
            try:
                self._queue.put_nowait(None)
            except queue.Full:
                pass
            if timeout > 0.0:
                self._writer_thread.join(timeout=timeout)
            if not self._writer_thread.is_alive():
                self._writer_thread = None
        self._fp = None
        self._writer = None

        if self._dropped_count > 0:
            msg = f"RawLogger closed with {self._dropped_count} dropped row(s) due to queue saturation."
            _LOG.warning(msg)
            if self._on_warning is not None:
                self._on_warning(msg)
            self._dropped_count = 0

    # ------------------------------------------------------------------
    # Drain observation (used by MainWindow.closeEvent to avoid losing
    # rows when the interpreter would kill the daemon writer thread)
    # ------------------------------------------------------------------

    def is_draining(self) -> bool:
        """True if close() returned with the writer thread still alive
        (i.e. it didn't finish within the bounded join). Callers that need
        a strong "data on disk" guarantee should follow up with
        await_drain() before allowing the process to exit.
        """
        return self._writer_thread is not None and self._writer_thread.is_alive()

    def pending_rows(self) -> int:
        """Approximate number of rows the writer thread still has to drain.
        For UI status displays — exact value can change between this call
        and the next as the writer processes items.
        """
        return self._queue.qsize()

    def await_drain(self, timeout: float | None = None) -> bool:
        """Block until the writer thread exits or ``timeout`` elapses.
        Returns True if the writer completed (data is on disk), False if
        the timeout expired with the writer still running.
        """
        t = self._writer_thread
        if t is None:
            return True
        t.join(timeout=timeout)
        finished = not t.is_alive()
        if finished:
            self._writer_thread = None
        return finished

    def set_flush_interval(self, seconds: float) -> None:
        try:
            interval = float(seconds)
        except (TypeError, ValueError):
            interval = _FLUSH_INTERVAL
        if interval < 0:
            interval = 0.0
        self._flush_interval = interval

    def log(self, direction: str, raw: bytes, timestamp: datetime | None = None, delta_t_ms: float = 0.0) -> None:
        # Surface any error the writer thread captured since the last call,
        # so on_error fires on the correct (UI) thread.
        self._pump_pending_error()
        if self._disabled:
            return
        try:
            if self._writer_thread is None:
                self.open()
            if self._writer_thread is None:
                return
            if isinstance(timestamp, (int, float)):
                ts_dt = datetime.fromtimestamp(timestamp)
            elif isinstance(timestamp, datetime):
                ts_dt = timestamp
            else:
                ts_dt = datetime.now()
            ts = ts_dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            try:
                self._queue.put_nowait((ts, direction.upper(), raw, delta_t_ms))
            except queue.Full:
                # Disk is stuck; drop the row rather than block the UI. Count
                # how many we dropped so the writer thread can log a warning
                # when it next gets to run.
                self._dropped_count += 1
                now = time.monotonic()
                if now - self._last_drop_warning_time > 2.0:
                    self._last_drop_warning_time = now
                    msg = f"RawLogger queue is full — dropping log rows. Total dropped: {self._dropped_count}"
                    _LOG.warning(msg)
                    if self._on_warning is not None:
                        self._on_warning(msg)
        except Exception as exc:
            self._handle_error("write", exc)

    # ------------------------------------------------------------------
    # Writer thread
    # ------------------------------------------------------------------

    def _writer_loop(self) -> None:
        """Drain the queue and write rows to disk until stopped."""
        fp = self._fp
        writer = self._writer
        try:
            while not self._stop_event.is_set():
                try:
                    item = self._queue.get(timeout=0.1)
                except queue.Empty:
                    self._maybe_periodic_flush(fp)
                    continue
                if item is None:
                    break
                if not self._write_one(writer, item):
                    return
                self._maybe_periodic_flush(fp)
        finally:
            # Drain anything enqueued between the stop signal and the join.
            # NB: never `return` from inside this `finally` — that would
            # silently swallow any exception escaping the main try block.
            # `break` lets the final flush still run and any pending
            # exception still propagate.
            while True:
                try:
                    item = self._queue.get_nowait()
                except queue.Empty:
                    break
                if item is None:
                    continue
                if not self._write_one(writer, item):
                    break
            self._save_and_close_workbook(writer, fp)

    def _save_and_close_workbook(self, writer, fp) -> None:
        if fp is not None:
            try:
                fp.flush()
            except Exception as exc:
                self._record_error("flush", exc)
            try:
                fp.close()
            except Exception:
                pass
        self._compile_workbook()

    def _compile_workbook(self) -> None:
        if not hasattr(self, "_tmp_data_path") or not self._tmp_data_path.exists():
            return

        import json
        import shutil
        from openpyxl import Workbook

        MAX_ROWS_PER_SHEET = 1_000_000
        excel_success = False

        wb = Workbook(write_only=True)
        try:
            meta_ws = wb.create_sheet(title="Metadata")
            meta_ws.append(["Key", "Value"])
            metadata = {}
            if self._tmp_meta_path.exists():
                try:
                    with self._tmp_meta_path.open("r", encoding="utf-8") as f:
                        metadata = json.load(f)
                except Exception:
                    metadata = self._metadata
            for key in sorted(metadata):
                value = str(metadata[key]).replace("\n", " ").strip()
                meta_ws.append([key, value])

            sheet_index = 1
            data_ws = wb.create_sheet(title="Data")
            header_row = None
            current_sheet_rows = 0

            with self._tmp_data_path.open("r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    if header_row is None:
                        header_row = row
                        data_ws.append(row)
                        current_sheet_rows += 1
                        continue

                    if current_sheet_rows >= MAX_ROWS_PER_SHEET:
                        sheet_index += 1
                        data_ws = wb.create_sheet(title=f"Data_{sheet_index}")
                        data_ws.append(header_row)
                        current_sheet_rows = 1

                    typed_row = []
                    for cell in row:
                        if cell == "":
                            typed_row.append("")
                        else:
                            try:
                                if "." in cell or "e" in cell.lower():
                                    typed_row.append(float(cell))
                                else:
                                    typed_row.append(int(cell))
                            except ValueError:
                                typed_row.append(cell)
                    data_ws.append(typed_row)
                    current_sheet_rows += 1

            self.path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(self.path)
            excel_success = True
        except Exception as exc:
            self._record_error("compile", exc)
        finally:
            try:
                wb.close()
            except Exception:
                pass

        if excel_success and self.path.exists() and self.path.stat().st_size > 0:
            try:
                self._tmp_data_path.unlink(missing_ok=True)
                self._tmp_meta_path.unlink(missing_ok=True)
            except Exception:
                pass
        else:
            try:
                csv_target = self.path.with_suffix(".csv")
                shutil.copy2(self._tmp_data_path, csv_target)
                _LOG.info("Preserved raw temp log as CSV fallback: %s", csv_target)
                self._tmp_data_path.unlink(missing_ok=True)
                self._tmp_meta_path.unlink(missing_ok=True)
            except Exception as copy_exc:
                _LOG.error("Failed to copy CSV fallback for %s: %s", self.path, copy_exc)

    @classmethod
    def recover_temp_files(cls, data_path: Path, meta_path: Path, target_path: Path) -> None:
        """Build the final xlsx file from leftover temp files with multi-sheet support and CSV fallback."""
        if not data_path.exists() or data_path.stat().st_size == 0:
            return

        import json
        import csv
        import shutil
        from openpyxl import Workbook

        MAX_ROWS_PER_SHEET = 1_000_000
        excel_success = False

        wb = Workbook(write_only=True)
        try:
            meta_ws = wb.create_sheet(title="Metadata")
            meta_ws.append(["Key", "Value"])
            metadata = {}
            if meta_path.exists():
                try:
                    with meta_path.open("r", encoding="utf-8") as f:
                        metadata = json.load(f)
                except Exception:
                    pass
            for key in sorted(metadata):
                value = str(metadata[key]).replace("\n", " ").strip()
                meta_ws.append([key, value])

            sheet_index = 1
            data_ws = wb.create_sheet(title="Data")
            header_row = None
            current_sheet_rows = 0

            with data_path.open("r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    if header_row is None:
                        header_row = row
                        data_ws.append(row)
                        current_sheet_rows += 1
                        continue

                    if current_sheet_rows >= MAX_ROWS_PER_SHEET:
                        sheet_index += 1
                        data_ws = wb.create_sheet(title=f"Data_{sheet_index}")
                        data_ws.append(header_row)
                        current_sheet_rows = 1

                    typed_row = []
                    for cell in row:
                        if cell == "":
                            typed_row.append("")
                        else:
                            try:
                                if "." in cell or "e" in cell.lower():
                                    typed_row.append(float(cell))
                                else:
                                    typed_row.append(int(cell))
                            except ValueError:
                                typed_row.append(cell)
                    data_ws.append(typed_row)
                    current_sheet_rows += 1

            target_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target_path)
            excel_success = True
        except Exception as exc:
            _LOG.error("Failed to recover raw log as Excel from %s: %s", data_path, exc, exc_info=True)
        finally:
            try:
                wb.close()
            except Exception:
                pass

        if excel_success and target_path.exists() and target_path.stat().st_size > 0:
            data_path.unlink(missing_ok=True)
            meta_path.unlink(missing_ok=True)
        else:
            try:
                csv_target = target_path.with_suffix(".csv")
                shutil.copy2(data_path, csv_target)
                _LOG.info("Recovered raw temp log as CSV fallback: %s", csv_target)
                data_path.unlink(missing_ok=True)
                meta_path.unlink(missing_ok=True)
            except Exception as copy_exc:
                _LOG.error("Failed to copy raw temp log to CSV fallback %s: %s", target_path, copy_exc)

    def _write_one(self, writer, item: tuple) -> bool:
        if writer is None:
            return False
        try:
            ts, direction, raw, delta_t_ms = item
            hex_str = raw.hex(" ").upper() if self._hex_format == "hex" else raw.hex().upper()
            writer.writerow([ts, direction, hex_str, f"{delta_t_ms:.1f}"])
            return True
        except Exception as exc:
            self._record_error("write", exc)
            return False

    def _maybe_periodic_flush(self, fp) -> None:
        if fp is None:
            return
        if self._flush_interval <= 0:
            try:
                fp.flush()
                self._last_flush = time.monotonic()
            except Exception as exc:
                self._record_error("flush", exc)
            return
        now = time.monotonic()
        if now - self._last_flush >= self._flush_interval:
            try:
                fp.flush()
                self._last_flush = now
            except Exception as exc:
                self._record_error("flush", exc)

    def _record_error(self, context: str, exc: Exception) -> None:
        _LOG.error("RawLogger %s error (writer thread)", context, exc_info=True)
        with self._error_lock:
            if self._pending_error is None:
                self._pending_error = f"Raw log {context} error for {self.path.name}: {exc}"

    def _pump_pending_error(self) -> None:
        with self._error_lock:
            err = self._pending_error
            self._pending_error = None
        if err is not None:
            self._disabled = True
            if self._on_error is not None:
                self._on_error(err)

    # ------------------------------------------------------------------
    # Helpers used by both threads
    # ------------------------------------------------------------------

    def _read_existing_header(self) -> list[str]:
        with self.path.open("r", encoding="utf-8") as fp:
            for line in fp:
                stripped = line.strip()
                if not stripped or stripped.startswith("#"):
                    continue
                return [c.strip() for c in stripped.split(",")] if stripped else []
        return []

    def _write_metadata(self) -> None:
        if not self._metadata or self._fp is None:
            return
        for key in sorted(self._metadata):
            value = str(self._metadata[key]).replace("\n", " ").strip()
            self._fp.write(f"# {key}: {value}\n")

    def _handle_error(self, context: str, exc: Exception) -> None:
        """UI-thread synchronous error path (used when ``log()`` itself
        raises before enqueueing). Background-thread errors go through
        ``_record_error`` + ``_pump_pending_error`` instead.
        """
        if self._disabled:
            return
        self._disabled = True
        self.close()
        _LOG.error("RawLogger %s error", context, exc_info=True)
        if self._on_error is not None:
            self._on_error(f"Raw log {context} error for {self.path.name}: {exc}")
